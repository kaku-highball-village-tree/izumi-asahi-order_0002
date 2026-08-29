"""JSONの商品データをTSVとCSVへ変換するコマンドラインツール。"""

from __future__ import annotations

import csv
import json
import os
import sys
import tempfile
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError


USAGE = "使用方法：py src/json_to_tsv_csv_Cmd.py 入力ファイル.json"
OUTPUT_COLUMNS = [
    "productCode",
    "productName",
    "spec",
    "futeiKan",
    "taxRate",
    "temperatureZone",
    "priceUnit",
    "salePrice",
    "orderUnit",
    "packQty",
    "packagingBoxes",
    "productWeightKg",
    "stockType",
    "stockTypeLabel",
    "remarks",
    "deleted",
    "suppliers",
    "customers",
    "createdAt",
    "updatedAt",
]


class JsonConversionError(Exception):
    """利用者へ日本語で表示する変換エラーを表す。"""


def get_input_path(arguments: Sequence[str]) -> Path:
    """引数を検証し、入力JSONファイルのパスを返す。"""
    if not arguments:
        raise JsonConversionError(
            f"エラー：入力JSONファイルを指定してください。\n{USAGE}"
        )
    if len(arguments) > 1:
        raise JsonConversionError(
            f"エラー：指定できる入力ファイルは1つだけです。\n{USAGE}"
        )
    return Path(arguments[0])


def validate_input_file(input_path: Path) -> None:
    """入力パスの存在、ファイル種別および拡張子を検証する。"""
    if not input_path.exists():
        raise JsonConversionError(
            "エラー：入力ファイルが見つかりません。\n"
            f"入力ファイル：{input_path}"
        )
    if not input_path.is_file():
        raise JsonConversionError("エラー：指定されたパスはファイルではありません。")
    if input_path.suffix.lower() != ".json":
        raise JsonConversionError("エラー：JSONファイルを指定してください。")


def create_output_paths(input_path: Path) -> tuple[Path, Path, Path, Path]:
    """TSV、CSV、XLSXおよび警告ファイルのパスを作成する。"""
    return (
        input_path.with_suffix(".tsv"),
        input_path.with_suffix(".csv"),
        input_path.with_suffix(".xlsx"),
        input_path.with_name(f"{input_path.stem}_warning.txt"),
    )


def load_json_file(input_path: Path) -> Any:
    """UTF-8（BOMあり・なし）でJSONファイルを読み込む。"""
    try:
        with input_path.open("r", encoding="utf-8-sig") as input_file:
            return json.load(input_file)
    except UnicodeDecodeError as error:
        raise JsonConversionError(
            "エラー：入力ファイルをUTF-8として読み込めません。"
        ) from error
    except json.JSONDecodeError as error:
        raise JsonConversionError(
            "エラー：JSONの形式が正しくありません。\n"
            f"行：{error.lineno}\n"
            f"列：{error.colno}\n"
            f"詳細：{error.msg}"
        ) from error
    except OSError as error:
        raise JsonConversionError(
            "エラー：入力ファイルを読み込めません。\n"
            f"入力ファイル：{input_path}\n"
            f"詳細：{error}"
        ) from error


def get_items(data: Any) -> list[dict[str, Any]]:
    """トップレベルとitemsの構造を検証して商品一覧を返す。"""
    if not isinstance(data, dict):
        raise JsonConversionError(
            "エラー：JSONのトップレベルはオブジェクトである必要があります。"
        )
    if "items" not in data:
        raise JsonConversionError("エラー：JSONにitemsが存在しません。")

    items = data["items"]
    if not isinstance(items, list):
        raise JsonConversionError("エラー：itemsは配列である必要があります。")

    for index, item in enumerate(items):
        if not isinstance(item, dict):
            raise JsonConversionError(
                f"エラー：items[{index}]がオブジェクトではありません。"
            )
    return items


def format_cell_value(value: Any) -> str:
    """JSONの値をTSV・CSV用のセル文字列へ変換する。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    return str(value)


def create_output_rows(items: list[dict[str, Any]]) -> list[list[str]]:
    """指定された列順でヘッダーと全商品行を作成する。"""
    rows = [OUTPUT_COLUMNS.copy()]
    for item in items:
        rows.append(
            [
                format_cell_value(item[column]) if column in item else ""
                for column in OUTPUT_COLUMNS
            ]
        )
    return rows


def write_warning_file(
    warning_path: Path,
    existing_paths: list[Path],
    tsv_path: Path,
    csv_path: Path,
    xlsx_path: Path,
) -> None:
    """既存出力についての警告ファイルを、未作成の場合だけ作成する。"""
    lines = ["警告：出力ファイルがすでに存在するため、変換を中止しました。"]
    lines.extend(f"既存ファイル：{path}" for path in existing_paths)
    lines.extend(
        (
            f"TSV出力：{tsv_path}",
            f"CSV出力：{csv_path}",
            f"XLSX出力：{xlsx_path}",
        )
    )

    try:
        with warning_path.open("x", encoding="utf-8", newline="\n") as warning_file:
            warning_file.write("\n".join(lines) + "\n")
    except FileExistsError:
        print(f"警告：警告ファイルはすでに存在するため上書きしません：{warning_path}")
    except OSError as error:
        print(f"警告：警告ファイルを作成できません：{warning_path}")
        print(f"詳細：{error}")


def validate_output_paths(
    tsv_path: Path, csv_path: Path, xlsx_path: Path, warning_path: Path
) -> None:
    """既存の出力を検出し、警告を記録して変換を中止する。"""
    existing_paths = [
        path for path in (tsv_path, csv_path, xlsx_path) if path.exists()
    ]
    if not existing_paths:
        return

    write_warning_file(
        warning_path, existing_paths, tsv_path, csv_path, xlsx_path
    )
    details = "\n".join(f"既存ファイル：{path}" for path in existing_paths)
    raise JsonConversionError(
        "警告：出力ファイルがすでに存在するため、変換を中止しました。\n"
        f"{details}"
    )


def write_temporary_file(
    output_path: Path, rows: list[list[str]], delimiter: str
) -> Path:
    """出力先と同じフォルダーに一意な一時ファイルを書き込む。"""
    temporary_path: Path | None = None
    try:
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{output_path.name}.", suffix=".tmp", dir=output_path.parent
        )
        temporary_path = Path(temporary_name)
        with os.fdopen(
            descriptor, "w", encoding="utf-8-sig", newline=""
        ) as output_file:
            writer = csv.writer(
                output_file, delimiter=delimiter, lineterminator="\r\n"
            )
            writer.writerows(rows)
        return temporary_path
    except (OSError, csv.Error) as error:
        if temporary_path is not None:
            try:
                temporary_path.unlink()
            except OSError:
                pass
        raise JsonConversionError(
            "エラー：出力ファイルを作成できません。\n"
            f"出力先：{output_path}\n"
            f"詳細：{error}"
        ) from error


def write_temporary_xlsx(
    output_path: Path, items: list[dict[str, Any]]
) -> Path:
    """商品データをXLSXの一時ファイルへ書き込む。"""
    temporary_path: Path | None = None
    try:
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=f".{output_path.name}.", suffix=".xlsx", dir=output_path.parent
        )
        os.close(descriptor)
        temporary_path = Path(temporary_name)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "items"
        worksheet.append(OUTPUT_COLUMNS)

        for item in items:
            row_number = worksheet.max_row + 1
            for column_number, column_name in enumerate(OUTPUT_COLUMNS, start=1):
                value = item[column_name] if column_name in item else None
                cell = worksheet.cell(row=row_number, column=column_number)

                if column_name == "productCode":
                    cell.value = "" if value is None else str(value)
                    cell.number_format = "@"
                elif value is None:
                    cell.value = None
                elif isinstance(value, bool):
                    cell.value = "true" if value else "false"
                elif isinstance(value, (dict, list)):
                    cell.value = json.dumps(
                        value, ensure_ascii=False, separators=(",", ":")
                    )
                else:
                    cell.value = value

        workbook.save(temporary_path)
        workbook.close()
        return temporary_path
    except (OSError, ValueError, TypeError, IllegalCharacterError) as error:
        if temporary_path is not None:
            try:
                temporary_path.unlink()
            except OSError:
                pass
        raise JsonConversionError(
            "エラー：XLSXファイルを作成できません。\n"
            f"出力先：{output_path}\n"
            f"詳細：{error}"
        ) from error


def finalize_output_files(file_pairs: list[tuple[Path, Path]]) -> None:
    """一時ファイルを、上書きせず正式な出力ファイルとして確定する。"""
    finalized_paths: list[Path] = []
    try:
        for temporary_path, output_path in file_pairs:
            # ハードリンク作成は既存の出力先を上書きせず、同一ファイルシステム内で
            # 一時ファイルを正式名から参照できるようにする。
            os.link(temporary_path, output_path)
            finalized_paths.append(output_path)
        for temporary_path, _ in file_pairs:
            temporary_path.unlink()
    except OSError as error:
        for output_path in finalized_paths:
            try:
                output_path.unlink()
            except OSError:
                pass
        for temporary_path, _ in file_pairs:
            try:
                temporary_path.unlink()
            except OSError:
                pass
        raise JsonConversionError(
            "エラー：出力ファイルを確定できません。\n"
            f"詳細：{error}"
        ) from error


def write_output_files(
    rows: list[list[str]],
    items: list[dict[str, Any]],
    tsv_path: Path,
    csv_path: Path,
    xlsx_path: Path,
) -> None:
    """3形式を書き込み、すべて成功した場合だけ正式名へ確定する。"""
    temporary_pairs: list[tuple[Path, Path]] = []
    try:
        temporary_pairs.append(
            (write_temporary_file(tsv_path, rows, "\t"), tsv_path)
        )
        temporary_pairs.append(
            (write_temporary_file(csv_path, rows, ","), csv_path)
        )
        temporary_pairs.append(
            (write_temporary_xlsx(xlsx_path, items), xlsx_path)
        )
    except JsonConversionError:
        for temporary_path, _ in temporary_pairs:
            try:
                temporary_path.unlink()
            except OSError:
                pass
        raise

    finalize_output_files(temporary_pairs)


def main() -> None:
    """JSONの検証からTSV・CSV出力までの処理全体を実行する。"""
    try:
        input_path = get_input_path(sys.argv[1:])
        validate_input_file(input_path)
        tsv_path, csv_path, xlsx_path, warning_path = create_output_paths(input_path)

        data = load_json_file(input_path)
        items = get_items(data)
        rows = create_output_rows(items)
        validate_output_paths(tsv_path, csv_path, xlsx_path, warning_path)
        write_output_files(rows, items, tsv_path, csv_path, xlsx_path)

        print("変換が完了しました。")
        print(f"入力ファイル : {input_path}")
        print(f"商品件数     : {len(items)}")
        print(f"TSV出力      : {tsv_path}")
        print(f"CSV出力      : {csv_path}")
        print(f"XLSX出力     : {xlsx_path}")
    except JsonConversionError as error:
        print(error)
        sys.exit(1)


if __name__ == "__main__":
    main()
