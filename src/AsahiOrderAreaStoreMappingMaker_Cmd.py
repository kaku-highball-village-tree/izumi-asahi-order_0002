# -- coding: utf-8 --
###############################################################
#
# AsahiOrderAreaStoreMappingMaker_Cmd.py
#
# pip install openpyxl
#
###############################################################

import csv
import os
import re
import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


TARGET_WORKSHEET_NAMES: tuple[str, str] = ("本州マグロ(週間)", "割り")
SUPPORTED_EXTENSIONS: set[str] = {".xlsx"}
MAX_BACKUP_NUMBER: int = 9999
AREA_STORE_MAPPING_FILE_NAME: str = "AsahiOrderAreaStoreMapping_週間_step0001.tsv"
AREA_STORE_MAPPING_HEADERS: tuple[str, str, str] = (
    "配送センター名",
    "店舗コード",
    "店舗名",
)
ALLOCATION_MAPPING_FILE_NAME: str = "AsahiOrderAreaStoreMapping_割り_step0001.tsv"
ALLOCATION_MAPPING_HEADERS: tuple[str, str, str, str, str] = (
    "エリア名",
    "店舗コード",
    "店舗略称",
    "APEX店舗コード",
    "APEX店舗名",
)
ALLOCATION_STORE_CODE_MISMATCH_FILE_NAME: str = (
    "AsahiOrderAreaStoreMapping_割り_step0002_店舗コード不一致.tsv"
)
ALLOCATION_STORE_NAME_MISMATCH_FILE_NAME: str = (
    "AsahiOrderAreaStoreMapping_割り_step0002_店舗名不一致.tsv"
)
ALLOCATION_FORMAL_STORE_NAME_FILE_NAME: str = (
    "AsahiOrderAreaStoreMapping_割り_step0002_正式店舗名.txt"
)
ALLOCATION_STEP0002_FILE_NAME: str = "AsahiOrderAreaStoreMapping_割り_step0002.tsv"
ALLOCATION_STEP0002_ERROR_FILE_NAME: str = (
    "AsahiOrderAreaStoreMapping_割り_step0002_error.txt"
)
ALLOCATION_FORMAL_STORE_NAME_HEADERS: tuple[str, str] = ("店舗コード", "店舗略称")
ALLOCATION_STEP0003_FILE_NAME: str = "AsahiOrderAreaStoreMapping_割り_step0003.tsv"
ALLOCATION_STEP0003_ERROR_FILE_NAME: str = (
    "AsahiOrderAreaStoreMapping_割り_step0003_error.txt"
)
ALLOCATION_STEP0003_HEADERS: tuple[str, str, str] = (
    "エリア名",
    "店舗コード",
    "店舗略称",
)
ALLOCATION_STEP0004_FILE_NAME: str = "AsahiOrderAreaStoreMapping_割り_step0004.tsv"
ALLOCATION_STEP0004_ERROR_FILE_NAME: str = (
    "AsahiOrderAreaStoreMapping_割り_step0004_error.txt"
)
ALLOCATION_STEP0004_HEADERS: tuple[str, str, str, str] = (
    "配送センター名",
    "エリア名",
    "店舗コード",
    "店舗略称",
)
ALLOCATION_STEP0005_FILE_NAME: str = "AsahiOrderAreaStoreMapping_割り_step0005.tsv"
ALLOCATION_STEP0005_ERROR_FILE_NAME: str = (
    "AsahiOrderAreaStoreMapping_割り_step0005_error.txt"
)
ALLOCATION_STEP0005_EXCLUDED_AREAS: set[str] = {"熊本", "九州", "北九州"}
ALLOCATION_STEP0005_CENTER_NAMES: dict[str, str] = {
    "広島": "①広島センター",
    "岡山": "②広島センター(岡山・四国転送分)",
    "四国／岡山": "②広島センター(岡山・四国転送分)",
}
AREA_STORE_MAPPING_TEXT_FILE_NAME: str = "AsahiOrderAreaStoreMapping_対応表.txt"
AREA_STORE_MAPPING_TSV_FILE_NAME: str = "AsahiOrderAreaStoreMapping_対応表.tsv"
AREA_STORE_MAPPING_ERROR_FILE_NAME: str = "AsahiOrderAreaStoreMapping_対応表_error.txt"
FINAL_AREA_STORE_MAPPING_HEADERS: tuple[str, str, str] = (
    "配送センター名",
    "店舗コード",
    "店舗略称",
)
FINAL_DISTRIBUTION_CENTER_NAMES: dict[str, str] = {
    "①広島センター": "広島センター",
    "②広島センター(岡山・四国転送分)": "広島センター(岡山・四国転送分)",
}
CIRCLED_NUMBERS: str = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"


def write_error_text(pszOutputFileFullPath: str, pszErrorMessage: str) -> None:
    """エラーメッセージをUTF-8のテキストファイルへ上書き保存します。"""
    pszDirectoryFullPath: str = os.path.dirname(pszOutputFileFullPath)
    if pszDirectoryFullPath != "":
        os.makedirs(pszDirectoryFullPath, exist_ok=True)
    with open(pszOutputFileFullPath, mode="w", encoding="utf-8", newline="") as objFile:
        objFile.write(pszErrorMessage.rstrip("\n") + "\n")


def get_error_file_full_path(pszInputFileFullPath: str) -> str:
    """入力ファイルと同じフォルダーに作る_error.txtのパスを返します。"""
    pszDirectoryFullPath: str = os.path.dirname(os.path.abspath(pszInputFileFullPath))
    pszBaseNameWithoutExtension: str = os.path.splitext(
        os.path.basename(pszInputFileFullPath)
    )[0]
    return os.path.join(pszDirectoryFullPath, pszBaseNameWithoutExtension + "_error.txt")


def report_processing_error(
    pszInputFileFullPath: str, pszProcessName: str, pszDetailMessage: str
) -> None:
    """標準エラーと入力ファイル用_error.txtへ同じエラーを出力します。"""
    pszErrorMessage: str = (
        "処理結果: エラー\n"
        + "入力ファイル: "
        + os.path.abspath(pszInputFileFullPath)
        + "\n発生した処理: "
        + pszProcessName
        + "\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    try:
        write_error_text(get_error_file_full_path(pszInputFileFullPath), pszErrorMessage)
    except OSError as objException:
        print(
            "Error: _error.txtの保存にも失敗しました。Detail = " + str(objException),
            file=sys.stderr,
        )
    try:
        pszWarningFileFullPath: str = get_warning_file_full_path(
            pszInputFileFullPath
        )
        if os.path.exists(pszWarningFileFullPath):
            os.remove(pszWarningFileFullPath)
    except OSError as objException:
        print(
            "Error: 古い_warning.txtを削除できませんでした。Detail = "
            + str(objException),
            file=sys.stderr,
        )


def remove_old_error_file(pszInputFileFullPath: str) -> None:
    """正常終了後、以前の処理で作られた_error.txtがあれば削除します。"""
    pszErrorFileFullPath: str = get_error_file_full_path(pszInputFileFullPath)
    if os.path.exists(pszErrorFileFullPath):
        os.remove(pszErrorFileFullPath)


def get_warning_file_full_path(pszInputFileFullPath: str) -> str:
    """入力ファイルと同じフォルダーに作る_warning.txtのパスを返します。"""
    pszDirectoryFullPath: str = os.path.dirname(os.path.abspath(pszInputFileFullPath))
    pszBaseNameWithoutExtension: str = os.path.splitext(
        os.path.basename(pszInputFileFullPath)
    )[0]
    return os.path.join(
        pszDirectoryFullPath, pszBaseNameWithoutExtension + "_warning.txt"
    )


def validate_input_path(pszInputFileFullPath: str) -> str:
    """入力パスと拡張子を検証し、絶対パスを返します。"""
    pszAbsolutePath: str = os.path.abspath(pszInputFileFullPath)
    if not os.path.exists(pszAbsolutePath):
        raise ValueError("入力ファイルが見つかりません。Path = " + pszAbsolutePath)
    if not os.path.isfile(pszAbsolutePath):
        raise ValueError("入力パスがファイルではありません。Path = " + pszAbsolutePath)
    pszExtension: str = os.path.splitext(pszAbsolutePath)[1].lower()
    if pszExtension not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            "入力ファイルの拡張子は.xlsxではありません。Path = " + pszAbsolutePath
        )
    return pszAbsolutePath


def normalize_cell_value(objValue: object) -> object:
    """空セルを空文字へ変換し、それ以外のセル値は変更せず返します。"""
    if objValue is None:
        return ""
    return objValue


def is_empty_cell_value(objValue: object) -> bool:
    """末尾空行・空列の判定対象となる空セルかを返します。"""
    return objValue is None or objValue == ""


def read_worksheet_rows(objWorksheet: Worksheet) -> tuple[list[list[object]], int]:
    """シートのセル値を読み、末尾の完全空行・空列を除いて返します。"""
    iLastDataRow: int = 0
    iLastDataColumn: int = 0
    for objRow in objWorksheet.iter_rows(
        min_row=1,
        max_row=objWorksheet.max_row,
        min_col=1,
        max_col=objWorksheet.max_column,
    ):
        for objCell in objRow:
            if not is_empty_cell_value(objCell.value):
                iLastDataRow = max(iLastDataRow, objCell.row)
                iLastDataColumn = max(iLastDataColumn, objCell.column)
    if iLastDataRow == 0 or iLastDataColumn == 0:
        return [], 0
    listRows: list[list[object]] = []
    for objRow in objWorksheet.iter_rows(
        min_row=1,
        max_row=iLastDataRow,
        min_col=1,
        max_col=iLastDataColumn,
    ):
        listRows.append([normalize_cell_value(objCell.value) for objCell in objRow])
    return listRows, iLastDataColumn


def read_excel_worksheets(
    pszInputFileFullPath: str,
) -> tuple[dict[str, tuple[list[list[object]], int]], list[str]]:
    """対象シートを1回のExcel読込で取得し、存在しないシート名も返します。"""
    objWorkbook: Workbook = load_workbook(pszInputFileFullPath, data_only=True)
    try:
        dictWorksheetResults: dict[str, tuple[list[list[object]], int]] = {}
        listMissingWorksheetNames: list[str] = []
        for pszWorksheetName in TARGET_WORKSHEET_NAMES:
            if pszWorksheetName not in objWorkbook.sheetnames:
                listMissingWorksheetNames.append(pszWorksheetName)
                continue
            dictWorksheetResults[pszWorksheetName] = read_worksheet_rows(
                objWorkbook[pszWorksheetName]
            )
        return dictWorksheetResults, listMissingWorksheetNames
    finally:
        objWorkbook.close()


def get_output_file_path(
    pszInputFileFullPath: str, pszWorksheetName: str
) -> Path:
    """入力ファイル名とシート名からTSVの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    return objInputPath.with_name(objInputPath.stem + "_" + pszWorksheetName + ".tsv")


def create_temporary_path(objOutputPath: Path) -> Path:
    """出力先と同じフォルダーに同じ拡張子の一時パスを作ります。"""
    iFileDescriptor, pszTemporaryPath = tempfile.mkstemp(
        prefix=objOutputPath.stem + "_",
        suffix=objOutputPath.suffix,
        dir=objOutputPath.parent,
    )
    os.close(iFileDescriptor)
    return Path(pszTemporaryPath)


def save_tsv_rows(objOutputPath: Path, listRows: list[list[object]]) -> None:
    """セル値をUTF-8 BOMなし、タブ区切り、CRLFのTSVへ保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerows(listRows)


def read_tsv_rows(objInputPath: Path) -> list[list[str]]:
    """UTF-8のTSVを先頭から終端まで読み取ります。"""
    with objInputPath.open(mode="r", encoding="utf-8", newline="") as objFile:
        return list(csv.reader(objFile, delimiter="\t", strict=True))


def is_allocation_store_data_row(listValues: list[str]) -> bool:
    """割り対応表の2～6列目がすべて空白以外の店舗行か返します。"""
    if len(listValues) < 6:
        return False
    return all(pszValue.strip() != "" for pszValue in listValues[1:6])


def build_allocation_mapping_rows(listRows: list[list[str]]) -> list[list[str]]:
    """単位行より後ろから割り対応表の2～6列目を抽出します。"""
    iUnitRowIndex: int | None = None
    for iRowIndex, listValues in enumerate(listRows):
        if len(listValues) >= 6 and listValues[5] == "単位":
            iUnitRowIndex = iRowIndex
            break
    if iUnitRowIndex is None:
        raise ValueError("割りTSV内に6列目が「単位」の行が見つかりません。")
    iDataStartRowIndex: int | None = None
    for iRowIndex in range(iUnitRowIndex + 1, len(listRows)):
        if is_allocation_store_data_row(listRows[iRowIndex]):
            iDataStartRowIndex = iRowIndex
            break
    if iDataStartRowIndex is None:
        raise ValueError(
            "6列目が「単位」の行より後ろに店舗データが見つかりません。"
        )
    return [
        listValues[1:6]
        for listValues in listRows[iDataStartRowIndex:]
        if is_allocation_store_data_row(listValues)
    ]


def get_allocation_mapping_output_path(objInputTsvPath: Path) -> Path:
    """割りTSVと同じフォルダーに作る割り対応表TSVのパスを返します。"""
    return objInputTsvPath.with_name(ALLOCATION_MAPPING_FILE_NAME)


def save_allocation_mapping_tsv(
    objOutputPath: Path, listMappingRows: list[list[str]]
) -> None:
    """5列ヘッダーと割り対応データをUTF-8 TSVへ保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(ALLOCATION_MAPPING_HEADERS)
        objWriter.writerows(listMappingRows)


def process_allocation_mapping_file(objInputTsvPath: Path) -> tuple[Path, int]:
    """割りTSVから割り対応表を独立して作成します。"""
    if not objInputTsvPath.exists() or not objInputTsvPath.is_file():
        raise ValueError("割りTSVが見つかりません。Path = " + str(objInputTsvPath))
    listRows: list[list[str]] = read_tsv_rows(objInputTsvPath)
    try:
        listMappingRows: list[list[str]] = build_allocation_mapping_rows(listRows)
    except ValueError as objException:
        raise ValueError(
            str(objException) + " Path = " + str(objInputTsvPath)
        ) from objException
    objOutputPath: Path = get_allocation_mapping_output_path(objInputTsvPath)
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_allocation_mapping_tsv(objTemporaryPath, listMappingRows)
        os.replace(objTemporaryPath, objOutputPath)
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return objOutputPath, len(listMappingRows)


def get_allocation_mismatch_output_paths(
    objStep0001Path: Path,
) -> tuple[Path, Path]:
    """割りstep0001と同じフォルダーに作る2種類の不一致TSVパスを返します。"""
    return (
        objStep0001Path.with_name(ALLOCATION_STORE_CODE_MISMATCH_FILE_NAME),
        objStep0001Path.with_name(ALLOCATION_STORE_NAME_MISMATCH_FILE_NAME),
    )


def build_allocation_mismatch_rows(
    listRows: list[list[str]],
) -> tuple[list[list[str]], list[list[str]]]:
    """割りstep0001を検証し、店舗コード不一致行と店舗名不一致行を返します。"""
    if not listRows:
        raise ValueError("割りstep0001 TSVが空です。")
    if tuple(listRows[0]) != ALLOCATION_MAPPING_HEADERS:
        raise ValueError(
            "割りstep0001 TSVの項目名行が正しくありません。期待値 = "
            + "\\t".join(ALLOCATION_MAPPING_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listRows[0])
        )
    listStoreCodeMismatchRows: list[list[str]] = []
    listStoreNameMismatchRows: list[list[str]] = []
    for iRowNumber, listValues in enumerate(listRows[1:], start=2):
        if not listValues:
            continue
        if len(listValues) != len(ALLOCATION_MAPPING_HEADERS):
            raise ValueError(
                "割りstep0001 TSVのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = "
                + str(len(ALLOCATION_MAPPING_HEADERS))
                + "、実際の列数 = "
                + str(len(listValues))
            )
        if listValues[1] != listValues[3]:
            listStoreCodeMismatchRows.append(listValues)
        elif listValues[2] != listValues[4]:
            listStoreNameMismatchRows.append(listValues)
    return listStoreCodeMismatchRows, listStoreNameMismatchRows


def replace_allocation_mismatch_files(
    dictTemporaryPaths: dict[Path, Path], dictBackupPaths: dict[Path, Path]
) -> None:
    """2種類の不一致TSVを一括置換し、失敗時は処理前の状態へ戻します。"""
    listRenamedOutputs: list[tuple[Path, Path]] = []
    listReplacedOutputs: list[Path] = []
    try:
        for objOutputPath, objBackupPath in dictBackupPaths.items():
            if objBackupPath.exists():
                raise FileExistsError(
                    "バックアップ先がすでに存在します。Path = " + str(objBackupPath)
                )
            os.rename(objOutputPath, objBackupPath)
            listRenamedOutputs.append((objOutputPath, objBackupPath))
        for objOutputPath, objTemporaryPath in dictTemporaryPaths.items():
            os.replace(objTemporaryPath, objOutputPath)
            listReplacedOutputs.append(objOutputPath)
    except Exception:
        for objOutputPath in reversed(listReplacedOutputs):
            if objOutputPath.exists():
                objOutputPath.unlink()
        for objOutputPath, objBackupPath in reversed(listRenamedOutputs):
            if objBackupPath.exists():
                os.rename(objBackupPath, objOutputPath)
        raise


def process_allocation_mismatch_files(
    objStep0001Path: Path,
) -> tuple[Path, int, Path, int, dict[Path, Path]]:
    """割りstep0001から2種類の不一致TSVを作成します。"""
    if not objStep0001Path.exists() or not objStep0001Path.is_file():
        raise ValueError(
            "割りstep0001 TSVが見つかりません。Path = " + str(objStep0001Path)
        )
    listRows: list[list[str]] = read_tsv_rows(objStep0001Path)
    listStoreCodeMismatchRows, listStoreNameMismatchRows = (
        build_allocation_mismatch_rows(listRows)
    )
    objStoreCodeMismatchPath, objStoreNameMismatchPath = (
        get_allocation_mismatch_output_paths(objStep0001Path)
    )
    dictOutputRows: dict[Path, list[list[str]]] = {
        objStoreCodeMismatchPath: listStoreCodeMismatchRows,
        objStoreNameMismatchPath: listStoreNameMismatchRows,
    }
    dictBackupPaths: dict[Path, Path] = {
        objOutputPath: get_next_backup_path(objOutputPath)
        for objOutputPath in dictOutputRows
        if objOutputPath.exists()
    }
    dictTemporaryPaths: dict[Path, Path] = {}
    try:
        for objOutputPath, listMismatchRows in dictOutputRows.items():
            objTemporaryPath: Path = create_temporary_path(objOutputPath)
            dictTemporaryPaths[objOutputPath] = objTemporaryPath
            save_tsv_rows(
                objTemporaryPath,
                [list(ALLOCATION_MAPPING_HEADERS)] + listMismatchRows,
            )
        replace_allocation_mismatch_files(dictTemporaryPaths, dictBackupPaths)
    finally:
        for objTemporaryPath in dictTemporaryPaths.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    return (
        objStoreCodeMismatchPath,
        len(listStoreCodeMismatchRows),
        objStoreNameMismatchPath,
        len(listStoreNameMismatchRows),
        dictBackupPaths,
    )


def get_allocation_formal_store_name_paths(
    objStep0001Path: Path,
) -> tuple[Path, Path, Path]:
    """正式店舗名入力、最終step0002出力、専用エラーの各パスを返します。"""
    return (
        objStep0001Path.with_name(ALLOCATION_FORMAL_STORE_NAME_FILE_NAME),
        objStep0001Path.with_name(ALLOCATION_STEP0002_FILE_NAME),
        objStep0001Path.with_name(ALLOCATION_STEP0002_ERROR_FILE_NAME),
    )


def build_allocation_step0002_rows(
    listStep0001Rows: list[list[str]], listFormalStoreNameRows: list[list[str]]
) -> tuple[list[list[str]], int, int]:
    """step0001全店舗を維持し、指定店舗の店舗略称だけを正式名へ変更します。"""
    if not listStep0001Rows:
        raise ValueError("割りstep0001 TSVが空です。")
    if tuple(listStep0001Rows[0]) != ALLOCATION_MAPPING_HEADERS:
        raise ValueError(
            "割りstep0001 TSVの項目名行が正しくありません。期待値 = "
            + "\\t".join(ALLOCATION_MAPPING_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listStep0001Rows[0])
        )
    if not listFormalStoreNameRows:
        raise ValueError("正式店舗名ファイルが空です。")
    if tuple(listFormalStoreNameRows[0]) != ALLOCATION_FORMAL_STORE_NAME_HEADERS:
        raise ValueError(
            "正式店舗名ファイルの項目名行が正しくありません。期待値 = "
            + "\\t".join(ALLOCATION_FORMAL_STORE_NAME_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listFormalStoreNameRows[0])
        )

    dictFormalStoreNames: dict[str, tuple[str, int]] = {}
    for iRowNumber, listValues in enumerate(listFormalStoreNameRows[1:], start=2):
        if not listValues:
            continue
        if len(listValues) != len(ALLOCATION_FORMAL_STORE_NAME_HEADERS):
            raise ValueError(
                "正式店舗名ファイルのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = 2、実際の列数 = "
                + str(len(listValues))
            )
        pszStoreCode, pszFormalStoreName = listValues
        if pszStoreCode == "" or pszFormalStoreName == "":
            raise ValueError(
                "正式店舗名ファイルに空の店舗コードまたは店舗略称があります。行 = "
                + str(iRowNumber)
            )
        if pszStoreCode in dictFormalStoreNames:
            raise ValueError(
                "正式店舗名ファイルに店舗コードの重複があります。店舗コード = "
                + pszStoreCode
                + "、行 = "
                + str(dictFormalStoreNames[pszStoreCode][1])
                + ", "
                + str(iRowNumber)
            )
        dictFormalStoreNames[pszStoreCode] = (pszFormalStoreName, iRowNumber)

    dictStep0001StoreCounts: dict[str, int] = {}
    listOutputRows: list[list[str]] = [list(ALLOCATION_MAPPING_HEADERS)]
    iChangedRowCount: int = 0
    for iRowNumber, listValues in enumerate(listStep0001Rows[1:], start=2):
        if not listValues:
            continue
        if len(listValues) != len(ALLOCATION_MAPPING_HEADERS):
            raise ValueError(
                "割りstep0001 TSVのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = 5、実際の列数 = "
                + str(len(listValues))
            )
        pszStoreCode: str = listValues[1]
        if pszStoreCode in dictFormalStoreNames:
            dictStep0001StoreCounts[pszStoreCode] = (
                dictStep0001StoreCounts.get(pszStoreCode, 0) + 1
            )
            pszFormalStoreName: str = dictFormalStoreNames[pszStoreCode][0]
            listOutputRow: list[str] = listValues.copy()
            if listOutputRow[2] != pszFormalStoreName:
                listOutputRow[2] = pszFormalStoreName
                iChangedRowCount += 1
            listOutputRows.append(listOutputRow)
        else:
            listOutputRows.append(listValues.copy())

    for pszStoreCode, (_, iFormalRowNumber) in dictFormalStoreNames.items():
        iMatchCount: int = dictStep0001StoreCounts.get(pszStoreCode, 0)
        if iMatchCount == 0:
            raise ValueError(
                "正式店舗名の店舗コードが割りstep0001 TSVに見つかりません。店舗コード = "
                + pszStoreCode
                + "、正式店舗名ファイル行 = "
                + str(iFormalRowNumber)
            )
        if iMatchCount > 1:
            raise ValueError(
                "正式店舗名の店舗コードが割りstep0001 TSVに複数あります。店舗コード = "
                + pszStoreCode
                + "、件数 = "
                + str(iMatchCount)
            )
    return listOutputRows, len(dictFormalStoreNames), iChangedRowCount


def replace_allocation_step0002_file(
    objOutputPath: Path, objTemporaryPath: Path, objBackupPath: Path | None
) -> None:
    """最終step0002を置換し、失敗時は既存出力を復元します。"""
    bBackupCreated: bool = False
    try:
        if objBackupPath is not None:
            if objBackupPath.exists():
                raise FileExistsError(
                    "バックアップ先がすでに存在します。Path = " + str(objBackupPath)
                )
            os.rename(objOutputPath, objBackupPath)
            bBackupCreated = True
        os.replace(objTemporaryPath, objOutputPath)
    except Exception:
        if bBackupCreated and objBackupPath is not None and objBackupPath.exists():
            if objOutputPath.exists():
                objOutputPath.unlink()
            os.rename(objBackupPath, objOutputPath)
        raise


def process_allocation_formal_store_names(
    objStep0001Path: Path,
) -> tuple[Path, Path, int, int, int, Path | None]:
    """正式店舗名をstep0001全店舗へ反映し、最終step0002を作成します。"""
    objFormalStoreNamePath, objOutputPath, _ = get_allocation_formal_store_name_paths(
        objStep0001Path
    )
    if not objStep0001Path.exists() or not objStep0001Path.is_file():
        raise ValueError(
            "割りstep0001 TSVが見つかりません。Path = " + str(objStep0001Path)
        )
    if not objFormalStoreNamePath.exists() or not objFormalStoreNamePath.is_file():
        raise ValueError(
            "正式店舗名ファイルが見つかりません。Path = "
            + str(objFormalStoreNamePath)
        )
    listStep0001Rows: list[list[str]] = read_tsv_rows(objStep0001Path)
    listFormalStoreNameRows: list[list[str]] = read_tsv_rows(objFormalStoreNamePath)
    listOutputRows, iFormalStoreNameCount, iChangedRowCount = (
        build_allocation_step0002_rows(listStep0001Rows, listFormalStoreNameRows)
    )
    objBackupPath: Path | None = (
        get_next_backup_path(objOutputPath) if objOutputPath.exists() else None
    )
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_tsv_rows(objTemporaryPath, listOutputRows)
        if read_tsv_rows(objTemporaryPath) != listOutputRows:
            raise ValueError("割りstep0002 TSVの保存後検証に失敗しました。")
        replace_allocation_step0002_file(
            objOutputPath, objTemporaryPath, objBackupPath
        )
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return (
        objFormalStoreNamePath,
        objOutputPath,
        len(listOutputRows) - 1,
        iFormalStoreNameCount,
        iChangedRowCount,
        objBackupPath,
    )


def report_allocation_step0002_error(
    objStep0001Path: Path, objFormalStoreNamePath: Path, pszDetailMessage: str
) -> Path:
    """処理Aのエラーを標準エラーとstep0002専用エラーファイルへ出力します。"""
    _, _, objErrorPath = get_allocation_formal_store_name_paths(objStep0001Path)
    pszErrorMessage: str = (
        "処理結果: エラー\n"
        + "入力ファイル: "
        + str(objFormalStoreNamePath)
        + "\n発生した処理: 割り正式店舗名反映処理\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    write_error_text(str(objErrorPath), pszErrorMessage)
    return objErrorPath


def get_allocation_step0003_paths(objStep0002Path: Path) -> tuple[Path, Path]:
    """最終step0003出力と専用エラーの各パスを返します。"""
    return (
        objStep0002Path.with_name(ALLOCATION_STEP0003_FILE_NAME),
        objStep0002Path.with_name(ALLOCATION_STEP0003_ERROR_FILE_NAME),
    )


def build_allocation_step0003_rows(
    listStep0002Rows: list[list[str]],
) -> tuple[list[list[str]], int]:
    """step0002を検証し、APEXの2列を除いた3列の全店舗行を返します。"""
    if not listStep0002Rows:
        raise ValueError("割りstep0002 TSVが空です。")
    if tuple(listStep0002Rows[0]) != ALLOCATION_MAPPING_HEADERS:
        raise ValueError(
            "割りstep0002 TSVの項目名行が正しくありません。期待値 = "
            + "\\t".join(ALLOCATION_MAPPING_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listStep0002Rows[0])
        )
    listOutputRows: list[list[str]] = [list(ALLOCATION_STEP0003_HEADERS)]
    iInputDataRowCount: int = 0
    for iRowNumber, listValues in enumerate(listStep0002Rows[1:], start=2):
        if not listValues:
            continue
        if len(listValues) != len(ALLOCATION_MAPPING_HEADERS):
            raise ValueError(
                "割りstep0002 TSVのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = 5、実際の列数 = "
                + str(len(listValues))
            )
        listOutputRows.append(listValues[:3])
        iInputDataRowCount += 1
    if len(listOutputRows) - 1 != iInputDataRowCount:
        raise ValueError(
            "割りstep0002とstep0003のデータ行数が一致しません。入力行数 = "
            + str(iInputDataRowCount)
            + "、出力行数 = "
            + str(len(listOutputRows) - 1)
        )
    return listOutputRows, iInputDataRowCount


def process_allocation_step0003_file(
    objStep0002Path: Path,
) -> tuple[Path, int, int, Path | None]:
    """step0002を再読込し、APEXの2列を除いたstep0003を作成します。"""
    if not objStep0002Path.exists() or not objStep0002Path.is_file():
        raise ValueError(
            "割りstep0002 TSVが見つかりません。Path = " + str(objStep0002Path)
        )
    listStep0002Rows: list[list[str]] = read_tsv_rows(objStep0002Path)
    listOutputRows, iInputDataRowCount = build_allocation_step0003_rows(
        listStep0002Rows
    )
    objOutputPath, _ = get_allocation_step0003_paths(objStep0002Path)
    objBackupPath: Path | None = (
        get_next_backup_path(objOutputPath) if objOutputPath.exists() else None
    )
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_tsv_rows(objTemporaryPath, listOutputRows)
        if read_tsv_rows(objTemporaryPath) != listOutputRows:
            raise ValueError("割りstep0003 TSVの保存後検証に失敗しました。")
        replace_allocation_step0002_file(
            objOutputPath, objTemporaryPath, objBackupPath
        )
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return (
        objOutputPath,
        iInputDataRowCount,
        len(listOutputRows) - 1,
        objBackupPath,
    )


def report_allocation_step0003_error(
    objStep0002Path: Path, pszDetailMessage: str
) -> Path:
    """step0003作成エラーを標準エラーと専用エラーファイルへ出力します。"""
    _, objErrorPath = get_allocation_step0003_paths(objStep0002Path)
    pszErrorMessage: str = (
        "処理結果: エラー\n"
        + "入力ファイル: "
        + str(objStep0002Path)
        + "\n発生した処理: 割りstep0003作成処理\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    write_error_text(str(objErrorPath), pszErrorMessage)
    return objErrorPath


def get_allocation_step0004_paths(objStep0003Path: Path) -> tuple[Path, Path]:
    """割りstep0004出力と専用エラーの各パスを返します。"""
    return (
        objStep0003Path.with_name(ALLOCATION_STEP0004_FILE_NAME),
        objStep0003Path.with_name(ALLOCATION_STEP0004_ERROR_FILE_NAME),
    )


def build_allocation_step0004_rows(
    listWeeklyRows: list[list[str]], listAllocationStep0003Rows: list[list[str]]
) -> tuple[list[list[str]], int, int]:
    """週間店舗コードから配送センター名を取得し、割り全店舗の左端へ追加します。"""
    if not listWeeklyRows:
        raise ValueError("週間step0001 TSVが空です。")
    if tuple(listWeeklyRows[0]) != AREA_STORE_MAPPING_HEADERS:
        raise ValueError(
            "週間step0001 TSVの項目名行が正しくありません。期待値 = "
            + "\\t".join(AREA_STORE_MAPPING_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listWeeklyRows[0])
        )
    dictDistributionCenters: dict[str, str] = {}
    for iRowNumber, listValues in enumerate(listWeeklyRows[1:], start=2):
        if not listValues:
            continue
        if len(listValues) != len(AREA_STORE_MAPPING_HEADERS):
            raise ValueError(
                "週間step0001 TSVのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = 3、実際の列数 = "
                + str(len(listValues))
            )
        pszDistributionCenterName, pszStoreCode, _ = listValues
        if pszDistributionCenterName == "" or pszStoreCode == "":
            raise ValueError(
                "週間step0001 TSVに空の配送センター名または店舗コードがあります。行 = "
                + str(iRowNumber)
            )
        pszExistingCenterName: str | None = dictDistributionCenters.get(pszStoreCode)
        if (
            pszExistingCenterName is not None
            and pszExistingCenterName != pszDistributionCenterName
        ):
            raise ValueError(
                "週間step0001 TSVの同じ店舗コードに異なる配送センター名があります。店舗コード = "
                + pszStoreCode
                + "、配送センター名 = "
                + pszExistingCenterName
                + ", "
                + pszDistributionCenterName
            )
        dictDistributionCenters[pszStoreCode] = pszDistributionCenterName

    if not listAllocationStep0003Rows:
        raise ValueError("割りstep0003 TSVが空です。")
    if tuple(listAllocationStep0003Rows[0]) != ALLOCATION_STEP0003_HEADERS:
        raise ValueError(
            "割りstep0003 TSVの項目名行が正しくありません。期待値 = "
            + "\\t".join(ALLOCATION_STEP0003_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listAllocationStep0003Rows[0])
        )
    listOutputRows: list[list[str]] = [list(ALLOCATION_STEP0004_HEADERS)]
    iMatchedRowCount: int = 0
    iUnmatchedRowCount: int = 0
    for iRowNumber, listValues in enumerate(
        listAllocationStep0003Rows[1:], start=2
    ):
        if not listValues:
            continue
        if len(listValues) != len(ALLOCATION_STEP0003_HEADERS):
            raise ValueError(
                "割りstep0003 TSVのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = 3、実際の列数 = "
                + str(len(listValues))
            )
        pszDistributionCenterName = dictDistributionCenters.get(listValues[1], "")
        if pszDistributionCenterName == "":
            iUnmatchedRowCount += 1
        else:
            iMatchedRowCount += 1
        listOutputRows.append([pszDistributionCenterName] + listValues.copy())
    if len(listOutputRows) - 1 != iMatchedRowCount + iUnmatchedRowCount:
        raise ValueError(
            "割りstep0003とstep0004のデータ行数が一致しません。入力行数 = "
            + str(iMatchedRowCount + iUnmatchedRowCount)
            + "、出力行数 = "
            + str(len(listOutputRows) - 1)
        )
    return listOutputRows, iMatchedRowCount, iUnmatchedRowCount


def process_allocation_step0004_file(
    objWeeklyStep0001Path: Path, objAllocationStep0003Path: Path
) -> tuple[Path, int, int, int, Path | None]:
    """週間step0001と割りstep0003を再読込し、割りstep0004を作成します。"""
    if not objWeeklyStep0001Path.exists() or not objWeeklyStep0001Path.is_file():
        raise ValueError(
            "週間step0001 TSVが見つかりません。Path = "
            + str(objWeeklyStep0001Path)
        )
    if not objAllocationStep0003Path.exists() or not objAllocationStep0003Path.is_file():
        raise ValueError(
            "割りstep0003 TSVが見つかりません。Path = "
            + str(objAllocationStep0003Path)
        )
    listWeeklyRows: list[list[str]] = read_tsv_rows(objWeeklyStep0001Path)
    listAllocationStep0003Rows: list[list[str]] = read_tsv_rows(
        objAllocationStep0003Path
    )
    listOutputRows, iMatchedRowCount, iUnmatchedRowCount = (
        build_allocation_step0004_rows(
            listWeeklyRows, listAllocationStep0003Rows
        )
    )
    objOutputPath, _ = get_allocation_step0004_paths(objAllocationStep0003Path)
    objBackupPath: Path | None = (
        get_next_backup_path(objOutputPath) if objOutputPath.exists() else None
    )
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_tsv_rows(objTemporaryPath, listOutputRows)
        if read_tsv_rows(objTemporaryPath) != listOutputRows:
            raise ValueError("割りstep0004 TSVの保存後検証に失敗しました。")
        replace_allocation_step0002_file(
            objOutputPath, objTemporaryPath, objBackupPath
        )
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return (
        objOutputPath,
        len(listOutputRows) - 1,
        iMatchedRowCount,
        iUnmatchedRowCount,
        objBackupPath,
    )


def report_allocation_step0004_error(
    objWeeklyStep0001Path: Path,
    objAllocationStep0003Path: Path,
    pszDetailMessage: str,
) -> Path:
    """step0004作成エラーを標準エラーと専用エラーファイルへ出力します。"""
    _, objErrorPath = get_allocation_step0004_paths(objAllocationStep0003Path)
    pszErrorMessage: str = (
        "処理結果: エラー\n週間入力ファイル: "
        + str(objWeeklyStep0001Path)
        + "\n割り入力ファイル: "
        + str(objAllocationStep0003Path)
        + "\n発生した処理: 割りstep0004作成処理\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    write_error_text(str(objErrorPath), pszErrorMessage)
    return objErrorPath


def get_allocation_step0005_paths(objStep0004Path: Path) -> tuple[Path, Path]:
    """割りstep0005出力と専用エラーの各パスを返します。"""
    return (
        objStep0004Path.with_name(ALLOCATION_STEP0005_FILE_NAME),
        objStep0004Path.with_name(ALLOCATION_STEP0005_ERROR_FILE_NAME),
    )


def build_allocation_step0005_rows(
    listStep0004Rows: list[list[str]],
) -> tuple[list[list[str]], int, int, int]:
    """指定エリアを除外し、空の配送センター名をエリア別に補完します。"""
    if not listStep0004Rows:
        raise ValueError("割りstep0004 TSVが空です。")
    if tuple(listStep0004Rows[0]) != ALLOCATION_STEP0004_HEADERS:
        raise ValueError(
            "割りstep0004 TSVの項目名行が正しくありません。期待値 = "
            + "\\t".join(ALLOCATION_STEP0004_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listStep0004Rows[0])
        )
    listOutputRows: list[list[str]] = [list(ALLOCATION_STEP0004_HEADERS)]
    iInputRowCount: int = 0
    iDeletedRowCount: int = 0
    iFilledRowCount: int = 0
    iUnchangedRowCount: int = 0
    for iRowNumber, listValues in enumerate(listStep0004Rows[1:], start=2):
        if not listValues:
            continue
        if len(listValues) != len(ALLOCATION_STEP0004_HEADERS):
            raise ValueError(
                "割りstep0004 TSVのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = 4、実際の列数 = "
                + str(len(listValues))
            )
        iInputRowCount += 1
        pszAreaName: str = listValues[1]
        if pszAreaName in ALLOCATION_STEP0005_EXCLUDED_AREAS:
            iDeletedRowCount += 1
            continue
        listOutputRow: list[str] = listValues.copy()
        pszCenterName: str | None = ALLOCATION_STEP0005_CENTER_NAMES.get(pszAreaName)
        if pszCenterName is not None and listOutputRow[0].strip() == "":
            listOutputRow[0] = pszCenterName
            iFilledRowCount += 1
        else:
            iUnchangedRowCount += 1
        listOutputRows.append(listOutputRow)
    if iInputRowCount != iDeletedRowCount + iFilledRowCount + iUnchangedRowCount:
        raise ValueError(
            "割りstep0005の分類件数が入力行数と一致しません。入力行数 = "
            + str(iInputRowCount)
        )
    if len(listOutputRows) - 1 != iInputRowCount - iDeletedRowCount:
        raise ValueError(
            "割りstep0005の出力行数が正しくありません。期待行数 = "
            + str(iInputRowCount - iDeletedRowCount)
            + "、実際の行数 = "
            + str(len(listOutputRows) - 1)
        )
    return listOutputRows, iDeletedRowCount, iFilledRowCount, iUnchangedRowCount


def process_allocation_step0005_file(
    objStep0004Path: Path,
) -> tuple[Path, int, int, int, int, int, Path | None]:
    """割りstep0004を再読込し、エリアの除外・センター補完済みTSVを作ります。"""
    if not objStep0004Path.exists() or not objStep0004Path.is_file():
        raise ValueError(
            "割りstep0004 TSVが見つかりません。Path = " + str(objStep0004Path)
        )
    listStep0004Rows: list[list[str]] = read_tsv_rows(objStep0004Path)
    listOutputRows, iDeletedRowCount, iFilledRowCount, iUnchangedRowCount = (
        build_allocation_step0005_rows(listStep0004Rows)
    )
    iInputRowCount: int = iDeletedRowCount + iFilledRowCount + iUnchangedRowCount
    objOutputPath, _ = get_allocation_step0005_paths(objStep0004Path)
    objBackupPath: Path | None = (
        get_next_backup_path(objOutputPath) if objOutputPath.exists() else None
    )
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_tsv_rows(objTemporaryPath, listOutputRows)
        if read_tsv_rows(objTemporaryPath) != listOutputRows:
            raise ValueError("割りstep0005 TSVの保存後検証に失敗しました。")
        replace_allocation_step0002_file(
            objOutputPath, objTemporaryPath, objBackupPath
        )
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return (
        objOutputPath,
        iInputRowCount,
        iDeletedRowCount,
        iFilledRowCount,
        iUnchangedRowCount,
        len(listOutputRows) - 1,
        objBackupPath,
    )


def report_allocation_step0005_error(
    objStep0004Path: Path, pszDetailMessage: str
) -> Path:
    """step0005作成エラーを標準エラーと専用エラーファイルへ出力します。"""
    _, objErrorPath = get_allocation_step0005_paths(objStep0004Path)
    pszErrorMessage: str = (
        "処理結果: エラー\n入力ファイル: "
        + str(objStep0004Path)
        + "\n発生した処理: 割りstep0005作成処理\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    write_error_text(str(objErrorPath), pszErrorMessage)
    return objErrorPath


def get_final_area_store_mapping_paths(
    objStep0005Path: Path,
) -> tuple[Path, Path, Path]:
    """最終対応表TXT・TSVと専用エラーの各パスを返します。"""
    return (
        objStep0005Path.with_name(AREA_STORE_MAPPING_TEXT_FILE_NAME),
        objStep0005Path.with_name(AREA_STORE_MAPPING_TSV_FILE_NAME),
        objStep0005Path.with_name(AREA_STORE_MAPPING_ERROR_FILE_NAME),
    )


def build_final_area_store_mapping_rows(
    listStep0005Rows: list[list[str]],
) -> tuple[list[list[str]], int, int]:
    """step0005からエリア名を除き、指定配送センター名を修正します。"""
    if not listStep0005Rows:
        raise ValueError("割りstep0005 TSVが空です。")
    if tuple(listStep0005Rows[0]) != ALLOCATION_STEP0004_HEADERS:
        raise ValueError(
            "割りstep0005 TSVの項目名行が正しくありません。期待値 = "
            + "\\t".join(ALLOCATION_STEP0004_HEADERS)
            + "、実際の値 = "
            + "\\t".join(listStep0005Rows[0])
        )
    listOutputRows: list[list[str]] = [list(FINAL_AREA_STORE_MAPPING_HEADERS)]
    iFirstCenterChangedCount: int = 0
    iSecondCenterChangedCount: int = 0
    for iRowNumber, listValues in enumerate(listStep0005Rows[1:], start=2):
        if not listValues:
            continue
        if len(listValues) != len(ALLOCATION_STEP0004_HEADERS):
            raise ValueError(
                "割りstep0005 TSVのデータ行の列数が正しくありません。行 = "
                + str(iRowNumber)
                + "、期待列数 = 4、実際の列数 = "
                + str(len(listValues))
            )
        pszDistributionCenterName: str = listValues[0]
        pszUpdatedCenterName: str = FINAL_DISTRIBUTION_CENTER_NAMES.get(
            pszDistributionCenterName, pszDistributionCenterName
        )
        if pszDistributionCenterName == "①広島センター":
            iFirstCenterChangedCount += 1
        elif pszDistributionCenterName == "②広島センター(岡山・四国転送分)":
            iSecondCenterChangedCount += 1
        listOutputRows.append(
            [pszUpdatedCenterName, listValues[2], listValues[3]]
        )
    iInputDataRowCount: int = sum(
        1 for listValues in listStep0005Rows[1:] if listValues
    )
    if len(listOutputRows) - 1 != iInputDataRowCount:
        raise ValueError(
            "割りstep0005と対応表のデータ行数が一致しません。入力行数 = "
            + str(iInputDataRowCount)
            + "、出力行数 = "
            + str(len(listOutputRows) - 1)
        )
    return listOutputRows, iFirstCenterChangedCount, iSecondCenterChangedCount


def get_next_final_mapping_backup_path(objOutputPath: Path) -> Path:
    """最終対応表に対応する次の.bk0001形式の連番バックアップを返します。"""
    objPattern: re.Pattern[str] = re.compile(
        r"^"
        + re.escape(objOutputPath.name)
        + r"\.bk([0-9]{4})"
        + re.escape(objOutputPath.suffix)
        + r"$"
    )
    listBackupNumbers: list[int] = []
    for objCandidatePath in objOutputPath.parent.iterdir():
        objMatch: re.Match[str] | None = objPattern.fullmatch(objCandidatePath.name)
        if objMatch is None or not objCandidatePath.is_file():
            continue
        iBackupNumber: int = int(objMatch.group(1))
        if 1 <= iBackupNumber <= MAX_BACKUP_NUMBER:
            listBackupNumbers.append(iBackupNumber)
    iBackupNumber: int = 1 if not listBackupNumbers else max(listBackupNumbers) + 1
    if iBackupNumber > MAX_BACKUP_NUMBER:
        raise ValueError(
            "対応表のバックアップ番号が最大値9999に到達しています。Path = "
            + str(objOutputPath)
        )
    return objOutputPath.with_name(
        objOutputPath.name + f".bk{iBackupNumber:04d}" + objOutputPath.suffix
    )


def replace_final_area_store_mapping_files(
    dictTemporaryPaths: dict[Path, Path], dictBackupPaths: dict[Path, Path]
) -> None:
    """対応表TXT・TSVを一括置換し、失敗時は両方の既存出力を復元します。"""
    listRenamedOutputs: list[tuple[Path, Path]] = []
    listReplacedOutputs: list[Path] = []
    try:
        for objOutputPath, objBackupPath in dictBackupPaths.items():
            if objBackupPath.exists():
                raise FileExistsError(
                    "バックアップ先がすでに存在します。Path = " + str(objBackupPath)
                )
            os.rename(objOutputPath, objBackupPath)
            listRenamedOutputs.append((objOutputPath, objBackupPath))
        for objOutputPath, objTemporaryPath in dictTemporaryPaths.items():
            os.replace(objTemporaryPath, objOutputPath)
            listReplacedOutputs.append(objOutputPath)
    except Exception:
        for objOutputPath in reversed(listReplacedOutputs):
            if objOutputPath.exists():
                objOutputPath.unlink()
        for objOutputPath, objBackupPath in reversed(listRenamedOutputs):
            if objBackupPath.exists():
                os.rename(objBackupPath, objOutputPath)
        raise


def process_final_area_store_mapping_files(
    objStep0005Path: Path,
) -> tuple[Path, Path, int, int, int, dict[Path, Path]]:
    """割りstep0005を再読込し、同一内容の最終対応表TXT・TSVを作ります。"""
    if not objStep0005Path.exists() or not objStep0005Path.is_file():
        raise ValueError(
            "割りstep0005 TSVが見つかりません。Path = " + str(objStep0005Path)
        )
    listStep0005Rows: list[list[str]] = read_tsv_rows(objStep0005Path)
    listOutputRows, iFirstCenterChangedCount, iSecondCenterChangedCount = (
        build_final_area_store_mapping_rows(listStep0005Rows)
    )
    objTextPath, objTsvPath, _ = get_final_area_store_mapping_paths(
        objStep0005Path
    )
    dictTemporaryPaths: dict[Path, Path] = {}
    dictBackupPaths: dict[Path, Path] = {
        objOutputPath: get_next_final_mapping_backup_path(objOutputPath)
        for objOutputPath in (objTextPath, objTsvPath)
        if objOutputPath.exists()
    }
    try:
        for objOutputPath in (objTextPath, objTsvPath):
            objTemporaryPath: Path = create_temporary_path(objOutputPath)
            dictTemporaryPaths[objOutputPath] = objTemporaryPath
            save_tsv_rows(objTemporaryPath, listOutputRows)
            if read_tsv_rows(objTemporaryPath) != listOutputRows:
                raise ValueError(
                    "対応表の保存後検証に失敗しました。Path = "
                    + str(objOutputPath)
                )
        if (
            read_tsv_rows(dictTemporaryPaths[objTextPath])
            != read_tsv_rows(dictTemporaryPaths[objTsvPath])
        ):
            raise ValueError("対応表TXTとTSVの内容が一致しません。")
        replace_final_area_store_mapping_files(
            dictTemporaryPaths, dictBackupPaths
        )
    finally:
        for objTemporaryPath in dictTemporaryPaths.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    return (
        objTextPath,
        objTsvPath,
        len(listOutputRows) - 1,
        iFirstCenterChangedCount,
        iSecondCenterChangedCount,
        dictBackupPaths,
    )


def report_final_area_store_mapping_error(
    objStep0005Path: Path, pszDetailMessage: str
) -> Path:
    """対応表作成エラーを標準エラーと専用エラーファイルへ出力します。"""
    _, _, objErrorPath = get_final_area_store_mapping_paths(objStep0005Path)
    pszErrorMessage: str = (
        "処理結果: エラー\n入力ファイル: "
        + str(objStep0005Path)
        + "\n発生した処理: 朝日注文エリア店舗対応表作成処理\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    write_error_text(str(objErrorPath), pszErrorMessage)
    return objErrorPath


def get_row_value(listValues: list[str], iColumnIndex: int) -> str:
    """指定列が存在すれば値を返し、列不足なら空文字を返します。"""
    if iColumnIndex >= len(listValues):
        return ""
    return listValues[iColumnIndex]


def normalize_distribution_center_name(pszValue: str) -> str:
    """配送センター名の前後空白を除き、括弧を半角へ統一します。"""
    return pszValue.strip().replace("（", "(").replace("）", ")")


def is_distribution_center_name(pszValue: str) -> bool:
    """丸数字で始まりセンターを含む配送センター見出しか返します。"""
    pszNormalizedValue: str = normalize_distribution_center_name(pszValue)
    return (
        pszNormalizedValue != ""
        and pszNormalizedValue[0] in CIRCLED_NUMBERS
        and "センター" in pszNormalizedValue
    )


def find_distribution_centers(
    listRows: list[list[str]],
) -> list[tuple[int, int, str]]:
    """全セルから配送センター見出しの行・列・正規化名を返します。"""
    listCenters: list[tuple[int, int, str]] = []
    for iRowIndex, listValues in enumerate(listRows):
        for iColumnIndex, pszValue in enumerate(listValues):
            if is_distribution_center_name(pszValue):
                listCenters.append(
                    (
                        iRowIndex,
                        iColumnIndex,
                        normalize_distribution_center_name(pszValue),
                    )
                )
    if not listCenters:
        raise ValueError("配送センター見出しが見つかりません。")
    return listCenters


def find_store_groups(listRows: list[list[str]]) -> list[tuple[int, int, int]]:
    """店舗コードと末尾が店舗名の隣接ヘッダーをすべて返します。"""
    listStoreGroups: list[tuple[int, int, int]] = []
    for iRowIndex, listValues in enumerate(listRows):
        for iColumnIndex, pszValue in enumerate(listValues):
            if pszValue != "店舗コード":
                continue
            pszStoreNameHeader: str = get_row_value(listValues, iColumnIndex + 1)
            if pszStoreNameHeader.endswith("店舗名"):
                listStoreGroups.append(
                    (iRowIndex, iColumnIndex, iColumnIndex + 1)
                )
    if not listStoreGroups:
        raise ValueError("店舗コード・店舗名の組が見つかりません。")
    return listStoreGroups


def assign_distribution_center(
    tupleStoreGroup: tuple[int, int, int],
    listCenters: list[tuple[int, int, str]],
) -> tuple[int, int, str]:
    """店舗グループの上方で最も近い見出し行から所属センターを返します。"""
    iHeaderRowIndex, iStoreCodeColumnIndex, _ = tupleStoreGroup
    listPreviousCenters: list[tuple[int, int, str]] = [
        tupleCenter
        for tupleCenter in listCenters
        if tupleCenter[0] <= iHeaderRowIndex
    ]
    if not listPreviousCenters:
        raise ValueError(
            "店舗グループより上に配送センター見出しがありません。行 = "
            + str(iHeaderRowIndex + 1)
            + "、列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    iNearestCenterRowIndex: int = max(
        tupleCenter[0] for tupleCenter in listPreviousCenters
    )
    listSameRowCenters: list[tuple[int, int, str]] = sorted(
        (
            tupleCenter
            for tupleCenter in listPreviousCenters
            if tupleCenter[0] == iNearestCenterRowIndex
            and tupleCenter[1] <= iStoreCodeColumnIndex
        ),
        key=lambda tupleCenter: tupleCenter[1],
    )
    if not listSameRowCenters:
        raise ValueError(
            "店舗グループを配送センターへ関連付けできません。行 = "
            + str(iHeaderRowIndex + 1)
            + "、列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    return listSameRowCenters[-1]


def extract_store_group_rows(
    listRows: list[list[str]],
    tupleStoreGroup: tuple[int, int, int],
    pszDistributionCenterName: str,
) -> list[list[str]]:
    """店舗グループのヘッダー直後から小計まで店舗コード・店舗名を返します。"""
    iHeaderRowIndex, iStoreCodeColumnIndex, iStoreNameColumnIndex = tupleStoreGroup
    listMappingRows: list[list[str]] = []
    bFoundSubtotal: bool = False
    for listValues in listRows[iHeaderRowIndex + 1 :]:
        pszStoreCode: str = get_row_value(listValues, iStoreCodeColumnIndex)
        pszStoreName: str = get_row_value(listValues, iStoreNameColumnIndex)
        if pszStoreCode.strip() == "小計":
            bFoundSubtotal = True
            break
        if pszStoreCode.strip() == "" or pszStoreName.strip() == "":
            continue
        if pszStoreCode == "店舗コード" or pszStoreName.endswith("店舗名"):
            continue
        listMappingRows.append(
            [pszDistributionCenterName, pszStoreCode, pszStoreName]
        )
    if not bFoundSubtotal:
        raise ValueError(
            "店舗グループの小計行が見つかりません。配送センター = "
            + pszDistributionCenterName
            + "、ヘッダー行 = "
            + str(iHeaderRowIndex + 1)
            + "、店舗コード列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    if not listMappingRows:
        raise ValueError(
            "店舗グループに店舗データが見つかりません。配送センター = "
            + pszDistributionCenterName
            + "、ヘッダー行 = "
            + str(iHeaderRowIndex + 1)
            + "、店舗コード列 = "
            + str(iStoreCodeColumnIndex + 1)
        )
    return listMappingRows


def build_area_store_mapping_rows(
    listRows: list[list[str]],
) -> tuple[list[list[str]], int, int]:
    """配送センターと全店舗グループを検出して3列の対応表を返します。"""
    listCenters: list[tuple[int, int, str]] = find_distribution_centers(listRows)
    listStoreGroups: list[tuple[int, int, int]] = find_store_groups(listRows)
    listAssignedGroups: list[
        tuple[tuple[int, int, str], tuple[int, int, int]]
    ] = [
        (assign_distribution_center(tupleStoreGroup, listCenters), tupleStoreGroup)
        for tupleStoreGroup in listStoreGroups
    ]
    listAssignedGroups.sort(
        key=lambda tupleAssignment: (
            tupleAssignment[0][0],
            tupleAssignment[0][1],
            tupleAssignment[1][1],
            tupleAssignment[1][0],
        )
    )
    listMappingRows: list[list[str]] = []
    for tupleCenter, tupleStoreGroup in listAssignedGroups:
        listMappingRows.extend(
            extract_store_group_rows(
                listRows,
                tupleStoreGroup,
                tupleCenter[2],
            )
        )
    if not listMappingRows:
        raise ValueError("店舗データが見つかりません。")
    iCenterCount: int = len(
        {(tupleCenter[0], tupleCenter[1]) for tupleCenter, _ in listAssignedGroups}
    )
    return listMappingRows, iCenterCount, len(listAssignedGroups)


def save_area_store_mapping_tsv(
    objOutputPath: Path, listMappingRows: list[list[str]]
) -> None:
    """3列ヘッダーと週間店舗対応データをUTF-8 TSVへ保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(AREA_STORE_MAPPING_HEADERS)
        objWriter.writerows(listMappingRows)


def get_area_store_mapping_output_path(objInputTsvPath: Path) -> Path:
    """週間TSVと同じフォルダーに作る店舗対応表TSVのパスを返します。"""
    return objInputTsvPath.with_name(AREA_STORE_MAPPING_FILE_NAME)


def process_area_store_mapping_file(
    objInputTsvPath: Path,
) -> tuple[Path, int, int, int]:
    """本州マグロ週間TSVから店舗対応表を独立して作成します。"""
    if not objInputTsvPath.exists() or not objInputTsvPath.is_file():
        raise ValueError("週間TSVが見つかりません。Path = " + str(objInputTsvPath))
    listRows: list[list[str]] = read_tsv_rows(objInputTsvPath)
    try:
        listMappingRows, iCenterCount, iGroupCount = build_area_store_mapping_rows(
            listRows
        )
    except ValueError as objException:
        raise ValueError(
            str(objException) + " Path = " + str(objInputTsvPath)
        ) from objException
    objOutputPath: Path = get_area_store_mapping_output_path(objInputTsvPath)
    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_area_store_mapping_tsv(objTemporaryPath, listMappingRows)
        os.replace(objTemporaryPath, objOutputPath)
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()
    return objOutputPath, len(listMappingRows), iCenterCount, iGroupCount


def find_backup_numbers(objOutputPath: Path) -> list[int]:
    """通常名TSVに対応する既存バックアップの4桁番号を返します。"""
    objPattern: re.Pattern[str] = re.compile(
        r"^" + re.escape(objOutputPath.name) + r"\.bk([0-9]{4})\.tsv$"
    )
    listBackupNumbers: list[int] = []
    for objCandidatePath in objOutputPath.parent.iterdir():
        objMatch: re.Match[str] | None = objPattern.fullmatch(objCandidatePath.name)
        if objMatch is None or not objCandidatePath.is_file():
            continue
        iBackupNumber: int = int(objMatch.group(1))
        if 1 <= iBackupNumber <= MAX_BACKUP_NUMBER:
            listBackupNumbers.append(iBackupNumber)
    return listBackupNumbers


def get_next_backup_path(objOutputPath: Path) -> Path:
    """既存最大番号の次となる.bk%04d.tsvパスを返します。"""
    listBackupNumbers: list[int] = find_backup_numbers(objOutputPath)
    iBackupNumber: int = 1 if not listBackupNumbers else max(listBackupNumbers) + 1
    if iBackupNumber > MAX_BACKUP_NUMBER:
        raise ValueError(
            "バックアップ番号が最大値9999に到達しています。Path = "
            + str(objOutputPath)
        )
    return objOutputPath.with_name(
        objOutputPath.name + f".bk{iBackupNumber:04d}.tsv"
    )


def rename_output_to_backup(objOutputPath: Path) -> Path | None:
    """通常名TSVがあれば次の連番バックアップ名へ変更します。"""
    if not objOutputPath.exists():
        return None
    objBackupPath: Path = get_next_backup_path(objOutputPath)
    if objBackupPath.exists():
        raise FileExistsError(
            "バックアップ先がすでに存在します。Path = " + str(objBackupPath)
        )
    os.rename(objOutputPath, objBackupPath)
    return objBackupPath


def build_warning_text(
    pszInputFileFullPath: str,
    listMissingWorksheetNames: list[str],
    dictOutputPaths: dict[str, Path],
    dictBackupPaths: dict[str, Path],
    dictMappingOutputPaths: dict[str, Path],
    dictMappingBackupPaths: dict[str, Path],
) -> str:
    """未検出シート、作成TSV、旧TSVバックアップを含む警告文を返します。"""
    listLines: list[str] = [
        "処理結果: 警告",
        "入力ファイル: " + os.path.abspath(pszInputFileFullPath),
        "未検出シート: " + ", ".join(listMissingWorksheetNames),
        "警告内容: 対象シートが見つからないため、このシートのTSVは作成しませんでした。",
    ]
    for pszWorksheetName in TARGET_WORKSHEET_NAMES:
        if pszWorksheetName in listMissingWorksheetNames:
            continue
        listLines.append("作成したTSV: " + str(dictOutputPaths[pszWorksheetName]))
    for pszWorksheetName in listMissingWorksheetNames:
        if pszWorksheetName not in dictBackupPaths:
            continue
        listLines.append(
            "旧TSVの変更前パス: " + str(dictOutputPaths[pszWorksheetName])
        )
        listLines.append(
            "旧TSVのバックアップパス: " + str(dictBackupPaths[pszWorksheetName])
        )
    for pszWorksheetName in TARGET_WORKSHEET_NAMES:
        objMappingOutputPath: Path = dictMappingOutputPaths[pszWorksheetName]
        if pszWorksheetName in listMissingWorksheetNames:
            listLines.append(
                pszWorksheetName
                + "対応表: 対象シートがないため作成しませんでした。"
            )
            if pszWorksheetName in dictMappingBackupPaths:
                listLines.append(
                    "旧対応表の変更前パス: " + str(objMappingOutputPath)
                )
                listLines.append(
                    "旧対応表のバックアップパス: "
                    + str(dictMappingBackupPaths[pszWorksheetName])
                )
        else:
            listLines.append("作成予定の対応表: " + str(objMappingOutputPath))
    return "\n".join(listLines) + "\n"


def replace_output_files(
    dictTemporaryPaths: dict[Path, Path],
    dictMissingOutputBackups: dict[Path, Path],
    objWarningPath: Path,
    objTemporaryWarningPath: Path | None,
) -> None:
    """TSV・警告を置換し、欠落シートの旧TSVを連番名へ変更します。"""
    setManagedPaths: set[Path] = set(dictTemporaryPaths.keys()) | {objWarningPath}
    dictRollbackPaths: dict[Path, Path] = {}
    listRenamedMissingOutputs: list[tuple[Path, Path]] = []
    try:
        for objManagedPath in setManagedPaths:
            if not objManagedPath.exists():
                continue
            objRollbackPath: Path = create_temporary_path(objManagedPath)
            shutil.copy2(objManagedPath, objRollbackPath)
            dictRollbackPaths[objManagedPath] = objRollbackPath
        for objOutputPath, objBackupPath in dictMissingOutputBackups.items():
            if objBackupPath.exists():
                raise FileExistsError(
                    "バックアップ先がすでに存在します。Path = " + str(objBackupPath)
                )
            os.rename(objOutputPath, objBackupPath)
            listRenamedMissingOutputs.append((objOutputPath, objBackupPath))
        for objOutputPath, objTemporaryPath in dictTemporaryPaths.items():
            os.replace(objTemporaryPath, objOutputPath)
        if objTemporaryWarningPath is None:
            if objWarningPath.exists():
                objWarningPath.unlink()
        else:
            os.replace(objTemporaryWarningPath, objWarningPath)
    except Exception:
        for objManagedPath in setManagedPaths:
            objRollbackPath = dictRollbackPaths.get(objManagedPath)
            if objRollbackPath is not None and objRollbackPath.exists():
                os.replace(objRollbackPath, objManagedPath)
            elif objManagedPath.exists():
                objManagedPath.unlink()
        for objOutputPath, objBackupPath in reversed(listRenamedMissingOutputs):
            if objBackupPath.exists():
                os.rename(objBackupPath, objOutputPath)
        raise
    finally:
        for objRollbackPath in dictRollbackPaths.values():
            if objRollbackPath.exists():
                objRollbackPath.unlink()


def process_input_file(pszInputFileFullPath: str) -> None:
    """存在する対象Excelシートのセル値から調査用TSVを作成します。"""
    pszValidatedPath: str = validate_input_path(pszInputFileFullPath)
    dictWorksheetResults, listMissingWorksheetNames = read_excel_worksheets(
        pszValidatedPath
    )
    dictOutputPaths: dict[str, Path] = {
        pszWorksheetName: get_output_file_path(pszValidatedPath, pszWorksheetName)
        for pszWorksheetName in TARGET_WORKSHEET_NAMES
    }
    dictMappingOutputPaths: dict[str, Path] = {
        "割り": get_allocation_mapping_output_path(dictOutputPaths["割り"]),
        "本州マグロ(週間)": get_area_store_mapping_output_path(
            dictOutputPaths["本州マグロ(週間)"]
        ),
    }
    dictBackupPaths: dict[str, Path] = {}
    for pszWorksheetName in listMissingWorksheetNames:
        objMissingOutputPath: Path = dictOutputPaths[pszWorksheetName]
        if objMissingOutputPath.exists():
            dictBackupPaths[pszWorksheetName] = get_next_backup_path(
                objMissingOutputPath
            )
    dictMappingBackupPaths: dict[str, Path] = {}
    for pszWorksheetName in listMissingWorksheetNames:
        objMappingOutputPath: Path = dictMappingOutputPaths[pszWorksheetName]
        if objMappingOutputPath.exists():
            dictMappingBackupPaths[pszWorksheetName] = get_next_backup_path(
                objMappingOutputPath
            )
    dictTemporaryPaths: dict[Path, Path] = {}
    objWarningPath: Path = Path(get_warning_file_full_path(pszValidatedPath))
    objTemporaryWarningPath: Path | None = None
    try:
        for pszWorksheetName, (listRows, _) in dictWorksheetResults.items():
            objOutputPath: Path = dictOutputPaths[pszWorksheetName]
            objTemporaryPath: Path = create_temporary_path(objOutputPath)
            dictTemporaryPaths[objOutputPath] = objTemporaryPath
            save_tsv_rows(objTemporaryPath, listRows)
        if dictWorksheetResults and listMissingWorksheetNames:
            objTemporaryWarningPath = create_temporary_path(objWarningPath)
            write_error_text(
                str(objTemporaryWarningPath),
                build_warning_text(
                    pszValidatedPath,
                    listMissingWorksheetNames,
                    dictOutputPaths,
                    dictBackupPaths,
                    dictMappingOutputPaths,
                    dictMappingBackupPaths,
                ),
            )
        dictMissingOutputBackups: dict[Path, Path] = {
            dictOutputPaths[pszWorksheetName]: objBackupPath
            for pszWorksheetName, objBackupPath in dictBackupPaths.items()
        }
        for pszWorksheetName, objMappingBackupPath in (
            dictMappingBackupPaths.items()
        ):
            dictMissingOutputBackups[
                dictMappingOutputPaths[pszWorksheetName]
            ] = objMappingBackupPath
        replace_output_files(
            dictTemporaryPaths,
            dictMissingOutputBackups,
            objWarningPath,
            objTemporaryWarningPath,
        )
    finally:
        for objTemporaryPath in dictTemporaryPaths.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
        if objTemporaryWarningPath is not None and objTemporaryWarningPath.exists():
            objTemporaryWarningPath.unlink()
    if not dictWorksheetResults:
        raise ValueError(
            "対象シートが見つかりません。対象シート = "
            + ", ".join(TARGET_WORKSHEET_NAMES)
        )
    objCreatedAllocationMappingPath: Path | None = None
    iAllocationMappingRowCount: int = 0
    objStoreCodeMismatchPath: Path | None = None
    iStoreCodeMismatchCount: int = 0
    objStoreNameMismatchPath: Path | None = None
    iStoreNameMismatchCount: int = 0
    dictAllocationMismatchBackupPaths: dict[Path, Path] = {}
    objFormalStoreNamePath: Path | None = None
    objAllocationStep0002Path: Path | None = None
    iAllocationStep0002RowCount: int = 0
    iFormalStoreNameCount: int = 0
    iFormalStoreNameChangedCount: int = 0
    objAllocationStep0002BackupPath: Path | None = None
    objAllocationStep0003Path: Path | None = None
    iAllocationStep0003InputRowCount: int = 0
    iAllocationStep0003OutputRowCount: int = 0
    objAllocationStep0003BackupPath: Path | None = None
    objAllocationStep0004Path: Path | None = None
    iAllocationStep0004RowCount: int = 0
    iAllocationStep0004MatchedRowCount: int = 0
    iAllocationStep0004UnmatchedRowCount: int = 0
    objAllocationStep0004BackupPath: Path | None = None
    objAllocationStep0005Path: Path | None = None
    iAllocationStep0005InputRowCount: int = 0
    iAllocationStep0005DeletedRowCount: int = 0
    iAllocationStep0005FilledRowCount: int = 0
    iAllocationStep0005UnchangedRowCount: int = 0
    iAllocationStep0005OutputRowCount: int = 0
    objAllocationStep0005BackupPath: Path | None = None
    objFinalMappingTextPath: Path | None = None
    objFinalMappingTsvPath: Path | None = None
    iFinalMappingRowCount: int = 0
    iFinalFirstCenterChangedCount: int = 0
    iFinalSecondCenterChangedCount: int = 0
    dictFinalMappingBackupPaths: dict[Path, Path] = {}
    objCreatedMappingPath: Path | None = None
    iMappingRowCount: int = 0
    iMappingCenterCount: int = 0
    iMappingGroupCount: int = 0
    listMappingResultLines: list[str] = []
    listMappingErrorLines: list[str] = []
    if "割り" in dictWorksheetResults:
        try:
            (
                objCreatedAllocationMappingPath,
                iAllocationMappingRowCount,
            ) = process_allocation_mapping_file(dictOutputPaths["割り"])
            (
                objStoreCodeMismatchPath,
                iStoreCodeMismatchCount,
                objStoreNameMismatchPath,
                iStoreNameMismatchCount,
                dictAllocationMismatchBackupPaths,
            ) = process_allocation_mismatch_files(objCreatedAllocationMappingPath)
            listMappingResultLines.append(
                "処理A（割り対応表）: 成功\n出力ファイル: "
                + str(objCreatedAllocationMappingPath)
                + "\n店舗コード不一致ファイル: "
                + str(objStoreCodeMismatchPath)
                + "\n店舗コード不一致件数: "
                + str(iStoreCodeMismatchCount)
                + "\n店舗名不一致ファイル: "
                + str(objStoreNameMismatchPath)
                + "\n店舗名不一致件数: "
                + str(iStoreNameMismatchCount)
            )
        except Exception as objException:
            try:
                objBackupPath = rename_output_to_backup(
                    dictMappingOutputPaths["割り"]
                )
                pszBackupDetail: str = (
                    ""
                    if objBackupPath is None
                    else "\n旧出力バックアップ: " + str(objBackupPath)
                )
            except Exception as objBackupException:
                pszBackupDetail = (
                    "\n旧出力のバックアップにも失敗しました。Detail = "
                    + str(objBackupException)
                )
            listMappingErrorLines.append(
                "処理A（割り対応表）: エラー\nエラー内容: "
                + str(objException)
                + pszBackupDetail
            )
    else:
        listMappingResultLines.append("処理A（割り対応表）: スキップ")
    if (
        objCreatedAllocationMappingPath is not None
        and objStoreCodeMismatchPath is not None
        and objStoreNameMismatchPath is not None
    ):
        try:
            (
                objFormalStoreNamePath,
                objAllocationStep0002Path,
                iAllocationStep0002RowCount,
                iFormalStoreNameCount,
                iFormalStoreNameChangedCount,
                objAllocationStep0002BackupPath,
            ) = process_allocation_formal_store_names(
                objCreatedAllocationMappingPath
            )
            _, _, objStep0002ErrorPath = get_allocation_formal_store_name_paths(
                objCreatedAllocationMappingPath
            )
            if objStep0002ErrorPath.exists():
                objStep0002ErrorPath.unlink()
            listMappingResultLines.append(
                "処理C（割り正式店舗名反映）: 成功\n出力ファイル: "
                + str(objAllocationStep0002Path)
            )
        except Exception as objException:
            objFormalStoreNamePath, _, _ = get_allocation_formal_store_name_paths(
                objCreatedAllocationMappingPath
            )
            try:
                objStep0002ErrorPath = report_allocation_step0002_error(
                    objCreatedAllocationMappingPath,
                    objFormalStoreNamePath,
                    str(objException),
                )
                pszErrorFileDetail: str = "\nエラーファイル: " + str(
                    objStep0002ErrorPath
                )
            except Exception as objErrorFileException:
                pszErrorFileDetail = (
                    "\nstep0002_error.txtの保存にも失敗しました。Detail = "
                    + str(objErrorFileException)
                )
            listMappingErrorLines.append(
                "処理C（割り正式店舗名反映）: エラー\nエラー内容: "
                + str(objException)
                + pszErrorFileDetail
            )
    elif "割り" not in dictWorksheetResults:
        listMappingResultLines.append("処理C（割り正式店舗名反映）: スキップ")
    if objAllocationStep0002Path is not None:
        try:
            (
                objAllocationStep0003Path,
                iAllocationStep0003InputRowCount,
                iAllocationStep0003OutputRowCount,
                objAllocationStep0003BackupPath,
            ) = process_allocation_step0003_file(objAllocationStep0002Path)
            _, objStep0003ErrorPath = get_allocation_step0003_paths(
                objAllocationStep0002Path
            )
            if objStep0003ErrorPath.exists():
                objStep0003ErrorPath.unlink()
            listMappingResultLines.append(
                "処理D（割りstep0003作成）: 成功\n出力ファイル: "
                + str(objAllocationStep0003Path)
            )
        except Exception as objException:
            try:
                objStep0003ErrorPath = report_allocation_step0003_error(
                    objAllocationStep0002Path, str(objException)
                )
                pszErrorFileDetail = "\nエラーファイル: " + str(
                    objStep0003ErrorPath
                )
            except Exception as objErrorFileException:
                pszErrorFileDetail = (
                    "\nstep0003_error.txtの保存にも失敗しました。Detail = "
                    + str(objErrorFileException)
                )
            listMappingErrorLines.append(
                "処理D（割りstep0003作成）: エラー\nエラー内容: "
                + str(objException)
                + pszErrorFileDetail
            )
    elif "割り" not in dictWorksheetResults:
        listMappingResultLines.append("処理D（割りstep0003作成）: スキップ")
    if "本州マグロ(週間)" in dictWorksheetResults:
        try:
            (
                objCreatedMappingPath,
                iMappingRowCount,
                iMappingCenterCount,
                iMappingGroupCount,
            ) = process_area_store_mapping_file(
                dictOutputPaths["本州マグロ(週間)"]
            )
            listMappingResultLines.append(
                "処理B（週間配送センター対応表）: 成功\n出力ファイル: "
                + str(objCreatedMappingPath)
            )
        except Exception as objException:
            try:
                objBackupPath = rename_output_to_backup(
                    dictMappingOutputPaths["本州マグロ(週間)"]
                )
                pszBackupDetail = (
                    ""
                    if objBackupPath is None
                    else "\n旧出力バックアップ: " + str(objBackupPath)
                )
            except Exception as objBackupException:
                pszBackupDetail = (
                    "\n旧出力のバックアップにも失敗しました。Detail = "
                    + str(objBackupException)
                )
            listMappingErrorLines.append(
                "処理B（週間配送センター対応表）: エラー\nエラー内容: "
                + str(objException)
                + pszBackupDetail
            )
    else:
        listMappingResultLines.append("処理B（週間配送センター対応表）: スキップ")
    if objCreatedMappingPath is not None and objAllocationStep0003Path is not None:
        try:
            (
                objAllocationStep0004Path,
                iAllocationStep0004RowCount,
                iAllocationStep0004MatchedRowCount,
                iAllocationStep0004UnmatchedRowCount,
                objAllocationStep0004BackupPath,
            ) = process_allocation_step0004_file(
                objCreatedMappingPath, objAllocationStep0003Path
            )
            _, objStep0004ErrorPath = get_allocation_step0004_paths(
                objAllocationStep0003Path
            )
            if objStep0004ErrorPath.exists():
                objStep0004ErrorPath.unlink()
            listMappingResultLines.append(
                "処理E（割りstep0004作成）: 成功\n出力ファイル: "
                + str(objAllocationStep0004Path)
            )
        except Exception as objException:
            try:
                objStep0004ErrorPath = report_allocation_step0004_error(
                    objCreatedMappingPath,
                    objAllocationStep0003Path,
                    str(objException),
                )
                pszErrorFileDetail = "\nエラーファイル: " + str(
                    objStep0004ErrorPath
                )
            except Exception as objErrorFileException:
                pszErrorFileDetail = (
                    "\nstep0004_error.txtの保存にも失敗しました。Detail = "
                    + str(objErrorFileException)
                )
            listMappingErrorLines.append(
                "処理E（割りstep0004作成）: エラー\nエラー内容: "
                + str(objException)
                + pszErrorFileDetail
            )
    elif objAllocationStep0003Path is not None:
        objExpectedWeeklyMappingPath: Path = dictMappingOutputPaths[
            "本州マグロ(週間)"
        ]
        pszStep0004ErrorMessage: str = (
            "今回の実行で週間step0001 TSVが作成されませんでした。Path = "
            + str(objExpectedWeeklyMappingPath)
        )
        try:
            objStep0004ErrorPath = report_allocation_step0004_error(
                objExpectedWeeklyMappingPath,
                objAllocationStep0003Path,
                pszStep0004ErrorMessage,
            )
            pszErrorFileDetail = "\nエラーファイル: " + str(
                objStep0004ErrorPath
            )
        except Exception as objErrorFileException:
            pszErrorFileDetail = (
                "\nstep0004_error.txtの保存にも失敗しました。Detail = "
                + str(objErrorFileException)
            )
        listMappingErrorLines.append(
            "処理E（割りstep0004作成）: エラー\nエラー内容: "
            + pszStep0004ErrorMessage
            + pszErrorFileDetail
        )
    else:
        listMappingResultLines.append("処理E（割りstep0004作成）: スキップ")
    if objAllocationStep0004Path is not None:
        try:
            (
                objAllocationStep0005Path,
                iAllocationStep0005InputRowCount,
                iAllocationStep0005DeletedRowCount,
                iAllocationStep0005FilledRowCount,
                iAllocationStep0005UnchangedRowCount,
                iAllocationStep0005OutputRowCount,
                objAllocationStep0005BackupPath,
            ) = process_allocation_step0005_file(objAllocationStep0004Path)
            _, objStep0005ErrorPath = get_allocation_step0005_paths(
                objAllocationStep0004Path
            )
            if objStep0005ErrorPath.exists():
                objStep0005ErrorPath.unlink()
            listMappingResultLines.append(
                "処理F（割りstep0005作成）: 成功\n出力ファイル: "
                + str(objAllocationStep0005Path)
            )
        except Exception as objException:
            try:
                objStep0005ErrorPath = report_allocation_step0005_error(
                    objAllocationStep0004Path, str(objException)
                )
                pszErrorFileDetail = "\nエラーファイル: " + str(
                    objStep0005ErrorPath
                )
            except Exception as objErrorFileException:
                pszErrorFileDetail = (
                    "\nstep0005_error.txtの保存にも失敗しました。Detail = "
                    + str(objErrorFileException)
                )
            listMappingErrorLines.append(
                "処理F（割りstep0005作成）: エラー\nエラー内容: "
                + str(objException)
                + pszErrorFileDetail
            )
    else:
        listMappingResultLines.append("処理F（割りstep0005作成）: スキップ")
    if objAllocationStep0005Path is not None:
        try:
            (
                objFinalMappingTextPath,
                objFinalMappingTsvPath,
                iFinalMappingRowCount,
                iFinalFirstCenterChangedCount,
                iFinalSecondCenterChangedCount,
                dictFinalMappingBackupPaths,
            ) = process_final_area_store_mapping_files(objAllocationStep0005Path)
            _, _, objFinalMappingErrorPath = get_final_area_store_mapping_paths(
                objAllocationStep0005Path
            )
            if objFinalMappingErrorPath.exists():
                objFinalMappingErrorPath.unlink()
            listMappingResultLines.append(
                "処理G（朝日注文エリア店舗対応表作成）: 成功\n出力TXT: "
                + str(objFinalMappingTextPath)
                + "\n出力TSV: "
                + str(objFinalMappingTsvPath)
            )
        except Exception as objException:
            try:
                objFinalMappingErrorPath = report_final_area_store_mapping_error(
                    objAllocationStep0005Path, str(objException)
                )
                pszErrorFileDetail = "\nエラーファイル: " + str(
                    objFinalMappingErrorPath
                )
            except Exception as objErrorFileException:
                pszErrorFileDetail = (
                    "\n対応表_error.txtの保存にも失敗しました。Detail = "
                    + str(objErrorFileException)
                )
            listMappingErrorLines.append(
                "処理G（朝日注文エリア店舗対応表作成）: エラー\nエラー内容: "
                + str(objException)
                + pszErrorFileDetail
            )
    else:
        listMappingResultLines.append(
            "処理G（朝日注文エリア店舗対応表作成）: スキップ"
        )
    if listMappingErrorLines:
        raise ValueError(
            "\n\n".join(listMappingErrorLines + listMappingResultLines)
        )
    remove_old_error_file(pszValidatedPath)
    print("朝日注文エリア店舗対応調査用TSVファイルを作成しました。")
    print("Input: " + pszValidatedPath)
    for pszWorksheetName in TARGET_WORKSHEET_NAMES:
        if pszWorksheetName not in dictWorksheetResults:
            continue
        listRows, iColumnCount = dictWorksheetResults[pszWorksheetName]
        print("Worksheet: " + pszWorksheetName)
        print("TSV: " + str(dictOutputPaths[pszWorksheetName]))
        print("Rows: " + str(len(listRows)))
        print("Columns: " + str(iColumnCount))
    for pszWorksheetName in listMissingWorksheetNames:
        print("Warning: 対象シートが見つかりません。Sheet = " + pszWorksheetName)
        if pszWorksheetName in dictBackupPaths:
            print("Backup: " + str(dictBackupPaths[pszWorksheetName]))
    if listMissingWorksheetNames:
        print("Warning File: " + str(objWarningPath))
    if objCreatedMappingPath is not None:
        print(
            "Weekly Area Store Mapping Input: "
            + str(dictOutputPaths["本州マグロ(週間)"])
        )
        print("Weekly Area Store Mapping TSV: " + str(objCreatedMappingPath))
        print("Weekly Area Store Mapping Rows: " + str(iMappingRowCount))
        print("Weekly Area Store Mapping Centers: " + str(iMappingCenterCount))
        print("Weekly Area Store Mapping Groups: " + str(iMappingGroupCount))
    if objCreatedAllocationMappingPath is not None:
        print("Allocation Area Store Mapping Result: Success")
        print("Allocation Area Store Mapping Input: " + str(dictOutputPaths["割り"]))
        print(
            "Allocation Area Store Mapping TSV: "
            + str(objCreatedAllocationMappingPath)
        )
        print(
            "Allocation Area Store Mapping Rows: "
            + str(iAllocationMappingRowCount)
        )
    if objStoreCodeMismatchPath is not None and objStoreNameMismatchPath is not None:
        print("Allocation Store Mismatch Result: Success")
        print("Allocation Store Mismatch Input: " + str(objCreatedAllocationMappingPath))
        print("Allocation Store Code Mismatch TSV: " + str(objStoreCodeMismatchPath))
        print("Allocation Store Code Mismatch Rows: " + str(iStoreCodeMismatchCount))
        print("Allocation Store Name Mismatch TSV: " + str(objStoreNameMismatchPath))
        print("Allocation Store Name Mismatch Rows: " + str(iStoreNameMismatchCount))
        for objOutputPath, objBackupPath in dictAllocationMismatchBackupPaths.items():
            print("Allocation Store Mismatch Previous TSV: " + str(objOutputPath))
            print("Allocation Store Mismatch Backup TSV: " + str(objBackupPath))
    if objAllocationStep0002Path is not None and objFormalStoreNamePath is not None:
        print("Allocation Formal Store Name Result: Success")
        print("Allocation Formal Store Name Input: " + str(objFormalStoreNamePath))
        print("Allocation Formal Store Name TSV: " + str(objAllocationStep0002Path))
        print(
            "Allocation Formal Store Name Rows: "
            + str(iAllocationStep0002RowCount)
        )
        print(
            "Allocation Formal Store Name Entries: " + str(iFormalStoreNameCount)
        )
        print(
            "Allocation Formal Store Name Changed Rows: "
            + str(iFormalStoreNameChangedCount)
        )
        if objAllocationStep0002BackupPath is not None:
            print(
                "Allocation Formal Store Name Backup TSV: "
                + str(objAllocationStep0002BackupPath)
            )
    if objAllocationStep0003Path is not None:
        print("Allocation Step0003 Result: Success")
        print("Allocation Step0003 Input: " + str(objAllocationStep0002Path))
        print("Allocation Step0003 TSV: " + str(objAllocationStep0003Path))
        print(
            "Allocation Step0003 Input Rows: "
            + str(iAllocationStep0003InputRowCount)
        )
        print(
            "Allocation Step0003 Output Rows: "
            + str(iAllocationStep0003OutputRowCount)
        )
        print("Allocation Step0003 Removed Columns: APEX店舗コード, APEX店舗名")
        if objAllocationStep0003BackupPath is not None:
            print(
                "Allocation Step0003 Backup TSV: "
                + str(objAllocationStep0003BackupPath)
            )
    if objAllocationStep0004Path is not None:
        print("Allocation Step0004 Result: Success")
        print("Allocation Step0004 Weekly Input: " + str(objCreatedMappingPath))
        print(
            "Allocation Step0004 Allocation Input: "
            + str(objAllocationStep0003Path)
        )
        print("Allocation Step0004 TSV: " + str(objAllocationStep0004Path))
        print("Allocation Step0004 Rows: " + str(iAllocationStep0004RowCount))
        print(
            "Allocation Step0004 Center Matched Rows: "
            + str(iAllocationStep0004MatchedRowCount)
        )
        print(
            "Allocation Step0004 Center Unmatched Rows: "
            + str(iAllocationStep0004UnmatchedRowCount)
        )
        if objAllocationStep0004BackupPath is not None:
            print(
                "Allocation Step0004 Backup TSV: "
                + str(objAllocationStep0004BackupPath)
            )
    if objAllocationStep0005Path is not None:
        print("Allocation Step0005 Result: Success")
        print("Allocation Step0005 Input: " + str(objAllocationStep0004Path))
        print("Allocation Step0005 TSV: " + str(objAllocationStep0005Path))
        print(
            "Allocation Step0005 Input Rows: "
            + str(iAllocationStep0005InputRowCount)
        )
        print(
            "Allocation Step0005 Deleted Rows: "
            + str(iAllocationStep0005DeletedRowCount)
        )
        print(
            "Allocation Step0005 Center Filled Rows: "
            + str(iAllocationStep0005FilledRowCount)
        )
        print(
            "Allocation Step0005 Unchanged Rows: "
            + str(iAllocationStep0005UnchangedRowCount)
        )
        print(
            "Allocation Step0005 Output Rows: "
            + str(iAllocationStep0005OutputRowCount)
        )
        if objAllocationStep0005BackupPath is not None:
            print(
                "Allocation Step0005 Backup TSV: "
                + str(objAllocationStep0005BackupPath)
            )
    if objFinalMappingTextPath is not None and objFinalMappingTsvPath is not None:
        print("Final Area Store Mapping Result: Success")
        print("Final Area Store Mapping Input: " + str(objAllocationStep0005Path))
        print("Final Area Store Mapping TXT: " + str(objFinalMappingTextPath))
        print("Final Area Store Mapping TSV: " + str(objFinalMappingTsvPath))
        print("Final Area Store Mapping Rows: " + str(iFinalMappingRowCount))
        print(
            "Final Area Store Mapping ①広島センター Changed Rows: "
            + str(iFinalFirstCenterChangedCount)
        )
        print(
            "Final Area Store Mapping ②広島センター(岡山・四国転送分) Changed Rows: "
            + str(iFinalSecondCenterChangedCount)
        )
        for objOutputPath, objBackupPath in dictFinalMappingBackupPaths.items():
            print("Final Area Store Mapping Previous File: " + str(objOutputPath))
            print("Final Area Store Mapping Backup File: " + str(objBackupPath))


def main() -> int:
    """引数を確認して処理し、成功0・失敗1の終了コードを返します。"""
    if len(sys.argv) != 2:
        pszScriptFileName: str = os.path.basename(__file__)
        pszErrorMessage: str = (
            "Error: 入力Excelファイルパスを1件指定してください。\n"
            + "Usage: python "
            + pszScriptFileName
            + " <input_xlsx_file_path>\n"
        )
        print(pszErrorMessage, file=sys.stderr, end="")
        pszErrorFileFullPath: str = (
            os.path.splitext(pszScriptFileName)[0] + "_error_argument.txt"
        )
        try:
            write_error_text(pszErrorFileFullPath, pszErrorMessage)
        except OSError as objException:
            print(
                "Error: 引数エラーファイルを保存できません。Detail = "
                + str(objException),
                file=sys.stderr,
            )
        return 1
    pszInputFileFullPath: str = sys.argv[1]
    try:
        process_input_file(pszInputFileFullPath)
    except Exception as objException:
        report_processing_error(
            pszInputFileFullPath,
            "朝日注文エリア店舗対応調査TSV作成処理",
            str(objException),
        )
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
