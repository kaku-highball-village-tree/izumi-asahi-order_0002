# -- coding: utf-8 --
###############################################################
#
# highlight_red_color_changed_order_quantity_Cmd.py
#
# pip install openpyxl
#
###############################################################

import csv
import io
import os
import re
import sys
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Color
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


EXPECTED_HEADERS: tuple[str, str, str, str] = (
    "productCode",
    "productName",
    "spec",
    "quantity",
)
RED_FONT_COLOR: str = "FFFF0000"
MAX_HISTORY_NUMBER: int = 9999


###############################################################
#
# OrderRow
#
###############################################################
class OrderRow:
    """ExcelまたはTSVから取得した1商品の4列を保持します。

    Excel由来の行では、数量を赤色にするためにquantity_cellも保持します。
    TSV由来の行ではquantity_cellはNoneです。
    """

    def __init__(
        self,
        product_code: str,
        product_name: str,
        spec: str,
        quantity: object,
        quantity_cell: Cell | None = None,
    ) -> None:
        """4列の値と、必要な場合はExcelのquantityセルを初期設定します。"""
        self.product_code: str = product_code
        self.product_name: str = product_name
        self.spec: str = spec
        self.quantity: object = quantity
        self.quantity_cell: Cell | None = quantity_cell


###############################################################
#
# write_error_text
#
###############################################################
def write_error_text(
    pszOutputFileFullPath: str,
    pszErrorMessage: str,
) -> None:
    """エラーメッセージをUTF-8のテキストファイルへ上書き保存します。

    Args:
        pszOutputFileFullPath: エラーファイルの保存先です。
        pszErrorMessage: 利用者へ知らせるエラー内容です。

    Returns:
        戻り値はありません。
    """
    pszDirectoryFullPath: str = os.path.dirname(pszOutputFileFullPath)
    if pszDirectoryFullPath != "":
        os.makedirs(pszDirectoryFullPath, exist_ok=True)
    with open(pszOutputFileFullPath, mode="w", encoding="utf-8", newline="") as objFile:
        objFile.write(pszErrorMessage.rstrip("\n") + "\n")


###############################################################
#
# get_error_file_full_path
#
###############################################################
def get_error_file_full_path(pszInputFileFullPath: str) -> str:
    """入力Excelと同じフォルダーに作る_error.txtのパスを返します。"""
    pszDirectoryFullPath: str = os.path.dirname(os.path.abspath(pszInputFileFullPath))
    pszBaseNameWithoutExtension: str = os.path.splitext(
        os.path.basename(pszInputFileFullPath)
    )[0]
    return os.path.join(
        pszDirectoryFullPath,
        pszBaseNameWithoutExtension + "_error.txt",
    )


###############################################################
#
# report_processing_error
#
###############################################################
def report_processing_error(
    pszInputFileFullPath: str,
    pszProcessName: str,
    pszDetailMessage: str,
) -> None:
    """標準エラーと入力Excel用_error.txtへ同じエラーを出力します。"""
    pszErrorMessage: str = (
        "処理結果: エラー\n"
        + "入力Excel: "
        + os.path.abspath(pszInputFileFullPath)
        + "\n"
        + "発生した処理: "
        + pszProcessName
        + "\n"
        + "エラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    try:
        write_error_text(get_error_file_full_path(pszInputFileFullPath), pszErrorMessage)
    except OSError as objException:
        print(
            "Error: _error.txtの保存にも失敗しました。Detail = " + str(objException),
            file=sys.stderr,
        )


###############################################################
#
# remove_old_error_file
#
###############################################################
def remove_old_error_file(pszInputFileFullPath: str) -> None:
    """正常終了後、以前の処理で作られた_error.txtがあれば削除します。"""
    pszErrorFileFullPath: str = get_error_file_full_path(pszInputFileFullPath)
    if os.path.exists(pszErrorFileFullPath):
        os.remove(pszErrorFileFullPath)


###############################################################
#
# validate_excel_path
#
###############################################################
def validate_excel_path(pszInputFileFullPath: str) -> str:
    """入力パスが存在する.xlsxファイルであることを確認します。

    Returns:
        検証後の絶対パスを返します。

    Raises:
        ValueError: パス、ファイル種別、拡張子に問題がある場合です。
    """
    pszAbsolutePath: str = os.path.abspath(pszInputFileFullPath)
    if not os.path.exists(pszAbsolutePath):
        raise ValueError("入力Excelファイルが見つかりません。Path = " + pszAbsolutePath)
    if not os.path.isfile(pszAbsolutePath):
        raise ValueError("入力パスがファイルではありません。Path = " + pszAbsolutePath)
    if os.path.splitext(pszAbsolutePath)[1].lower() != ".xlsx":
        raise ValueError("入力ファイルの拡張子は.xlsxではありません。Path = " + pszAbsolutePath)
    return pszAbsolutePath


###############################################################
#
# load_excel
#
###############################################################
def load_excel(pszInputFileFullPath: str) -> Workbook:
    """openpyxlでExcelを読み込み、既存書式を含むWorkbookを返します。"""
    return load_workbook(pszInputFileFullPath)


###############################################################
#
# normalize_header
#
###############################################################
def normalize_header(objValue: object) -> str:
    """見出し比較のため、セル値を前後空白のない文字列へ変換します。"""
    if objValue is None:
        return ""
    return str(objValue).strip()


###############################################################
#
# find_target_worksheet
#
###############################################################
def find_target_worksheet(objWorkbook: Workbook) -> Worksheet:
    """A1～D1が仕様どおりのシートを探し、1枚だけなら返します。

    Raises:
        ValueError: 該当シートが0枚または複数枚の場合です。
    """
    listTargetWorksheets: list[Worksheet] = []
    for objWorksheet in objWorkbook.worksheets:
        tupleHeaders: tuple[str, str, str, str] = tuple(
            normalize_header(objWorksheet.cell(row=1, column=iColumn).value)
            for iColumn in range(1, 5)
        )  # type: ignore[assignment]
        if tupleHeaders == EXPECTED_HEADERS:
            listTargetWorksheets.append(objWorksheet)

    if len(listTargetWorksheets) == 0:
        raise ValueError(
            "A1～D1がproductCode、productName、spec、quantityのシートが見つかりません。"
        )
    if len(listTargetWorksheets) > 1:
        pszSheetNames: str = ", ".join(
            objWorksheet.title for objWorksheet in listTargetWorksheets
        )
        raise ValueError(
            "対象シートが複数見つかりました。対象シート = " + pszSheetNames
        )
    return listTargetWorksheets[0]


###############################################################
#
# normalize_text_cell
#
###############################################################
def normalize_text_cell(objValue: object) -> str:
    """TSVへ保存する文字列セルを、Noneなら空文字として正規化します。"""
    if objValue is None:
        return ""
    return str(objValue)


###############################################################
#
# normalize_quantity_for_comparison
#
###############################################################
def normalize_quantity_for_comparison(objValue: object) -> Decimal | None:
    """数量を比較用Decimalへ変換し、空欄はNoneとして区別します。

    1、1.0、文字列の「1」は同じDecimalになります。数式や数値に変換
    できない文字列は誤判定を避けるためValueErrorにします。
    """
    if objValue is None:
        return None
    if isinstance(objValue, str):
        pszValue: str = objValue.strip()
        if pszValue == "":
            return None
        if pszValue.startswith("="):
            raise ValueError("quantityに数式が入力されています。Value = " + pszValue)
    elif isinstance(objValue, bool):
        raise ValueError("quantityに真偽値が入力されています。Value = " + str(objValue))
    else:
        pszValue = str(objValue)

    try:
        objQuantity: Decimal = Decimal(pszValue)
    except (InvalidOperation, ValueError) as objException:
        raise ValueError("quantityを数値として解釈できません。Value = " + pszValue) from objException
    if not objQuantity.is_finite():
        raise ValueError("quantityに有限でない数値が入力されています。Value = " + pszValue)
    return objQuantity


###############################################################
#
# quantity_to_tsv_text
#
###############################################################
def quantity_to_tsv_text(objValue: object) -> str:
    """Excelの数量を、値を変えずTSVへ書き込める文字列に変換します。"""
    if objValue is None:
        return ""
    return str(objValue)


###############################################################
#
# validate_unique_product_codes
#
###############################################################
def validate_unique_product_codes(listOrderRows: list[OrderRow], pszSourceName: str) -> None:
    """商品コードが重複していないことを確認します。

    Raises:
        ValueError: 同じproductCodeが複数行に存在する場合です。
    """
    setSeenCodes: set[str] = set()
    setDuplicateCodes: set[str] = set()
    for objOrderRow in listOrderRows:
        if objOrderRow.product_code in setSeenCodes:
            setDuplicateCodes.add(objOrderRow.product_code)
        setSeenCodes.add(objOrderRow.product_code)
    if setDuplicateCodes:
        raise ValueError(
            pszSourceName
            + "でproductCodeが重複しています。productCode = "
            + ", ".join(sorted(setDuplicateCodes))
        )


###############################################################
#
# read_excel_rows
#
###############################################################
def read_excel_rows(objWorksheet: Worksheet) -> list[OrderRow]:
    """対象シートの2行目以降からA～D列の発注データを読み取ります。"""
    listOrderRows: list[OrderRow] = []
    for iRow in range(2, objWorksheet.max_row + 1):
        listValues: list[object] = [
            objWorksheet.cell(row=iRow, column=iColumn).value
            for iColumn in range(1, 5)
        ]
        if all(objValue is None or str(objValue).strip() == "" for objValue in listValues):
            continue

        pszProductCode: str = normalize_text_cell(listValues[0]).strip()
        if pszProductCode == "":
            raise ValueError(
                str(iRow) + "行目はproductCodeが空ですが、ほかの列に値があります。"
            )
        normalize_quantity_for_comparison(listValues[3])
        listOrderRows.append(
            OrderRow(
                pszProductCode,
                normalize_text_cell(listValues[1]),
                normalize_text_cell(listValues[2]),
                listValues[3],
                objWorksheet.cell(row=iRow, column=4),
            )
        )

    validate_unique_product_codes(listOrderRows, "Excel")
    return listOrderRows


###############################################################
#
# build_history_patterns
#
###############################################################
def build_history_patterns(pszExcelBaseName: str) -> tuple[re.Pattern[str], re.Pattern[str]]:
    """対象Excel専用の履歴TSVとlatest TSVを判定する正規表現を返します。"""
    pszEscapedBaseName: str = re.escape(pszExcelBaseName)
    objHistoryPattern: re.Pattern[str] = re.compile(
        r"^" + pszEscapedBaseName + r"_(\d{4})\.tsv$"
    )
    objLatestPattern: re.Pattern[str] = re.compile(
        r"^" + pszEscapedBaseName + r"_(\d{4})_latest\.tsv$"
    )
    return objHistoryPattern, objLatestPattern


###############################################################
#
# find_history_files
#
###############################################################
def find_history_files(
    pszInputFileFullPath: str,
) -> tuple[dict[int, Path], dict[int, Path]]:
    """Excelと同じフォルダーから、対応する履歴とlatestを検索します。"""
    objExcelPath: Path = Path(pszInputFileFullPath)
    objHistoryPattern, objLatestPattern = build_history_patterns(objExcelPath.stem)
    dictHistoryFiles: dict[int, Path] = {}
    dictLatestFiles: dict[int, Path] = {}

    for objPath in objExcelPath.parent.iterdir():
        if not objPath.is_file():
            continue
        objHistoryMatch = objHistoryPattern.fullmatch(objPath.name)
        if objHistoryMatch is not None:
            dictHistoryFiles[int(objHistoryMatch.group(1))] = objPath
            continue
        objLatestMatch = objLatestPattern.fullmatch(objPath.name)
        if objLatestMatch is not None:
            dictLatestFiles[int(objLatestMatch.group(1))] = objPath
    return dictHistoryFiles, dictLatestFiles


###############################################################
#
# validate_history_state
#
###############################################################
def validate_history_state(
    dictHistoryFiles: dict[int, Path],
    dictLatestFiles: dict[int, Path],
) -> tuple[bool, Path | None, int]:
    """履歴状態を検証し、初回判定、比較用パス、次番号を返します。

    自動修復は禁止されているため、不整合があればValueErrorにします。
    """
    if not dictHistoryFiles and not dictLatestFiles:
        return True, None, 1
    if dictHistoryFiles and not dictLatestFiles:
        raise ValueError("履歴TSVが存在しますが、比較用latest TSVがありません。")
    if not dictHistoryFiles and dictLatestFiles:
        raise ValueError("比較用latest TSVが存在しますが、履歴TSVがありません。")
    if len(dictLatestFiles) > 1:
        pszFileNames: str = ", ".join(
            objPath.name for _, objPath in sorted(dictLatestFiles.items())
        )
        raise ValueError("比較用latest TSVが複数存在します。Files = " + pszFileNames)

    iLatestNumber: int = next(iter(dictLatestFiles))
    if iLatestNumber not in dictHistoryFiles:
        raise ValueError(
            "latest TSVと同じ番号の履歴TSVがありません。History number = "
            + f"{iLatestNumber:04d}"
        )
    iMaximumHistoryNumber: int = max(dictHistoryFiles)
    if iLatestNumber != iMaximumHistoryNumber:
        raise ValueError(
            "latest TSVの番号が最新履歴番号と一致しません。latest = "
            + f"{iLatestNumber:04d}"
            + ", history max = "
            + f"{iMaximumHistoryNumber:04d}"
        )
    objHistoryBytes: bytes = dictHistoryFiles[iLatestNumber].read_bytes()
    objLatestBytes: bytes = dictLatestFiles[iLatestNumber].read_bytes()
    if objHistoryBytes != objLatestBytes:
        raise ValueError("同じ番号の履歴TSVとlatest TSVの内容が異なります。")
    if iMaximumHistoryNumber >= MAX_HISTORY_NUMBER:
        raise ValueError("履歴番号が上限9999に達しています。")
    return False, dictLatestFiles[iLatestNumber], iMaximumHistoryNumber + 1


###############################################################
#
# load_previous_tsv
#
###############################################################
def load_previous_tsv(objTsvPath: Path) -> list[OrderRow]:
    """最新latest TSVを読み、4列の前回発注データとして返します。"""
    with objTsvPath.open(mode="r", encoding="utf-8-sig", newline="") as objFile:
        listRows: list[list[str]] = list(csv.reader(objFile, delimiter="\t"))
    if not listRows:
        raise ValueError("比較用TSVが空です。File = " + str(objTsvPath))
    if tuple(cell.strip() for cell in listRows[0]) != EXPECTED_HEADERS:
        raise ValueError("比較用TSVの見出しが正しくありません。File = " + str(objTsvPath))

    listOrderRows: list[OrderRow] = []
    for iRowNumber, listRow in enumerate(listRows[1:], start=2):
        if len(listRow) != 4:
            raise ValueError(
                "比較用TSVが4列ではありません。Row = " + str(iRowNumber)
            )
        if all(pszCell.strip() == "" for pszCell in listRow):
            continue
        pszProductCode: str = listRow[0].strip()
        if pszProductCode == "":
            raise ValueError(
                "比較用TSVのproductCodeが空です。Row = " + str(iRowNumber)
            )
        normalize_quantity_for_comparison(listRow[3])
        listOrderRows.append(
            OrderRow(
                pszProductCode,
                listRow[1],
                listRow[2],
                listRow[3],
            )
        )
    validate_unique_product_codes(listOrderRows, "比較用TSV")
    return listOrderRows


###############################################################
#
# compare_quantity
#
###############################################################
def compare_quantity(
    listCurrentRows: list[OrderRow],
    listPreviousRows: list[OrderRow],
) -> list[Cell]:
    """productCodeで照合し、数量が変わった現在Excelのセルを返します。

    前回に存在しない新規商品は比較できないため、赤色対象にはしません。
    """
    dictPreviousRows: dict[str, OrderRow] = {
        objRow.product_code: objRow for objRow in listPreviousRows
    }
    listChangedCells: list[Cell] = []
    for objCurrentRow in listCurrentRows:
        objPreviousRow: OrderRow | None = dictPreviousRows.get(objCurrentRow.product_code)
        if objPreviousRow is None:
            continue
        objCurrentQuantity = normalize_quantity_for_comparison(objCurrentRow.quantity)
        objPreviousQuantity = normalize_quantity_for_comparison(objPreviousRow.quantity)
        if objCurrentQuantity != objPreviousQuantity:
            if objCurrentRow.quantity_cell is None:
                raise ValueError("変更対象のquantityセルを取得できません。")
            listChangedCells.append(objCurrentRow.quantity_cell)
    return listChangedCells


###############################################################
#
# highlight_changed_quantity
#
###############################################################
def highlight_changed_quantity(listChangedCells: list[Cell]) -> None:
    """既存フォント属性を複製し、変更された数量セルの色だけを赤にします。"""
    for objCell in listChangedCells:
        objFont = copy(objCell.font)
        objFont.color = Color(rgb=RED_FONT_COLOR)
        objCell.font = objFont


###############################################################
#
# serialize_tsv
#
###############################################################
def serialize_tsv(listOrderRows: list[OrderRow]) -> str:
    """現在の4列を、UTF-8保存用のタブ区切り文字列へ変換します。"""
    objStringBuffer = io.StringIO(newline="")
    objWriter = csv.writer(objStringBuffer, delimiter="\t", lineterminator="\n")
    objWriter.writerow(EXPECTED_HEADERS)
    for objOrderRow in listOrderRows:
        objWriter.writerow(
            [
                objOrderRow.product_code,
                objOrderRow.product_name,
                objOrderRow.spec,
                quantity_to_tsv_text(objOrderRow.quantity),
            ]
        )
    return objStringBuffer.getvalue()


###############################################################
#
# build_snapshot_paths
#
###############################################################
def build_snapshot_paths(
    pszInputFileFullPath: str,
    iHistoryNumber: int,
) -> tuple[Path, Path]:
    """指定履歴番号の履歴TSVとlatest TSVの保存先を返します。"""
    objExcelPath: Path = Path(pszInputFileFullPath)
    pszNumber: str = f"{iHistoryNumber:04d}"
    objHistoryPath: Path = objExcelPath.with_name(
        objExcelPath.stem + "_" + pszNumber + ".tsv"
    )
    objLatestPath: Path = objExcelPath.with_name(
        objExcelPath.stem + "_" + pszNumber + "_latest.tsv"
    )
    return objHistoryPath, objLatestPath


###############################################################
#
# save_snapshot_files
#
###############################################################
def save_snapshot_files(
    pszInputFileFullPath: str,
    iHistoryNumber: int,
    pszTsvText: str,
) -> tuple[Path, Path]:
    """同一内容の履歴TSVとlatest TSVを上書き禁止で新規作成します。

    片方の作成に失敗した場合、未完成の今回分だけを片付け、以前の履歴と
    latestは変更しません。
    """
    objHistoryPath, objLatestPath = build_snapshot_paths(
        pszInputFileFullPath,
        iHistoryNumber,
    )
    listCreatedPaths: list[Path] = []
    try:
        with objHistoryPath.open(mode="x", encoding="utf-8", newline="") as objFile:
            objFile.write(pszTsvText)
        listCreatedPaths.append(objHistoryPath)
        with objLatestPath.open(mode="x", encoding="utf-8", newline="") as objFile:
            objFile.write(pszTsvText)
        listCreatedPaths.append(objLatestPath)
        if objHistoryPath.read_bytes() != objLatestPath.read_bytes():
            raise OSError("作成した履歴TSVとlatest TSVの内容が一致しません。")
    except Exception:
        for objCreatedPath in reversed(listCreatedPaths):
            try:
                objCreatedPath.unlink()
            except OSError:
                pass
        raise
    return objHistoryPath, objLatestPath


###############################################################
#
# save_excel
#
###############################################################
def save_excel(objWorkbook: Workbook, pszInputFileFullPath: str) -> None:
    """比較後のWorkbookを入力Excelへ上書き保存します。"""
    objWorkbook.save(pszInputFileFullPath)


###############################################################
#
# remove_previous_latest_tsv
#
###############################################################
def remove_previous_latest_tsv(objPreviousLatestPath: Path) -> None:
    """新しい履歴とlatestの作成成功後、直前のlatestだけを削除します。"""
    objPreviousLatestPath.unlink()


###############################################################
#
# process_excel_file
#
###############################################################
def process_excel_file(pszInputFileFullPath: str) -> None:
    """編集後Excelを前回履歴と比較し、赤色設定と履歴更新を行います。"""
    pszValidatedPath: str = validate_excel_path(pszInputFileFullPath)
    objWorkbook: Workbook = load_excel(pszValidatedPath)
    objWorksheet: Worksheet = find_target_worksheet(objWorkbook)
    listCurrentRows: list[OrderRow] = read_excel_rows(objWorksheet)
    dictHistoryFiles, dictLatestFiles = find_history_files(pszValidatedPath)
    bIsFirstRun, objPreviousLatestPath, iNextHistoryNumber = validate_history_state(
        dictHistoryFiles,
        dictLatestFiles,
    )
    pszTsvText: str = serialize_tsv(listCurrentRows)

    if bIsFirstRun:
        raise ValueError(
            "比較用履歴がありません。Excelを編集する前に--prepareを実行してください。"
        )

    if objPreviousLatestPath is None:
        raise ValueError("比較用latest TSVを特定できません。")
    listPreviousRows: list[OrderRow] = load_previous_tsv(objPreviousLatestPath)
    listChangedCells: list[Cell] = compare_quantity(listCurrentRows, listPreviousRows)
    highlight_changed_quantity(listChangedCells)
    save_excel(objWorkbook, pszValidatedPath)

    objHistoryPath, objLatestPath = save_snapshot_files(
        pszValidatedPath,
        iNextHistoryNumber,
        pszTsvText,
    )
    remove_previous_latest_tsv(objPreviousLatestPath)
    remove_old_error_file(pszValidatedPath)
    print("Excelと履歴TSVを更新しました。")
    print("Changed quantity cells: " + str(len(listChangedCells)))
    print("Excel: " + pszValidatedPath)
    print("History: " + str(objHistoryPath))
    print("Latest: " + str(objLatestPath))


###############################################################
#
# prepare_history_before_edit
#
###############################################################
def prepare_history_before_edit(pszInputFileFullPath: str) -> None:
    """Excelを開く前に履歴を検証し、初回だけ編集前履歴を作成します。

    Args:
        pszInputFileFullPath: これから担当者が編集するExcelのパスです。

    Returns:
        戻り値はありません。履歴がない場合は0001の履歴とlatestを作り、
        既に履歴がある場合は整合性検証だけを行います。

    Raises:
        ValueError: Excelまたは履歴の状態が仕様に合わない場合です。
        OSError: ExcelやTSVの読み書きに失敗した場合です。
    """
    pszValidatedPath: str = validate_excel_path(pszInputFileFullPath)
    objWorkbook: Workbook = load_excel(pszValidatedPath)
    objWorksheet: Worksheet = find_target_worksheet(objWorkbook)
    listCurrentRows: list[OrderRow] = read_excel_rows(objWorksheet)
    dictHistoryFiles, dictLatestFiles = find_history_files(pszValidatedPath)
    bIsFirstRun, _, iNextHistoryNumber = validate_history_state(
        dictHistoryFiles,
        dictLatestFiles,
    )

    if bIsFirstRun:
        pszTsvText: str = serialize_tsv(listCurrentRows)
        objHistoryPath, objLatestPath = save_snapshot_files(
            pszValidatedPath,
            iNextHistoryNumber,
            pszTsvText,
        )
        remove_old_error_file(pszValidatedPath)
        print("編集前の初回履歴を作成しました。")
        print("History: " + str(objHistoryPath))
        print("Latest: " + str(objLatestPath))
        return

    remove_old_error_file(pszValidatedPath)
    print("編集前の履歴確認が完了しました。")


###############################################################
#
# main
#
###############################################################
def main() -> int:
    """引数を確認して処理を実行し、成功0・失敗1の終了コードを返します。"""
    iArgumentCount: int = len(sys.argv)
    bIsPrepareMode: bool = iArgumentCount == 3 and sys.argv[1] == "--prepare"
    bIsCompareMode: bool = iArgumentCount == 2
    if not bIsPrepareMode and not bIsCompareMode:
        pszScriptFileName: str = os.path.basename(__file__)
        pszErrorMessage: str = (
            "Error: 実行モードとExcelファイルパスを正しく指定してください。\n"
            + "Usage: python "
            + pszScriptFileName
            + " <excel_file_path>\n"
            + "Prepare: python "
            + pszScriptFileName
            + " --prepare <excel_file_path>\n"
        )
        print(pszErrorMessage, file=sys.stderr, end="")
        pszErrorFileFullPath: str = os.path.splitext(pszScriptFileName)[0] + "_error_argument.txt"
        try:
            write_error_text(pszErrorFileFullPath, pszErrorMessage)
        except OSError as objException:
            print("Error: 引数エラーファイルを保存できません。Detail = " + str(objException), file=sys.stderr)
        return 1

    pszInputFileFullPath: str = sys.argv[2] if bIsPrepareMode else sys.argv[1]
    try:
        if bIsPrepareMode:
            prepare_history_before_edit(pszInputFileFullPath)
        else:
            process_excel_file(pszInputFileFullPath)
    except Exception as objException:
        report_processing_error(
            pszInputFileFullPath,
            "発注数量変更の強調処理",
            str(objException),
        )
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
