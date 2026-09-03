# -- coding: utf-8 --
###############################################################
#
# AsahiSingleOrderProductCodeSelector_Cmd.py
#
# pip install openpyxl
#
###############################################################

from __future__ import annotations

import csv
import difflib
import os
import re
import sys
import tempfile
import tkinter as tk
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import messagebox, ttk

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


STEP_HEADERS: tuple[str, ...] = (
    "納品日",
    "曜日",
    "配送パターン",
    "便",
    "Ｐ品番",
    "APEX品番",
    "商品名",
    "産地",
    "仕様",
    "伝票原価",
    "伝票売価",
    "値入",
    "売価",
    "単位",
)
WEEKDAYS: tuple[str, ...] = ("月", "火", "水", "木", "金", "土", "日")
SUPPORTED_EXTENSIONS: set[str] = {".xlsx", ".tsv"}
STEP0007_MARKER: str = "_step0007_"
OUTPUT_PREFIX: str = "ProductCodeSelector_step0001_"
SOURCE_PRODUCTS_FILE_NAME: str = "products_all_109_readable.tsv"
PRODUCTS_FILE_NAME: str = "products_all_109_readable_ABC.tsv"
PRODUCT_HEADERS: tuple[str, str, str] = ("productCode", "productName", "spec")
COLUMN_WIDTH_LIMITS: tuple[tuple[int, int], ...] = (
    (12, 14),
    (6, 8),
    (14, 24),
    (8, 12),
    (14, 22),
    (14, 22),
    (24, 50),
    (14, 30),
    (14, 30),
    (14, 18),
    (14, 18),
    (10, 16),
    (12, 18),
    (10, 18),
)
STORE_COLUMN_WIDTH_LIMITS: tuple[int, int] = (10, 24)
PRODUCT_CATEGORY_TERMS: dict[str, tuple[str, ...]] = {
    "まぐろ": ("まぐろ", "鮪", "本まぐろ", "本鮪", "クロマグロ", "黒まぐろ", "キハダ", "メバチ", "ビンナガ", "ビンチョウ", "トンボ", "ミナミマグロ", "南まぐろ", "インドまぐろ"),
    "かき": ("かき", "牡蠣", "真牡蠣", "マガキ"),
    "いか": ("いか", "烏賊", "紋甲いか", "モンゴウイカ", "あおりいか", "するめいか", "やりいか"),
    "ぶり": ("ぶり", "鰤", "はまち", "ワラサ", "イナダ"),
    "たい": ("たい", "鯛", "真鯛", "マダイ"),
    "えび": ("えび", "海老", "蝦"),
    "かに": ("かに", "蟹", "ずわいがに", "ずわい蟹"),
    "たら": ("たら", "鱈", "真たら", "真鱈", "マダラ"),
}
PRODUCT_ATTRIBUTE_TERMS: dict[str, tuple[str, ...]] = {
    "冷凍": ("冷凍",), "生": ("生",), "ボイル": ("ボイル", "ゆで", "茹で"),
    "蒸し": ("蒸し",), "塩": ("塩",), "熟成": ("熟成",),
    "ロイン": ("ロイン",), "フィレ": ("フィレ", "フィーレ", "フィレット"),
    "切り落とし": ("切り落とし", "切落し"), "カマ": ("カマ",),
    "ほほ肉": ("ほほ肉", "頬肉"), "頭肉": ("頭肉",), "あら": ("あら",),
    "ラウンド": ("ラウンド",), "スキンレス": ("スキンレス", "皮なし", "皮無し"),
    "骨取り": ("骨取り", "骨とり"), "刺身用": ("刺身用", "生食用"),
    "加熱用": ("加熱用",), "加工品用": ("加工品用",), "原料": ("原料",),
    "養殖": ("養殖",), "天然": ("天然",), "MEL認証": ("MEL認証",),
    "代用品": ("代用品",),
}
PRODUCT_ORIGIN_TERMS: dict[str, tuple[str, ...]] = {
    "北海道": ("北海道",), "千葉": ("千葉県", "千葉県産", "千葉産"),
    "鳥取": ("鳥取県", "鳥取県産", "鳥取産"),
    "岡山": ("岡山県", "岡山県産", "岡山産"),
    "広島": ("広島県", "広島県産", "広島産"),
    "徳島": ("徳島県", "徳島県産", "徳島産"),
    "愛媛": ("愛媛県", "愛媛県産", "愛媛産"),
    "高知": ("高知県", "高知県産", "高知産"),
    "長崎": ("長崎県", "長崎県産", "長崎産"),
    "鹿児島": ("鹿児島県", "鹿児島県産", "鹿児島産"),
    "和歌山": ("和歌山県", "和歌山県産", "和歌山産"),
    "瀬戸内": ("瀬戸内", "瀬戸内産"),
    "国産": ("国産", "国内産"), "カナダ": ("カナダ", "カナダ産"),
}
PRODUCT_BRAND_TERMS: dict[str, tuple[str, ...]] = {
    "日の出": ("日の出", "日の出まぐろ"),
    "地元めし": ("地元めし", "瀬戸内地元めし"),
}
WEAK_PRODUCT_ATTRIBUTE_GROUPS: frozenset[str] = frozenset(
    {"冷凍", "生", "原料", "養殖", "天然"}
)
UNKNOWN_CATEGORY_SIMILARITY_THRESHOLD: float = 0.60


class SelectionCancelledError(Exception):
    """商品選択がユーザー操作によってキャンセルされたことを表します。"""


class NoMatchingProductError(Exception):
    """担当者が商品マスターに該当商品なしと判断したことを表します。"""


class ProductCandidate:
    """商品マスターの1商品を保持します。"""

    def __init__(self, pszCode: str, pszName: str, pszSpec: str) -> None:
        self.code: str = pszCode
        self.name: str = pszName
        self.spec: str = pszSpec

    @property
    def display_text(self) -> str:
        """プルダウンに表示する文字列を返します。"""
        pszText: str = self.code + " - " + self.name
        if self.spec:
            pszText += " - " + self.spec
        return pszText


def configure_standard_streams() -> None:
    """標準出力と標準エラーをUTF-8へ統一します。"""
    for objStream in (sys.stdout, sys.stderr):
        objReconfigure = getattr(objStream, "reconfigure", None)
        if callable(objReconfigure):
            objReconfigure(encoding="utf-8", errors="replace")


def normalize_cell(objValue: object, iColumn: int) -> str:
    """XLSXとTSVを同じ論理値として扱える文字列へ変換します。"""
    if objValue is None:
        return ""
    if iColumn == 0 and isinstance(objValue, datetime):
        return objValue.date().strftime("%Y/%m/%d")
    if iColumn == 0 and isinstance(objValue, date):
        return objValue.strftime("%Y/%m/%d")
    if isinstance(objValue, bool):
        return "TRUE" if objValue else "FALSE"
    if isinstance(objValue, float) and objValue.is_integer():
        return str(int(objValue))
    return str(objValue)


def validate_input_path(pszInputFileFullPath: str) -> Path:
    """商品別のXLSXまたはTSV入力パスを検証します。"""
    objInputPath: Path = Path(pszInputFileFullPath).expanduser().resolve()
    if not objInputPath.is_file():
        raise ValueError("入力ファイルが見つかりません。Path = " + str(objInputPath))
    if objInputPath.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(
            "入力形式はXLSXまたはTSVではありません。Path = " + str(objInputPath)
        )
    if objInputPath.stem.count(STEP0007_MARKER) > 1:
        raise ValueError(
            "入力ファイル名に_step0007_を複数含めることはできません。Path = "
            + str(objInputPath)
        )
    return objInputPath


def read_excel_table(objInputPath: Path) -> tuple[list[list[str]], str]:
    """商品別step0007 XLSXの全セルとシート名を読み込みます。"""
    objWorkbook: Workbook = load_workbook(objInputPath, data_only=True)
    if len(objWorkbook.worksheets) != 1:
        raise ValueError("入力XLSXのワークシート数が1ではありません。")
    objWorksheet: Worksheet = objWorkbook.active
    listRows: list[list[str]] = [
        [
            normalize_cell(objWorksheet.cell(iRow, iColumn).value, iColumn - 1)
            for iColumn in range(1, objWorksheet.max_column + 1)
        ]
        for iRow in range(1, objWorksheet.max_row + 1)
    ]
    return listRows, objWorksheet.title


def read_tsv_table(objInputPath: Path) -> tuple[list[list[str]], str]:
    """商品別step0007 TSVをUTF-8として読み込みます。"""
    with objInputPath.open(mode="r", encoding="utf-8-sig", newline="") as objFile:
        listRawRows: list[list[str]] = list(
            csv.reader(objFile, delimiter="\t", strict=True)
        )
    listRows: list[list[str]] = [
        [normalize_cell(pszValue, iColumn) for iColumn, pszValue in enumerate(listRow)]
        for listRow in listRawRows
    ]
    return listRows, "ProductCodeSelector_step0001"


def validate_step0007_table(listRows: list[list[str]]) -> None:
    """商品別step0007の2行ヘッダーと月～日7行を検証します。"""
    if len(listRows) != 2 + len(WEEKDAYS):
        raise ValueError(
            "商品別step0007はヘッダー2行とデータ7行の合計9行ではありません。"
            + " 行数 = "
            + str(len(listRows))
        )
    iColumnCount: int = len(listRows[0])
    if iColumnCount < len(STEP_HEADERS):
        raise ValueError("商品別step0007の列数が14列未満です。")
    for iRow, listRow in enumerate(listRows, start=1):
        if len(listRow) != iColumnCount:
            raise ValueError(f"商品別step0007の{iRow}行目の列数が一致しません。")
    if any(pszValue.strip() for pszValue in listRows[0][: len(STEP_HEADERS)]):
        raise ValueError("商品別step0007の1行目A～N列が空欄ではありません。")
    if tuple(pszValue.strip() for pszValue in listRows[1][: len(STEP_HEADERS)]) != STEP_HEADERS:
        raise ValueError("商品別step0007の2行目A～N列が仕様どおりではありません。")
    listStoreCodes: list[str] = [
        pszValue.strip() for pszValue in listRows[0][len(STEP_HEADERS) :]
    ]
    if any(not pszCode for pszCode in listStoreCodes):
        raise ValueError("商品別step0007の店舗コードに空欄があります。")
    if len(set(listStoreCodes)) != len(listStoreCodes):
        raise ValueError("商品別step0007の店舗コードが重複しています。")

    listDataRows: list[list[str]] = listRows[2:]
    # 商品基本情報は月曜日行だけに入り、火～日のC～N列は空欄またはメモです。
    if not listDataRows[0][6].strip():
        raise ValueError("商品別step0007の商品名が空欄です。")
    for iDay, (listRow, pszExpectedWeekday) in enumerate(
        zip(listDataRows, WEEKDAYS)
    ):
        iFileRow: int = iDay + 3
        if listRow[1].strip() != pszExpectedWeekday:
            raise ValueError(
                f"商品別step0007の{iFileRow}行目の曜日が{pszExpectedWeekday}ではありません。"
            )
    try:
        objMonday: date = datetime.strptime(listDataRows[0][0], "%Y/%m/%d").date()
    except ValueError as objException:
        raise ValueError("商品別step0007の月曜日日付が不正です。") from objException
    if objMonday.weekday() != 0:
        raise ValueError("商品別step0007の開始日が月曜日ではありません。")
    for iDay, listRow in enumerate(listDataRows):
        try:
            objActualDate: date = datetime.strptime(listRow[0], "%Y/%m/%d").date()
        except ValueError as objException:
            raise ValueError(
                f"商品別step0007の{iDay + 3}行目の日付が不正です。"
            ) from objException
        if objActualDate != objMonday + timedelta(days=iDay):
            raise ValueError("商品別step0007の日付が月～日の連続日付ではありません。")
    if not any(
        pszValue.strip()
        for listRow in listDataRows
        for pszValue in listRow[len(STEP_HEADERS) :]
    ):
        raise ValueError("商品別step0007に発注数量がありません。")


def clear_product_codes(listRows: list[list[str]]) -> list[list[str]]:
    """データ7行のＰ品番とAPEX品番を空欄にしたコピーを返します。"""
    listOutputRows: list[list[str]] = [listRow.copy() for listRow in listRows]
    for listRow in listOutputRows[2:]:
        listRow[4] = ""
        listRow[5] = ""
    return listOutputRows


def sanitize_filename_part(pszValue: str) -> str:
    """Windowsで安全なファイル名部分へ変換します。"""
    pszSafeValue: str = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", pszValue.strip())
    pszSafeValue = re.sub(r"_+", "_", pszSafeValue).rstrip(" .")
    if not pszSafeValue:
        raise ValueError("出力ファイル名を作成できません。")
    return pszSafeValue


def get_output_paths(objInputPath: Path, pszProductName: str) -> tuple[Path, Path]:
    """入力名からProductCodeSelector step0001の出力パスを作ります。"""
    iMarkerCount: int = objInputPath.stem.count(STEP0007_MARKER)
    if iMarkerCount == 0:
        pszOutputStem: str = OUTPUT_PREFIX + sanitize_filename_part(objInputPath.stem)
        return (
            objInputPath.with_name(pszOutputStem + ".xlsx"),
            objInputPath.with_name(pszOutputStem + ".tsv"),
        )
    if iMarkerCount > 1:
        raise ValueError(
            "入力ファイル名に_step0007_を複数含めることはできません。Path = "
            + str(objInputPath)
        )
    _, pszIdentity = objInputPath.stem.split(STEP0007_MARKER, 1)
    if "_" not in pszIdentity:
        raise ValueError("step0007のファイル名に商品コード以降の情報がありません。")
    _, pszAfterCode = pszIdentity.split("_", 1)
    pszSafeProductName: str = sanitize_filename_part(pszProductName)
    pszExpectedProductPrefix: str = pszSafeProductName + "_"
    if not pszAfterCode.startswith(pszExpectedProductPrefix):
        raise ValueError(
            "step0007のファイル名の商品名が入力表の商品名と一致しません。"
        )
    pszSuffix: str = pszAfterCode[len(pszExpectedProductPrefix) :]
    if not pszSuffix:
        raise ValueError("step0007のファイル名に配送センター情報がありません。")
    pszOutputStem = OUTPUT_PREFIX + sanitize_filename_part(
        pszSafeProductName + "_" + pszSuffix
    )
    return (
        objInputPath.with_name(pszOutputStem + ".xlsx"),
        objInputPath.with_name(pszOutputStem + ".tsv"),
    )


def calculate_display_width(objValue: object, iColumn: int) -> int:
    """全角文字を2、半角文字を1としてセルの最大行表示幅を返します。"""
    if objValue is None:
        return 0
    if iColumn == 1 and isinstance(objValue, datetime):
        pszValue: str = objValue.date().strftime("%Y/%m/%d")
    elif iColumn == 1 and isinstance(objValue, date):
        pszValue = objValue.strftime("%Y/%m/%d")
    else:
        pszValue = str(objValue)
    pszValue = pszValue.replace("\t", "    ")
    return max(
        (
            sum(
                2 if unicodedata.east_asian_width(pszCharacter) in ("W", "F") else 1
                for pszCharacter in pszLine
            )
            for pszLine in pszValue.splitlines() or [""]
        ),
        default=0,
    )


def adjust_excel_column_widths(objWorksheet: Worksheet) -> None:
    """セル内容と列別の最小・最大幅に基づいて全列幅を設定します。"""
    for iColumn in range(1, objWorksheet.max_column + 1):
        if iColumn <= len(COLUMN_WIDTH_LIMITS):
            iMinimumWidth, iMaximumWidth = COLUMN_WIDTH_LIMITS[iColumn - 1]
        else:
            iMinimumWidth, iMaximumWidth = STORE_COLUMN_WIDTH_LIMITS
        iContentWidth: int = max(
            calculate_display_width(
                objWorksheet.cell(iRow, iColumn).value, iColumn
            )
            for iRow in range(1, objWorksheet.max_row + 1)
        )
        iColumnWidth: int = min(iMaximumWidth, max(iMinimumWidth, iContentWidth + 2))
        objWorksheet.column_dimensions[get_column_letter(iColumn)].width = iColumnWidth


def save_excel_table(
    objOutputPath: Path, listRows: list[list[str]], pszWorksheetTitle: str
) -> None:
    """ProductCodeSelector step0001 XLSXを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.title = pszWorksheetTitle
    for iRow, listRow in enumerate(listRows, start=1):
        listValues: list[object] = listRow.copy()
        if iRow >= 3 and listValues[0]:
            listValues[0] = datetime.strptime(str(listValues[0]), "%Y/%m/%d").date()
        for iColumn in range(len(STEP_HEADERS), len(listValues)):
            if iRow >= 3 and str(listValues[iColumn]).isdigit():
                listValues[iColumn] = int(str(listValues[iColumn]))
        objWorksheet.append(listValues)
        if iRow == 1:
            for iColumn in range(len(STEP_HEADERS) + 1, len(listValues) + 1):
                objWorksheet.cell(iRow, iColumn).number_format = "@"
        elif iRow >= 3:
            objWorksheet.cell(iRow, 1).number_format = "yyyy/mm/dd"
            objWorksheet.cell(iRow, 5).number_format = "@"
            objWorksheet.cell(iRow, 6).number_format = "@"
    adjust_excel_column_widths(objWorksheet)
    objWorkbook.save(objOutputPath)


def save_tsv_table(objOutputPath: Path, listRows: list[list[str]]) -> None:
    """ProductCodeSelector step0001 TSVを保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        csv.writer(objFile, delimiter="\t", lineterminator="\r\n").writerows(listRows)


def validate_outputs_match(objExcelPath: Path, objTsvPath: Path) -> None:
    """保存したXLSXとTSVの全セルが一致し、商品コードが空欄か確認します。"""
    listExcelRows, _ = read_excel_table(objExcelPath)
    listTsvRows, _ = read_tsv_table(objTsvPath)
    if listExcelRows != listTsvRows:
        raise ValueError("step0001のXLSXとTSVの内容が一致しません。")
    if any(listRow[4].strip() or listRow[5].strip() for listRow in listExcelRows[2:]):
        raise ValueError("step0001のＰ品番またはAPEX品番が空欄ではありません。")


def get_source_products_file_path() -> Path:
    """Cmdプログラムと同じフォルダーの元商品マスターパスを返します。"""
    return Path(__file__).resolve().parent / SOURCE_PRODUCTS_FILE_NAME


def get_products_file_path() -> Path:
    """Cmdプログラムと同じフォルダーの3列商品マスターパスを返します。"""
    return Path(__file__).resolve().parent / PRODUCTS_FILE_NAME


def create_abc_product_master(
    objSourcePath: Path, objOutputPath: Path
) -> None:
    """元商品マスターのA～C列だけを抽出したTSVを安全に作成します。"""
    if not objSourcePath.is_file():
        raise ValueError("元商品マスターが見つかりません。Path = " + str(objSourcePath))
    with objSourcePath.open(mode="r", encoding="utf-8-sig", newline="") as objFile:
        listSourceRows: list[list[str]] = list(
            csv.reader(objFile, delimiter="\t", strict=True)
        )
    listSourceRows = [
        listRow
        for listRow in listSourceRows
        if any(pszValue.strip() for pszValue in listRow)
    ]
    if not listSourceRows:
        raise ValueError("元商品マスターが空です。")
    if len(listSourceRows[0]) < len(PRODUCT_HEADERS):
        raise ValueError("元商品マスターの列数が3列未満です。")
    if tuple(pszValue.strip() for pszValue in listSourceRows[0][:3]) != PRODUCT_HEADERS:
        raise ValueError(
            "元商品マスターの先頭3列はproductCode、productName、specではありません。"
        )
    listOutputRows: list[list[str]] = [list(PRODUCT_HEADERS)]
    for iRow, listRow in enumerate(listSourceRows[1:], start=2):
        if len(listRow) < len(PRODUCT_HEADERS):
            raise ValueError(f"元商品マスターの{iRow}行目が3列未満です。")
        pszCode, pszName, pszSpec = (pszValue.strip() for pszValue in listRow[:3])
        if not pszCode or not pszName:
            raise ValueError(
                f"元商品マスターの{iRow}行目の商品コードまたは商品名が空欄です。"
            )
        listOutputRows.append([pszCode, pszName, pszSpec])

    objTemporaryPath: Path = create_temporary_path(objOutputPath)
    try:
        save_tsv_table(objTemporaryPath, listOutputRows)
        with objTemporaryPath.open(
            mode="r", encoding="utf-8-sig", newline=""
        ) as objFile:
            listSavedRows: list[list[str]] = list(
                csv.reader(objFile, delimiter="\t", strict=True)
            )
        if listSavedRows != listOutputRows:
            raise ValueError(
                "products_all_109_readable_ABC.tsvの保存内容が元商品マスターのA～C列と一致しません。"
            )
        read_product_candidates(objTemporaryPath)
        os.replace(objTemporaryPath, objOutputPath)
    finally:
        if objTemporaryPath.exists():
            objTemporaryPath.unlink()


def read_product_candidates(objProductsPath: Path) -> list[ProductCandidate]:
    """商品マスターを読み込み、商品コードの矛盾を検証します。"""
    if not objProductsPath.is_file():
        raise ValueError("商品マスターが見つかりません。Path = " + str(objProductsPath))
    with objProductsPath.open(mode="r", encoding="utf-8-sig", newline="") as objFile:
        listRows: list[list[str]] = list(csv.reader(objFile, delimiter="\t", strict=True))
    listRows = [listRow for listRow in listRows if any(pszValue.strip() for pszValue in listRow)]
    if not listRows:
        raise ValueError("商品マスターが空です。")
    if tuple(pszValue.strip() for pszValue in listRows[0]) != PRODUCT_HEADERS:
        raise ValueError("商品マスターのヘッダーがproductCode、productName、specではありません。")
    listCandidates: list[ProductCandidate] = []
    dictCodes: dict[str, tuple[str, str]] = {}
    for iRow, listRow in enumerate(listRows[1:], start=2):
        if len(listRow) != 3:
            raise ValueError(f"商品マスターの{iRow}行目が3列ではありません。")
        pszCode, pszName, pszSpec = (pszValue.strip() for pszValue in listRow)
        if not pszCode or not pszName:
            raise ValueError(f"商品マスターの{iRow}行目の商品コードまたは商品名が空欄です。")
        tupleDefinition: tuple[str, str] = (pszName, pszSpec)
        if pszCode in dictCodes:
            if dictCodes[pszCode] != tupleDefinition:
                raise ValueError("商品マスターの商品コード定義が矛盾しています。Code = " + pszCode)
            continue
        dictCodes[pszCode] = tupleDefinition
        listCandidates.append(ProductCandidate(pszCode, pszName, pszSpec))
    if not listCandidates:
        raise ValueError("商品マスターに商品がありません。")
    return listCandidates


def normalize_product_name(pszValue: str) -> str:
    """商品名を候補検索用に正規化します。"""
    pszNormalized: str = unicodedata.normalize("NFKC", pszValue).strip().casefold()
    listCharacters: list[str] = []
    for pszCharacter in pszNormalized:
        iCodePoint: int = ord(pszCharacter)
        if 0x30A1 <= iCodePoint <= 0x30F6:
            listCharacters.append(chr(iCodePoint - 0x60))
        else:
            listCharacters.append(pszCharacter)
    return " ".join("".join(listCharacters).replace("\u3000", " ").split())


def compact_product_text(pszValue: str) -> str:
    """空白と記号を除いた候補比較用文字列を返します。"""
    return "".join(
        pszCharacter
        for pszCharacter in normalize_product_name(pszValue)
        if pszCharacter.isalnum()
    )


def extract_term_groups(
    pszValue: str, dictTermGroups: dict[str, tuple[str, ...]]
) -> set[str]:
    """文字列に含まれるカテゴリまたは属性のグループ名を返します。"""
    pszCompactValue: str = compact_product_text(pszValue)
    return {
        pszGroup
        for pszGroup, tupleTerms in dictTermGroups.items()
        if any(compact_product_text(pszTerm) in pszCompactValue for pszTerm in tupleTerms)
    }


def evaluate_product_candidate(
    objCandidate: ProductCandidate, pszProductName: str, pszInputSpec: str
) -> tuple[int, list[str], list[str]]:
    """候補の関連度、一致理由、主な相違点を返します。"""
    pszInputNormalized: str = normalize_product_name(pszProductName)
    pszCandidateNormalized: str = normalize_product_name(objCandidate.name)
    pszInputCompact: str = compact_product_text(pszProductName)
    pszCandidateCompact: str = compact_product_text(objCandidate.name)
    iScore: int = 0
    listReasons: list[str] = []
    listDifferences: list[str] = []
    if objCandidate.name.strip() == pszProductName.strip():
        iScore += 1000
        listReasons.append("完全一致")
    elif pszCandidateNormalized == pszInputNormalized:
        iScore += 900
        listReasons.append("正規化一致")
    elif pszCandidateCompact == pszInputCompact:
        iScore += 800
        listReasons.append("空白・記号除去一致")
    elif pszInputCompact in pszCandidateCompact or pszCandidateCompact in pszInputCompact:
        iScore += 400
        listReasons.append("名称包含")

    setInputCategories: set[str] = extract_term_groups(
        pszProductName, PRODUCT_CATEGORY_TERMS
    )
    setCandidateCategories: set[str] = extract_term_groups(
        objCandidate.name, PRODUCT_CATEGORY_TERMS
    )
    setSharedCategories: set[str] = setInputCategories & setCandidateCategories
    if setSharedCategories:
        iScore += 300 * len(setSharedCategories)
        listReasons.append("魚介カテゴリ一致:" + ",".join(sorted(setSharedCategories)))

    setInputAttributes: set[str] = extract_term_groups(
        pszProductName, PRODUCT_ATTRIBUTE_TERMS
    )
    setCandidateAttributes: set[str] = extract_term_groups(
        objCandidate.name, PRODUCT_ATTRIBUTE_TERMS
    )
    setSharedAttributes: set[str] = setInputAttributes & setCandidateAttributes
    if setSharedAttributes:
        iScore += 100 * len(setSharedAttributes)
        listReasons.append("属性一致:" + ",".join(sorted(setSharedAttributes)))
    for pszDifference in sorted(setInputAttributes ^ setCandidateAttributes):
        listDifferences.append("属性差:" + pszDifference)

    for pszLabel, dictTerms, iPoints in (
        ("産地", PRODUCT_ORIGIN_TERMS, 100),
        ("ブランド", PRODUCT_BRAND_TERMS, 80),
    ):
        setInputTerms: set[str] = extract_term_groups(pszProductName, dictTerms)
        setCandidateTerms: set[str] = extract_term_groups(objCandidate.name, dictTerms)
        setSharedTerms: set[str] = setInputTerms & setCandidateTerms
        if setSharedTerms:
            iScore += iPoints * len(setSharedTerms)
            listReasons.append(pszLabel + "一致:" + ",".join(sorted(setSharedTerms)))
        if setInputTerms and setCandidateTerms and not setSharedTerms:
            listDifferences.append(pszLabel + "差")

    pszInputSpecNormalized: str = compact_product_text(pszInputSpec)
    pszCandidateSpecNormalized: str = compact_product_text(objCandidate.spec)
    if pszInputSpecNormalized and pszCandidateSpecNormalized:
        if pszInputSpecNormalized == pszCandidateSpecNormalized:
            iScore += 150
            listReasons.append("仕様一致")
        else:
            listDifferences.append("仕様差")
    fSimilarity: float = difflib.SequenceMatcher(
        None, pszInputCompact, pszCandidateCompact
    ).ratio()
    iScore += round(fSimilarity * 200)
    if fSimilarity >= 0.6 and not any(
        pszReason in listReasons
        for pszReason in ("完全一致", "正規化一致", "空白・記号除去一致")
    ):
        listReasons.append("名称類似")
    return iScore, listReasons, listDifferences


def find_product_candidates(
    listCandidates: list[ProductCandidate], pszProductName: str, pszInputSpec: str = ""
) -> list[ProductCandidate]:
    """完全一致で終了せず、関連する可能性がある商品を関連度順に返します。"""
    listEvaluated: list[tuple[int, ProductCandidate]] = []
    setInputCategories: set[str] = extract_term_groups(
        pszProductName, PRODUCT_CATEGORY_TERMS
    )
    pszInputCompact: str = compact_product_text(pszProductName)
    for objCandidate in listCandidates:
        iScore, listReasons, _ = evaluate_product_candidate(
            objCandidate, pszProductName, pszInputSpec
        )
        setCandidateCategories: set[str] = extract_term_groups(
            objCandidate.name, PRODUCT_CATEGORY_TERMS
        )
        fSimilarity: float = difflib.SequenceMatcher(
            None, pszInputCompact, compact_product_text(objCandidate.name)
        ).ratio()
        if setInputCategories:
            bIsRelated: bool = bool(setInputCategories & setCandidateCategories)
        else:
            pszCandidateNormalized: str = normalize_product_name(objCandidate.name)
            pszCandidateCompact: str = compact_product_text(objCandidate.name)
            bStrongNameMatch: bool = (
                objCandidate.name.strip() == pszProductName.strip()
                or pszCandidateNormalized == normalize_product_name(pszProductName)
                or pszCandidateCompact == pszInputCompact
                or pszInputCompact in pszCandidateCompact
                or pszCandidateCompact in pszInputCompact
            )
            setSharedStrongAttributes: set[str] = (
                extract_term_groups(pszProductName, PRODUCT_ATTRIBUTE_TERMS)
                & extract_term_groups(objCandidate.name, PRODUCT_ATTRIBUTE_TERMS)
            ) - WEAK_PRODUCT_ATTRIBUTE_GROUPS
            bSharedOriginOrBrand: bool = any(
                extract_term_groups(pszProductName, dictTerms)
                & extract_term_groups(objCandidate.name, dictTerms)
                for dictTerms in (PRODUCT_ORIGIN_TERMS, PRODUCT_BRAND_TERMS)
            )
            pszInputSpecNormalized: str = compact_product_text(pszInputSpec)
            pszCandidateSpecNormalized: str = compact_product_text(objCandidate.spec)
            bMatchingSpec: bool = bool(
                pszInputSpecNormalized
                and pszCandidateSpecNormalized
                and pszInputSpecNormalized == pszCandidateSpecNormalized
            )
            bIsRelated = bool(
                bStrongNameMatch
                or fSimilarity >= UNKNOWN_CATEGORY_SIMILARITY_THRESHOLD
                or setSharedStrongAttributes
                or bSharedOriginOrBrand
                or bMatchingSpec
            )
        if bIsRelated:
            listEvaluated.append((iScore, objCandidate))
    listEvaluated.sort(key=lambda tupleItem: (-tupleItem[0], tupleItem[1].code))
    return [objCandidate for _, objCandidate in listEvaluated]


def select_product_candidate(
    pszProductName: str,
    pszInputSpec: str,
    listRelatedCandidates: list[ProductCandidate],
    listAllCandidates: list[ProductCandidate],
    pszSundayValue: str,
) -> ProductCandidate:
    """関連候補・全商品を検索できる画面から担当者が選択します。"""
    objRoot = tk.Tk()
    objRoot.title("Asahi Single Order Product Code Selector step0002")
    objRoot.resizable(True, False)
    objSelectedCandidate: ProductCandidate | None = None
    bNoMatchingProduct: bool = False
    listVisibleCandidates: list[ProductCandidate] = []
    objFrame = ttk.Frame(objRoot, padding=12)
    objFrame.grid(row=0, column=0, sticky="nsew")
    ttk.Label(objFrame, text="入力商品名:").grid(row=0, column=0, sticky="w")
    ttk.Label(objFrame, text=pszProductName).grid(row=1, column=0, columnspan=3, sticky="w")
    ttk.Label(objFrame, text="入力仕様:").grid(row=2, column=0, sticky="w")
    ttk.Label(objFrame, text=pszInputSpec or "(空欄)").grid(
        row=3, column=0, columnspan=3, sticky="w", pady=(0, 8)
    )
    ttk.Label(objFrame, text="検索:").grid(row=4, column=0, sticky="w")
    objSearch = tk.StringVar()
    objSearchEntry = ttk.Entry(objFrame, textvariable=objSearch, width=45)
    objSearchEntry.grid(row=4, column=1, columnspan=2, sticky="ew")
    ttk.Label(objFrame, text="表示対象:").grid(row=5, column=0, sticky="w")
    objMode = tk.StringVar(value="推奨・関連候補")
    objModeComboBox = ttk.Combobox(
        objFrame,
        textvariable=objMode,
        values=("推奨・関連候補", "同じ魚介カテゴリ", "全商品"),
        state="readonly",
        width=20,
    )
    objModeComboBox.grid(row=5, column=1, sticky="w")
    objCount = tk.StringVar()
    ttk.Label(objFrame, textvariable=objCount).grid(row=5, column=2, sticky="e")
    ttk.Label(objFrame, text="商品候補:").grid(row=6, column=0, sticky="w")
    objSelection = tk.StringVar()
    objComboBox = ttk.Combobox(
        objFrame, textvariable=objSelection, state="readonly", width=90,
    )
    objComboBox.grid(row=7, column=0, columnspan=3, sticky="ew")
    objDetails = tk.StringVar()
    ttk.Label(objFrame, textvariable=objDetails, wraplength=760).grid(
        row=8, column=0, columnspan=3, sticky="w", pady=(6, 12)
    )

    def update_candidate_details(*_objArguments: object) -> None:
        iSelectedIndex: int = objComboBox.current()
        if iSelectedIndex < 0 or iSelectedIndex >= len(listVisibleCandidates):
            objDetails.set("")
            return
        _, listReasons, listDifferences = evaluate_product_candidate(
            listVisibleCandidates[iSelectedIndex], pszProductName, pszInputSpec
        )
        pszDetails: str = "一致理由: " + (", ".join(listReasons) or "全商品から表示")
        if listDifferences:
            pszDetails += "\n相違点: " + ", ".join(listDifferences)
        objDetails.set(pszDetails)

    def refresh_candidates(*_objArguments: object) -> None:
        nonlocal listVisibleCandidates
        listBaseCandidates: list[ProductCandidate]
        if objMode.get() == "全商品":
            listBaseCandidates = listAllCandidates
        elif objMode.get() == "同じ魚介カテゴリ":
            setInputCategories: set[str] = extract_term_groups(
                pszProductName, PRODUCT_CATEGORY_TERMS
            )
            listBaseCandidates = [
                objCandidate
                for objCandidate in listAllCandidates
                if setInputCategories
                & extract_term_groups(objCandidate.name, PRODUCT_CATEGORY_TERMS)
            ]
        else:
            listBaseCandidates = listRelatedCandidates
        listSearchTerms: list[str] = [
            compact_product_text(pszTerm)
            for pszTerm in objSearch.get().split()
            if compact_product_text(pszTerm)
        ]
        listVisibleCandidates = [
            objCandidate
            for objCandidate in listBaseCandidates
            if all(
                pszTerm
                in compact_product_text(
                    objCandidate.code + " " + objCandidate.name + " " + objCandidate.spec
                )
                for pszTerm in listSearchTerms
            )
        ]
        listDisplayValues: list[str] = [
            objCandidate.display_text for objCandidate in listVisibleCandidates
        ]
        objComboBox.configure(values=listDisplayValues)
        objCount.set("候補: " + str(len(listVisibleCandidates)) + "件")
        if listDisplayValues:
            objComboBox.current(0)
        else:
            objSelection.set("")
        update_candidate_details()

    def confirm_selection() -> None:
        nonlocal objSelectedCandidate
        iSelectedIndex: int = objComboBox.current()
        if iSelectedIndex < 0:
            messagebox.showerror("商品選択", "商品を選択してください。", parent=objRoot)
            return
        if pszSundayValue and not messagebox.askyesno(
            "商品名列の移動確認",
            "日曜日行の商品名セルに値があります。\n\n値: "
            + pszSundayValue
            + "\n\n商品名列を移動すると、この値はstep0002に残りません。\n続行しますか？",
            parent=objRoot,
        ):
            return
        objSelectedCandidate = listVisibleCandidates[iSelectedIndex]
        objRoot.destroy()

    def cancel_selection() -> None:
        objRoot.destroy()

    def select_no_matching_product() -> None:
        nonlocal bNoMatchingProduct
        if messagebox.askyesno(
            "該当商品なし",
            "商品マスターに該当商品がないものとしてstep0002を作成せず終了しますか？",
            parent=objRoot,
        ):
            bNoMatchingProduct = True
            objRoot.destroy()

    ttk.Button(objFrame, text="確定", command=confirm_selection).grid(row=9, column=0, sticky="e")
    ttk.Button(objFrame, text="該当商品なし", command=select_no_matching_product).grid(
        row=9, column=1
    )
    ttk.Button(objFrame, text="キャンセル", command=cancel_selection).grid(row=9, column=2, sticky="w")
    objSearch.trace_add("write", refresh_candidates)
    objModeComboBox.bind("<<ComboboxSelected>>", refresh_candidates)
    objComboBox.bind("<<ComboboxSelected>>", update_candidate_details)
    objRoot.protocol("WM_DELETE_WINDOW", cancel_selection)
    refresh_candidates()
    objSearchEntry.focus_set()
    objRoot.mainloop()
    if bNoMatchingProduct:
        raise NoMatchingProductError("商品マスターに該当商品なしが選択されました。")
    if objSelectedCandidate is None:
        raise SelectionCancelledError("商品選択がキャンセルされました。")
    return objSelectedCandidate


def get_step0002_output_paths(
    objStep0001ExcelPath: Path, objStep0001TsvPath: Path
) -> tuple[Path, Path]:
    """step0001の出力名からstep0002のXLSX・TSVパスを作ります。"""
    pszMarker: str = "ProductCodeSelector_step0001_"
    if not objStep0001ExcelPath.stem.startswith(pszMarker):
        raise ValueError("step0001の出力ファイル名ではありません。")
    pszStep0002Stem: str = objStep0001ExcelPath.stem.replace(
        pszMarker, "ProductCodeSelector_step0002_", 1
    )
    return (
        objStep0001ExcelPath.with_name(pszStep0002Stem + ".xlsx"),
        objStep0001TsvPath.with_name(pszStep0002Stem + ".tsv"),
    )


def build_step0002_rows(
    listStep0001Rows: list[list[str]], objCandidate: ProductCandidate
) -> list[list[str]]:
    """選択商品を月曜日へ設定し、元のG3～G8をG4～G9へ移動します。"""
    listRows: list[list[str]] = [listRow.copy() for listRow in listStep0001Rows]
    listOriginalProductNames: list[str] = [listRows[iRow][6] for iRow in range(2, 8)]
    for listRow in listRows[2:]:
        listRow[4] = ""
        listRow[5] = ""
    listRows[2][5] = objCandidate.code
    listRows[2][6] = objCandidate.name
    for iOffset, pszOriginalName in enumerate(listOriginalProductNames, start=3):
        listRows[iOffset][6] = pszOriginalName
    return listRows


def validate_step0002_outputs(
    objExcelPath: Path,
    objTsvPath: Path,
    listStep0001Rows: list[list[str]],
    objCandidate: ProductCandidate,
) -> None:
    """step0002のXLSX・TSV一致と、指定セル以外が不変であることを確認します。"""
    listExcelRows, _ = read_excel_table(objExcelPath)
    listTsvRows, _ = read_tsv_table(objTsvPath)
    if listExcelRows != listTsvRows:
        raise ValueError("step0002のXLSXとTSVの内容が一致しません。")
    listExpectedRows: list[list[str]] = build_step0002_rows(listStep0001Rows, objCandidate)
    if listExcelRows != listExpectedRows:
        raise ValueError("step0002の保存内容が仕様どおりではありません。")


def create_temporary_path(objOutputPath: Path) -> Path:
    """出力と同じフォルダーに一意な一時パスを作ります。"""
    iFileDescriptor, pszTemporaryPath = tempfile.mkstemp(
        prefix="." + objOutputPath.stem + ".", suffix=objOutputPath.suffix,
        dir=objOutputPath.parent,
    )
    os.close(iFileDescriptor)
    os.unlink(pszTemporaryPath)
    return Path(pszTemporaryPath)


def replace_output_pair(
    objTemporaryExcelPath: Path,
    objTemporaryTsvPath: Path,
    objExcelOutputPath: Path,
    objTsvOutputPath: Path,
) -> None:
    """XLSXとTSVをまとめて置換し、失敗時には以前の出力へ戻します。"""
    dictBackups: dict[Path, Path] = {}
    listReplaced: list[Path] = []
    try:
        for objOutputPath in (objExcelOutputPath, objTsvOutputPath):
            if objOutputPath.exists():
                objBackupPath = create_temporary_path(objOutputPath)
                os.replace(objOutputPath, objBackupPath)
                dictBackups[objOutputPath] = objBackupPath
        for objTemporaryPath, objOutputPath in (
            (objTemporaryExcelPath, objExcelOutputPath),
            (objTemporaryTsvPath, objTsvOutputPath),
        ):
            os.replace(objTemporaryPath, objOutputPath)
            listReplaced.append(objOutputPath)
    except Exception:
        for objOutputPath in reversed(listReplaced):
            if objOutputPath.exists():
                objOutputPath.unlink()
        for objOutputPath, objBackupPath in dictBackups.items():
            if objBackupPath.exists():
                os.replace(objBackupPath, objOutputPath)
        raise
    finally:
        for objBackupPath in dictBackups.values():
            if objBackupPath.exists():
                objBackupPath.unlink()


def get_error_path(objInputPath: Path) -> Path:
    """入力ファイルを基準にエラーテキストのパスを返します。"""
    return objInputPath.with_name(objInputPath.stem + "_error.txt")


def write_error_text(objErrorPath: Path, pszErrorMessage: str) -> None:
    """処理エラーをUTF-8テキストで保存します。"""
    pszText: str = (
        "処理名:\nProductCodeSelector step0001～step0002\n\n"
        + "エラー:\n"
        + pszErrorMessage
        + "\n"
    )
    objErrorPath.write_text(pszText, encoding="utf-8")


def process_input_file(
    pszInputFileFullPath: str,
) -> tuple[Path, Path, Path, Path, str, ProductCandidate]:
    """step0007からstep0001を作成し、商品選択後にstep0002を作成します。"""
    objInputPath: Path = validate_input_path(pszInputFileFullPath)
    create_abc_product_master(
        get_source_products_file_path(), get_products_file_path()
    )
    if objInputPath.suffix.lower() == ".xlsx":
        listRows, pszWorksheetTitle = read_excel_table(objInputPath)
    else:
        listRows, pszWorksheetTitle = read_tsv_table(objInputPath)
    validate_step0007_table(listRows)
    pszProductName: str = listRows[2][6].strip()
    objExcelOutputPath, objTsvOutputPath = get_output_paths(
        objInputPath, pszProductName
    )
    listOutputRows: list[list[str]] = clear_product_codes(listRows)
    objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath)
    objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath)
    try:
        save_excel_table(objTemporaryExcelPath, listOutputRows, pszWorksheetTitle)
        save_tsv_table(objTemporaryTsvPath, listOutputRows)
        validate_outputs_match(objTemporaryExcelPath, objTemporaryTsvPath)
        replace_output_pair(
            objTemporaryExcelPath,
            objTemporaryTsvPath,
            objExcelOutputPath,
            objTsvOutputPath,
        )
    finally:
        for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    listStep0001ExcelRows, pszStep0001WorksheetTitle = read_excel_table(objExcelOutputPath)
    listStep0001TsvRows, _ = read_tsv_table(objTsvOutputPath)
    if listStep0001ExcelRows != listStep0001TsvRows:
        raise ValueError("正式なstep0001のXLSXとTSVの内容が一致しません。")
    listAllCandidates: list[ProductCandidate] = read_product_candidates(
        get_products_file_path()
    )
    pszInputSpec: str = listStep0001ExcelRows[2][8].strip()
    listMatchedCandidates: list[ProductCandidate] = find_product_candidates(
        listAllCandidates, pszProductName, pszInputSpec
    )
    objSelectedCandidate: ProductCandidate = select_product_candidate(
        pszProductName,
        pszInputSpec,
        listMatchedCandidates,
        listAllCandidates,
        listStep0001ExcelRows[8][6].strip(),
    )
    listStep0002Rows: list[list[str]] = build_step0002_rows(
        listStep0001ExcelRows, objSelectedCandidate
    )
    objStep0002ExcelPath, objStep0002TsvPath = get_step0002_output_paths(
        objExcelOutputPath, objTsvOutputPath
    )
    objTemporaryStep0002ExcelPath: Path = create_temporary_path(objStep0002ExcelPath)
    objTemporaryStep0002TsvPath: Path = create_temporary_path(objStep0002TsvPath)
    try:
        save_excel_table(
            objTemporaryStep0002ExcelPath,
            listStep0002Rows,
            pszStep0001WorksheetTitle,
        )
        save_tsv_table(objTemporaryStep0002TsvPath, listStep0002Rows)
        validate_step0002_outputs(
            objTemporaryStep0002ExcelPath,
            objTemporaryStep0002TsvPath,
            listStep0001ExcelRows,
            objSelectedCandidate,
        )
        replace_output_pair(
            objTemporaryStep0002ExcelPath,
            objTemporaryStep0002TsvPath,
            objStep0002ExcelPath,
            objStep0002TsvPath,
        )
    finally:
        for objTemporaryPath in (
            objTemporaryStep0002ExcelPath,
            objTemporaryStep0002TsvPath,
        ):
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    return (
        objExcelOutputPath,
        objTsvOutputPath,
        objStep0002ExcelPath,
        objStep0002TsvPath,
        pszProductName,
        objSelectedCandidate,
    )


def parse_command_line_arguments() -> str:
    """1つの入力ファイルパスを解析します。"""
    if len(sys.argv) != 2 or sys.argv[1].startswith("--"):
        raise ValueError("入力ファイルパスは1つ指定してください。")
    return sys.argv[1]


def main() -> int:
    """成功0・失敗1・キャンセル2・該当商品なし3の終了コードを返します。"""
    configure_standard_streams()
    try:
        pszInputFileFullPath: str = parse_command_line_arguments()
    except ValueError as objException:
        pszMessage: str = (
            "Error: "
            + str(objException)
            + "\nUsage: python "
            + os.path.basename(__file__)
            + " <input_file_path>\n"
        )
        print(pszMessage, file=sys.stderr, end="")
        Path(os.path.splitext(os.path.basename(__file__))[0] + "_error_argument.txt").write_text(
            pszMessage, encoding="utf-8"
        )
        return 1
    try:
        (
            objStep0001ExcelPath,
            objStep0001TsvPath,
            objStep0002ExcelPath,
            objStep0002TsvPath,
            pszProductName,
            objSelectedCandidate,
        ) = process_input_file(pszInputFileFullPath)
    except SelectionCancelledError as objException:
        print("キャンセル: " + str(objException), file=sys.stderr)
        return 2
    except NoMatchingProductError as objException:
        print("該当商品なし: " + str(objException), file=sys.stderr)
        return 3
    except Exception as objException:
        pszMessage = "Error: " + str(objException)
        print(pszMessage, file=sys.stderr)
        try:
            write_error_text(
                get_error_path(Path(pszInputFileFullPath).expanduser().resolve()),
                str(objException),
            )
        except OSError as objWriteException:
            print(
                "Error: エラーファイルを保存できません。Detail = "
                + str(objWriteException),
                file=sys.stderr,
            )
        return 1
    print("ProductCodeSelector step0001～step0002の作成が完了しました。")
    print("商品名: " + pszProductName)
    print("選択商品: " + objSelectedCandidate.display_text)
    print("step0001 XLSX: " + str(objStep0001ExcelPath))
    print("step0001 TSV: " + str(objStep0001TsvPath))
    print("step0002 XLSX: " + str(objStep0002ExcelPath))
    print("step0002 TSV: " + str(objStep0002TsvPath))
    return 0


if __name__ == "__main__":
    sys.exit(main())
