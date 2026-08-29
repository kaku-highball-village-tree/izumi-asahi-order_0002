"""Excelファイル内の「BY」を含むシートをTSV出力するコマンドラインツール。"""

from __future__ import annotations

import csv
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

START_MESSAGE = "izumi-asahi-order_0001 started"
TARGET_SHEET_KEYWORD = "BY"
CODE_HEADER_VALUE = "コード"
INCORRECT_VENDOR_NAME = "ショクリュ-"
CORRECT_VENDOR_NAME = "ショクリュー"
TARGET_VENDOR_NAMES = {"中村角", "旭物産"}
USAGE = "Usage: python src/izumi-asahi-order_0001_Cmd.py <excel_file_path>"


def build_output_path(excel_path: Path, sheet_index: int) -> Path:
    """BYシートの順番に応じたTSV出力パスを作成する。"""
    if sheet_index == 1:
        return excel_path.with_suffix(".tsv")

    return excel_path.with_name(f"{excel_path.stem}_{sheet_index:04d}.tsv")


def build_step0001_output_path(tsv_path: Path) -> Path:
    """step0001 TSVの出力パスを作成する。"""
    return tsv_path.with_name(f"{tsv_path.stem}_step0001.tsv")


def build_error_output_path(step_output_path: Path) -> Path:
    """step0001処理のエラー出力パスを作成する。"""
    return step_output_path.with_name(f"{step_output_path.stem}_error.txt")


def build_step0002_output_path(step0001_path: Path) -> Path:
    """step0002 TSVの出力パスを作成する。"""
    if step0001_path.stem.endswith("_step0001"):
        output_stem = f"{step0001_path.stem.removesuffix('_step0001')}_step0002"
    else:
        output_stem = f"{step0001_path.stem}_step0002"

    return step0001_path.with_name(f"{output_stem}.tsv")


def build_step0003_output_path(step0002_path: Path) -> Path:
    """step0003 TSVの出力パスを作成する。"""
    if step0002_path.stem.endswith("_step0002"):
        output_stem = f"{step0002_path.stem.removesuffix('_step0002')}_step0003"
    else:
        output_stem = f"{step0002_path.stem}_step0003"

    return step0002_path.with_name(f"{output_stem}.tsv")


def build_warning_output_path(step_output_path: Path) -> Path:
    """処理のwarning出力パスを作成する。"""
    return step_output_path.with_name(f"{step_output_path.stem}_warning.txt")


def normalize_cell_value(value: object) -> str:
    """TSV出力用にセル値を文字列へ変換する。"""
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")

    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")

    return str(value)


def get_a_column_value(row: list[str]) -> str:
    """A列の値を取得し、存在しない場合は空文字を返す。"""
    if not row:
        return ""

    return row[0].strip()


def is_blank_row(row: list[str]) -> bool:
    """行内のすべてのセルが空または空白だけかどうかを判定する。"""
    return all(cell.strip() == "" for cell in row)


def write_sheet_to_tsv(rows: Iterable[tuple[object, ...]], output_path: Path) -> None:
    """Excelシートの行データをTSVファイルへ出力する。"""
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.writer(output_file, delimiter="\t", lineterminator="\n")
        for row in rows:
            writer.writerow([normalize_cell_value(cell) for cell in row])


def read_tsv(tsv_path: Path) -> list[list[str]]:
    """TSVファイルを読み込む。"""
    with tsv_path.open("r", encoding="utf-8-sig", newline="") as input_file:
        return list(csv.reader(input_file, delimiter="\t"))


def write_rows_to_tsv(rows: Iterable[list[str]], output_path: Path) -> None:
    """行リストをTSVファイルへ出力する。"""
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.writer(output_file, delimiter="\t", lineterminator="\n")
        writer.writerows(rows)


def write_error_file(error_path: Path, message: str) -> None:
    """エラー内容をテキストファイルへ出力する。"""
    with error_path.open("w", encoding="utf-8", newline="") as error_file:
        error_file.write(f"{message}\n")


def write_warning_file(warning_path: Path, message: str) -> None:
    """warning内容をテキストファイルへ出力する。"""
    with warning_path.open("w", encoding="utf-8", newline="") as warning_file:
        warning_file.write(f"{message}\n")


def find_code_row_index(rows: list[list[str]]) -> int | None:
    """A列が「コード」と完全一致する行番号を返す。"""
    for index, row in enumerate(rows):
        if get_a_column_value(row) == CODE_HEADER_VALUE:
            return index

    return None


def find_code_header_rows(tsv_path: Path) -> tuple[list[str], list[str], int]:
    """店舗コード行と見出し行を取得する。"""
    rows = read_tsv(tsv_path)
    code_row_index = find_code_row_index(rows)

    if code_row_index is None:
        raise ValueError("Error: A列が「コード」の行が見つかりません。")

    if code_row_index == 0:
        raise ValueError("Error: 「コード」行の1つ上の行が存在しません。")

    store_code_row = rows[code_row_index - 1]
    header_row = rows[code_row_index]
    return store_code_row, header_row, code_row_index


def extract_rows_after_code_until_blank(
    rows: list[list[str]], code_row_index: int
) -> list[list[str]]:
    """「コード」行の次の行から完全空白行の1つ上の行までを取得する。"""
    detail_rows: list[list[str]] = []

    for row in rows[code_row_index + 1 :]:
        if is_blank_row(row):
            break

        detail_rows.append(row)

    return detail_rows


def normalize_vendor_name_in_a_column(row: list[str]) -> list[str]:
    """A列の業者名表記を必要に応じて修正する。"""
    if row and row[0].strip() == INCORRECT_VENDOR_NAME:
        normalized_row = row.copy()
        normalized_row[0] = CORRECT_VENDOR_NAME
        return normalized_row

    return row


def extract_target_vendor_blocks(rows: list[list[str]]) -> tuple[list[list[str]], bool]:
    """先頭2行と対象業者のデータブロックだけを取得する。"""
    output_rows = rows[:2]
    is_target_block = False
    target_vendor_found = False

    for row in rows[2:]:
        vendor_name = get_a_column_value(row)
        if vendor_name:
            is_target_block = vendor_name in TARGET_VENDOR_NAMES
            if is_target_block:
                target_vendor_found = True

        if is_target_block:
            output_rows.append(row)

    return output_rows, target_vendor_found


def process_step0003_tsv(step0002_path: Path) -> Path | None:
    """step0002 TSVから対象業者のブロックを抽出しstep0003 TSVを出力する。"""
    step0003_output_path = build_step0003_output_path(step0002_path)
    error_output_path = build_error_output_path(step0003_output_path)
    warning_output_path = build_warning_output_path(step0003_output_path)

    if not step0002_path.exists():
        message = f"Error: file not found: {step0002_path}"
        print(message)
        write_error_file(error_output_path, message)
        return None

    try:
        rows = read_tsv(step0002_path)
        output_rows, target_vendor_found = extract_target_vendor_blocks(rows)
        write_rows_to_tsv(output_rows, step0003_output_path)
        if not target_vendor_found:
            message = "Warning: no vendor blocks for 「中村角」 or 「旭物産」 were found."
            print(message)
            write_warning_file(warning_output_path, message)
    except OSError as error:
        message = f"Error: failed to process TSV file: {step0002_path}: {error}"
        print(message)
        write_error_file(error_output_path, message)
        return None

    print(f"Exported step0003 TSV to '{step0003_output_path}'")
    return step0003_output_path


def process_step0002_tsv(step0001_path: Path) -> Path | None:
    """step0001 TSVのA列業者名を修正しstep0002 TSVを出力する。"""
    step0002_output_path = build_step0002_output_path(step0001_path)
    error_output_path = build_error_output_path(step0002_output_path)

    if not step0001_path.exists():
        message = f"Error: file not found: {step0001_path}"
        print(message)
        write_error_file(error_output_path, message)
        return None

    try:
        rows = read_tsv(step0001_path)
        normalized_rows = [normalize_vendor_name_in_a_column(row) for row in rows]
        write_rows_to_tsv(normalized_rows, step0002_output_path)
    except OSError as error:
        message = f"Error: failed to process TSV file: {step0001_path}: {error}"
        print(message)
        write_error_file(error_output_path, message)
        return None

    print(f"Exported step0002 TSV to '{step0002_output_path}'")
    return step0002_output_path


def process_step0001_tsv(tsv_path: Path) -> Path | None:
    """TSVから店舗コード行・見出し行・明細行を抽出しstep0001 TSVを出力する。"""
    step_output_path = build_step0001_output_path(tsv_path)
    error_output_path = build_error_output_path(step_output_path)

    try:
        rows = read_tsv(tsv_path)
        store_code_row, header_row, code_row_index = find_code_header_rows(tsv_path)
        detail_rows = extract_rows_after_code_until_blank(rows, code_row_index)
        write_rows_to_tsv(
            [store_code_row, header_row, *detail_rows],
            step_output_path,
        )
    except ValueError as error:
        message = str(error)
        print(message)
        write_error_file(error_output_path, message)
        return None

    print(f"Exported step0001 TSV to '{step_output_path}'")
    return step_output_path


def export_by_sheets_to_tsv(excel_path: Path) -> list[Path]:
    """シート名にBYを含むシートだけをTSVファイルとして出力する。"""
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    by_sheet_names = [
        sheet_name
        for sheet_name in workbook.sheetnames
        if TARGET_SHEET_KEYWORD in sheet_name
    ]

    if not by_sheet_names:
        print("No sheets containing 'BY' were found.")
        workbook.close()
        return []

    output_paths: list[Path] = []

    for index, sheet_name in enumerate(by_sheet_names, start=1):
        worksheet = workbook[sheet_name]
        output_path = build_output_path(excel_path, index)
        write_sheet_to_tsv(worksheet.iter_rows(values_only=True), output_path)
        output_paths.append(output_path)
        print(f"Exported sheet '{sheet_name}' to '{output_path}'")

    workbook.close()
    return output_paths


def main() -> int:
    """コマンドライン引数を読み取り、ExcelからTSVへの出力を実行する。"""
    print(START_MESSAGE)

    if len(sys.argv) < 2:
        print(USAGE)
        return 1

    excel_path = Path(sys.argv[1])

    if not excel_path.exists():
        print(f"Error: file not found: {excel_path}")
        return 1

    if not excel_path.is_file():
        print(f"Error: not a file: {excel_path}")
        return 1

    if excel_path.suffix.lower() != ".xlsx":
        print(f"Error: expected an .xlsx file: {excel_path}")
        return 1

    tsv_paths = export_by_sheets_to_tsv(excel_path)
    for tsv_path in tsv_paths:
        step0001_path = process_step0001_tsv(tsv_path)
        if step0001_path is not None:
            step0002_path = process_step0002_tsv(step0001_path)
            if step0002_path is not None:
                process_step0003_tsv(step0002_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
