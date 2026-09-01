# -- coding: utf-8 --
###############################################################
#
# AsahiOrderTemplateMaker_Cmd.py
#
# pip install openpyxl tkcalendar
#
###############################################################

import csv
import hashlib
import math
import os
import random
import re
import shutil
import sys
import tempfile
import tkinter as tk
from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import messagebox

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from tkcalendar import Calendar


EXPECTED_HEADERS: tuple[str, str, str] = ("productCode", "productName", "spec")
STEP0002_HEADERS: tuple[str, ...] = (
    "納品日",
    "曜日",
    "配送パターン",
    "便",
    "Ｐ品番",
    "APEX品番",
    "商品名",
    "産地",
    "仕様",
    "伝票原価",
    "伝票売価",
    "値入",
    "売価",
    "単位",
)
WEEKDAYS: tuple[str, ...] = ("月", "火", "水", "木", "金", "土", "日")
STORE_DEFINITIONS: tuple[tuple[str, str], ...] = (
    ("2", "光の森"), ("3", "学園"), ("4", "ゆめタウン呉"),
    ("6", "サンモール"), ("7", "大川"), ("10", "中津"),
    ("11", "夢彩都"), ("13", "安古市"), ("16", "倉敷"),
    ("17", "大牟田"), ("18", "南岩国"), ("19", "平島"),
    ("20", "蔵王"), ("21", "サンピアン"), ("22", "柳井"),
    ("24", "高梁"), ("25", "東広島"), ("26", "黒瀬"),
    ("28", "井原"), ("29", "江田島"), ("30", "八代"),
    ("31", "吉田"), ("33", "久世"), ("34", "長府"),
    ("35", "新南陽"), ("36", "山陽"), ("37", "浜田"),
    ("38", "益田"), ("39", "斐川"), ("40", "高松"),
    ("41", "行橋"), ("45", "川尻"), ("46", "武雄"),
    ("47", "YM浜田"), ("48", "美作"), ("49", "八本松"),
    ("50", "防府"), ("51", "邑久"), ("52", "遠賀"),
    ("53", "筑紫野"), ("54", "宇部"), ("55", "丹波"),
    ("56", "山口"), ("57", "大竹"), ("58", "博多"),
    ("59", "八女"), ("61", "赤間"), ("63", "神西"),
    ("66", "八木"), ("68", "府中"), ("69", "八幡"),
    ("70", "久留米"), ("74", "己斐"), ("76", "はません"),
    ("80", "尾道"), ("82", "沼田"), ("86", "シティモール"),
    ("101", "佐賀"), ("102", "別府"), ("103", "広島"),
    ("104", "出雲"), ("105", "三豊"), ("106", "丸亀"),
    ("107", "みゆき"), ("108", "ゆめシティ"), ("109", "うきは"),
    ("110", "津山"), ("111", "徳島"), ("112", "小倉東"),
    ("114", "下関駅"), ("115", "新宮"), ("116", "西栄"),
    ("117", "田崎"), ("118", "玉名"), ("119", "阿賀"),
    ("120", "二葉の里"), ("121", "下関"), ("122", "松橋"),
    ("123", "大江"), ("124", "柳川"), ("125", "廿日市"),
    ("126", "YMさが"), ("127", "筑後"), ("128", "徳山"),
    ("131", "ＬＥＣＴ"), ("132", "南行橋"), ("133", "江津"),
    ("134", "福津"), ("135", "姫路"), ("136", "下松"),
    ("139", "福山"), ("140", "城野"), ("141", "南小野田"),
    ("143", "日田"), ("144", "青山"), ("145", "西条"),
    ("146", "飯塚"), ("147", "新大村"), ("148", "祇園"),
    ("151", "五日市"),
)
SUPPORTED_EXTENSIONS: set[str] = {".xlsx", ".tsv", ".csv"}
AREA_STORE_MAPPING_FILE_NAME: str = "AsahiOrderAreaStoreMapping_対応表.txt"
# 0.25は25%の商品削除確率を意味します。
PRODUCT_DELETE_PROBABILITY: float = 0.25
MIN_PRODUCT_COUNT: int = 1
MAX_PRODUCT_COUNT: int = 10


class ProductRow:
    """テンプレートへ出力する1商品の3列を保持します。"""

    def __init__(self, product_code: str, product_name: str, spec: str) -> None:
        self.product_code: str = product_code
        self.product_name: str = product_name
        self.spec: str = spec


class ProductFilePlan:
    """商品別Step0003～Step0007の識別情報と出力パスを保持します。"""

    def __init__(
        self,
        iSourceRow: int,
        listStep0002Row: list[str],
        pszApexCode: str,
        pszProductName: str,
        pszSafeApexCode: str,
        pszSafeProductName: str,
    ) -> None:
        self.source_row: int = iSourceRow
        self.step0002_row: list[str] = listStep0002Row
        self.apex_code: str = pszApexCode
        self.product_name: str = pszProductName
        self.safe_apex_code: str = pszSafeApexCode
        self.safe_product_name: str = pszSafeProductName
        self.file_identity: str = pszSafeApexCode + "_" + pszSafeProductName
        self.output_paths: dict[str, Path] = {}
        self.center_output_paths: dict[str, dict[str, Path]] = {}


class ProductProcessingProgress:
    """商品別処理の進行状況と復旧結果を保持します。"""

    def __init__(self) -> None:
        self.process_name: str = "処理0003"
        self.completed_process_name: str = "なし"
        self.current_product: int = 0
        self.total_products: int = 0
        self.apex_code: str = "なし"
        self.product_name: str = "なし"
        self.center_name: str = "なし"
        self.planned_outputs: int = 0
        self.actual_output_targets: int = 0
        self.temporary_outputs: int = 0
        self.validated_outputs: int = 0
        self.committed_outputs: int = 0
        self.skipped_centers: int = 0
        self.warning_files: int = 0
        self.stale_files: int = 0
        self.temp_copies: int = 0
        self.renamed_files: int = 0
        self.restore_successes: int = 0
        self.restore_failures: int = 0
        self.temp_backup_directory: str = "なし"
        self.failed_restore_paths: list[str] = []


class Step0002Error(Exception):
    """処理0002で発生したエラーであることを呼び出し元へ伝えます。"""


class Step0003Error(Exception):
    """処理0003で発生したエラーであることを呼び出し元へ伝えます。"""


class Step0004Error(Exception):
    """処理0004で発生したエラーであることを呼び出し元へ伝えます。"""


class Step0005Error(Exception):
    """処理0005で発生したエラーであることを呼び出し元へ伝えます。"""


class Step0006Error(Exception):
    """処理0006で発生したエラーであることを呼び出し元へ伝えます。"""


class Step0007Error(Exception):
    """処理0007で発生したエラーであることを呼び出し元へ伝えます。"""


def write_error_text(pszOutputFileFullPath: str, pszErrorMessage: str) -> None:
    """エラーメッセージをUTF-8のテキストファイルへ上書き保存します。"""
    pszDirectoryFullPath: str = os.path.dirname(pszOutputFileFullPath)
    if pszDirectoryFullPath != "":
        os.makedirs(pszDirectoryFullPath, exist_ok=True)
    with open(pszOutputFileFullPath, mode="w", encoding="utf-8", newline="") as objFile:
        objFile.write(pszErrorMessage.rstrip("\n") + "\n")


def write_warning_text(pszOutputFileFullPath: str, pszWarningMessage: str) -> None:
    """警告メッセージをUTF-8のテキストファイルへ上書き保存します。"""
    pszDirectoryFullPath: str = os.path.dirname(pszOutputFileFullPath)
    if pszDirectoryFullPath != "":
        os.makedirs(pszDirectoryFullPath, exist_ok=True)
    with open(pszOutputFileFullPath, mode="w", encoding="utf-8", newline="") as objFile:
        objFile.write(pszWarningMessage.rstrip("\n") + "\n")


def get_error_file_full_path(pszInputFileFullPath: str) -> str:
    """入力ファイルと同じフォルダーに作る_error.txtのパスを返します。"""
    pszDirectoryFullPath: str = os.path.dirname(os.path.abspath(pszInputFileFullPath))
    pszBaseNameWithoutExtension: str = os.path.splitext(
        os.path.basename(pszInputFileFullPath)
    )[0]
    return os.path.join(pszDirectoryFullPath, pszBaseNameWithoutExtension + "_error.txt")


def get_step0005_warning_file_path(pszInputFileFullPath: str) -> Path:
    """入力ファイルと同じフォルダーのstep0005_warning.txtパスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath).resolve()
    return objInputPath.with_name(objInputPath.stem + "_step0005_warning.txt")


def report_processing_error(
    pszInputFileFullPath: str, pszProcessName: str, pszDetailMessage: str
) -> None:
    """標準エラーと入力ファイル用_error.txtへ同じエラーを出力します。"""
    pszErrorMessage: str = (
        "処理結果: エラー\n"
        + "入力ファイル: "
        + os.path.abspath(pszInputFileFullPath)
        + "\n発生した処理: "
        + pszProcessName
        + "\nエラー内容: "
        + pszDetailMessage
        + "\n"
    )
    print(pszErrorMessage, file=sys.stderr, end="")
    try:
        write_error_text(get_error_file_full_path(pszInputFileFullPath), pszErrorMessage)
    except OSError as objException:
        print(
            "Error: _error.txtの保存にも失敗しました。Detail = " + str(objException),
            file=sys.stderr,
        )


def remove_old_error_file(pszInputFileFullPath: str) -> None:
    """正常終了後、以前の処理で作られた_error.txtがあれば削除します。"""
    pszErrorFileFullPath: str = get_error_file_full_path(pszInputFileFullPath)
    if os.path.exists(pszErrorFileFullPath):
        os.remove(pszErrorFileFullPath)


def validate_input_path(pszInputFileFullPath: str) -> str:
    """入力パスと拡張子を検証し、絶対パスを返します。"""
    pszAbsolutePath: str = os.path.abspath(pszInputFileFullPath)
    if not os.path.exists(pszAbsolutePath):
        raise ValueError("入力ファイルが見つかりません。Path = " + pszAbsolutePath)
    if not os.path.isfile(pszAbsolutePath):
        raise ValueError("入力パスがファイルではありません。Path = " + pszAbsolutePath)
    pszExtension: str = os.path.splitext(pszAbsolutePath)[1].lower()
    if pszExtension not in SUPPORTED_EXTENSIONS:
        raise ValueError("入力ファイルの拡張子は未対応です。Path = " + pszAbsolutePath)
    return pszAbsolutePath


def normalize_header(objValue: object) -> str:
    """ヘッダー比較用に値を前後空白のない文字列へ変換します。"""
    if objValue is None:
        return ""
    return str(objValue).strip()


def normalize_text(objValue: object) -> str:
    """出力用にNoneを空文字、それ以外を文字列へ変換します。"""
    if objValue is None:
        return ""
    return str(objValue)


def get_this_week_monday(objToday: date | None = None) -> date:
    """基準日を含む週の月曜日を返します。"""
    objBaseDate: date = date.today() if objToday is None else objToday
    return objBaseDate - timedelta(days=objBaseDate.weekday())


def validate_start_monday(objStartMonday: date) -> None:
    """開始日が月曜日であることを確認します。"""
    if objStartMonday.weekday() != 0:
        raise ValueError(
            "開始日は月曜日を指定してください。Date = "
            + objStartMonday.isoformat()
        )


def parse_start_monday(pszValue: str) -> date:
    """YYYY-MM-DDを開始月曜日として解析・検証します。"""
    if len(pszValue) != 10:
        raise ValueError("開始日の形式はYYYY-MM-DDではありません。Value = " + pszValue)
    try:
        objStartMonday: date = date.fromisoformat(pszValue)
    except ValueError as objException:
        raise ValueError(
            "開始日の形式はYYYY-MM-DDではありません。Value = " + pszValue
        ) from objException
    if objStartMonday.isoformat() != pszValue:
        raise ValueError("開始日の形式はYYYY-MM-DDではありません。Value = " + pszValue)
    validate_start_monday(objStartMonday)
    return objStartMonday


def select_start_monday() -> date | None:
    """カレンダーを表示し、利用者が選んだ月曜日またはキャンセル時Noneを返します。"""
    objThisWeekMonday: date = get_this_week_monday()
    objNextWeekMonday: date = objThisWeekMonday + timedelta(days=7)
    objSelectedMonday: date | None = objNextWeekMonday
    objLastValidMonday: date = objNextWeekMonday

    objRoot = tk.Tk()
    objRoot.title("Asahi Order Template Maker - 開始月曜日の選択")
    objRoot.resizable(False, False)

    objInstructionLabel = tk.Label(
        objRoot,
        text="処理0004の開始月曜日を選択してください。",
        padx=10,
        pady=8,
    )
    objInstructionLabel.pack()

    objCalendar = Calendar(
        objRoot,
        selectmode="day",
        year=objNextWeekMonday.year,
        month=objNextWeekMonday.month,
        day=objNextWeekMonday.day,
        date_pattern="yyyy-mm-dd",
        firstweekday="monday",
    )
    objCalendar.pack(padx=10, pady=5)

    def set_selected_monday(objMonday: date) -> None:
        nonlocal objSelectedMonday, objLastValidMonday
        objSelectedMonday = objMonday
        objLastValidMonday = objMonday
        objCalendar.selection_set(objMonday)

    def on_calendar_selected(_objEvent: object = None) -> None:
        nonlocal objSelectedMonday
        objSelectedDate: date = objCalendar.selection_get()
        if objSelectedDate.weekday() != 0:
            objCalendar.selection_set(objLastValidMonday)
            messagebox.showerror(
                "Asahi Order Template Maker",
                "月曜日を選択してください。",
                parent=objRoot,
            )
            return
        objSelectedMonday = objSelectedDate

    def confirm_selection() -> None:
        nonlocal objSelectedMonday
        objSelectedDate: date = objCalendar.selection_get()
        if objSelectedDate.weekday() != 0:
            messagebox.showerror(
                "Asahi Order Template Maker",
                "月曜日を選択してください。",
                parent=objRoot,
            )
            return
        objSelectedMonday = objSelectedDate
        objRoot.destroy()

    def cancel_selection() -> None:
        nonlocal objSelectedMonday
        objSelectedMonday = None
        objRoot.destroy()

    objCalendar.bind("<<CalendarSelected>>", on_calendar_selected)
    objButtonFrame = tk.Frame(objRoot, padx=10, pady=10)
    objButtonFrame.pack(fill=tk.X)
    tk.Button(
        objButtonFrame,
        text="今週の月曜日",
        command=lambda: set_selected_monday(objThisWeekMonday),
    ).pack(side=tk.LEFT, padx=2)
    tk.Button(
        objButtonFrame,
        text="来週の月曜日",
        command=lambda: set_selected_monday(objNextWeekMonday),
    ).pack(side=tk.LEFT, padx=2)
    tk.Button(objButtonFrame, text="決定", command=confirm_selection).pack(
        side=tk.LEFT, padx=8
    )
    tk.Button(objButtonFrame, text="キャンセル", command=cancel_selection).pack(
        side=tk.LEFT, padx=2
    )
    objRoot.protocol("WM_DELETE_WINDOW", cancel_selection)
    objRoot.bind("<Escape>", lambda _objEvent: cancel_selection())
    objRoot.lift()
    objRoot.attributes("-topmost", True)
    objRoot.after_idle(lambda: objRoot.attributes("-topmost", False))
    objRoot.mainloop()
    return objSelectedMonday


def show_start_monday_cancelled_message() -> None:
    """開始月曜日の選択キャンセルにより処理を中止することを通知します。"""
    objRoot = tk.Tk()
    objRoot.withdraw()
    messagebox.showerror(
        "Asahi Order Template Maker",
        "開始月曜日の選択がキャンセルされました。\n"
        "処理0004を完了できないため、処理を中止します。",
        parent=objRoot,
    )
    objRoot.destroy()


def validate_headers(listValues: list[object], pszSourceName: str) -> None:
    """先頭3列が仕様どおりのヘッダーであることを確認します。"""
    if len(listValues) < 3:
        raise ValueError(pszSourceName + "のヘッダーは3列未満です。")
    tupleHeaders: tuple[str, str, str] = tuple(
        normalize_header(objValue) for objValue in listValues[:3]
    )  # type: ignore[assignment]
    if tupleHeaders != EXPECTED_HEADERS:
        raise ValueError(
            pszSourceName
            + "の先頭3列はproductCode、productName、specではありません。"
        )


def find_target_worksheet(objWorkbook: Workbook) -> Worksheet:
    """A1～C1が仕様どおりのシートを探し、1枚だけなら返します。"""
    listTargetWorksheets: list[Worksheet] = []
    for objWorksheet in objWorkbook.worksheets:
        listHeaders: list[object] = [
            objWorksheet.cell(row=1, column=iColumn).value for iColumn in range(1, 4)
        ]
        if tuple(normalize_header(objValue) for objValue in listHeaders) == EXPECTED_HEADERS:
            listTargetWorksheets.append(objWorksheet)
    if len(listTargetWorksheets) == 0:
        raise ValueError(
            "A1～C1がproductCode、productName、specのシートが見つかりません。"
        )
    if len(listTargetWorksheets) > 1:
        raise ValueError(
            "対象シートが複数見つかりました。対象シート = "
            + ", ".join(objWorksheet.title for objWorksheet in listTargetWorksheets)
        )
    return listTargetWorksheets[0]


def build_product_row(listValues: list[object], iRow: int) -> ProductRow | None:
    """先頭3列を検証し、空行ならNone、それ以外なら商品行を返します。"""
    if len(listValues) < 3:
        if all(normalize_text(objValue).strip() == "" for objValue in listValues):
            return None
        raise ValueError(str(iRow) + "行目は3列未満です。")
    listTexts: list[str] = [normalize_text(objValue) for objValue in listValues[:3]]
    if all(pszValue.strip() == "" for pszValue in listTexts):
        return None
    pszProductCode: str = listTexts[0].strip()
    if pszProductCode == "":
        raise ValueError(
            str(iRow) + "行目はproductCodeが空ですが、ほかの対象列に値があります。"
        )
    return ProductRow(pszProductCode, listTexts[1], listTexts[2])


def validate_unique_product_codes(
    listProductRows: list[ProductRow], pszSourceName: str
) -> None:
    """productCodeが重複していないことを確認します。"""
    setSeenCodes: set[str] = set()
    setDuplicateCodes: set[str] = set()
    for objProductRow in listProductRows:
        if objProductRow.product_code in setSeenCodes:
            setDuplicateCodes.add(objProductRow.product_code)
        setSeenCodes.add(objProductRow.product_code)
    if setDuplicateCodes:
        raise ValueError(
            pszSourceName
            + "でproductCodeが重複しています。productCode = "
            + ", ".join(sorted(setDuplicateCodes))
        )


def validate_product_count(listProductRows: list[ProductRow]) -> None:
    """商品数が1品目以上10品目以下であることを確認します。"""
    iProductCount: int = len(listProductRows)
    if iProductCount < MIN_PRODUCT_COUNT:
        raise ValueError(
            "入力ファイルの商品数が0品目です。"
            + "商品数は"
            + str(MIN_PRODUCT_COUNT)
            + "品目以上"
            + str(MAX_PRODUCT_COUNT)
            + "品目以下にしてください。"
        )
    if iProductCount > MAX_PRODUCT_COUNT:
        raise ValueError(
            "入力ファイルの商品数が"
            + str(MAX_PRODUCT_COUNT)
            + "品目を超えています。商品数 = "
            + str(iProductCount)
            + "、上限 = "
            + str(MAX_PRODUCT_COUNT)
        )


def read_excel_rows(pszInputFileFullPath: str) -> list[ProductRow]:
    """Excelの対象シートからテンプレート用商品行を読み取ります。"""
    objWorkbook: Workbook = load_workbook(pszInputFileFullPath, data_only=True)
    objWorksheet: Worksheet = find_target_worksheet(objWorkbook)
    listProductRows: list[ProductRow] = []
    for iRow in range(2, objWorksheet.max_row + 1):
        listValues: list[object] = [
            objWorksheet.cell(row=iRow, column=iColumn).value for iColumn in range(1, 4)
        ]
        objProductRow: ProductRow | None = build_product_row(listValues, iRow)
        if objProductRow is not None:
            listProductRows.append(objProductRow)
    validate_unique_product_codes(listProductRows, "Excel")
    return listProductRows


def read_delimited_rows_with_encoding(
    pszInputFileFullPath: str, pszEncoding: str, pszDelimiter: str
) -> list[list[str]]:
    """指定文字コードと区切り文字で全レコードを読み取ります。"""
    with open(
        pszInputFileFullPath, mode="r", encoding=pszEncoding, newline=""
    ) as objFile:
        return list(csv.reader(objFile, delimiter=pszDelimiter, strict=True))


def read_delimited_rows(pszInputFileFullPath: str) -> list[ProductRow]:
    """CSVまたはTSVをUTF-8優先、失敗時CP932で読み取ります。"""
    pszExtension: str = os.path.splitext(pszInputFileFullPath)[1].lower()
    pszDelimiter: str = "\t" if pszExtension == ".tsv" else ","
    pszSourceName: str = "TSV" if pszExtension == ".tsv" else "CSV"
    try:
        listRows: list[list[str]] = read_delimited_rows_with_encoding(
            pszInputFileFullPath, "utf-8-sig", pszDelimiter
        )
    except UnicodeDecodeError:
        listRows = read_delimited_rows_with_encoding(
            pszInputFileFullPath, "cp932", pszDelimiter
        )
    if not listRows:
        raise ValueError(pszSourceName + "ファイルが空です。")
    validate_headers(listRows[0], pszSourceName)
    listProductRows: list[ProductRow] = []
    for iRow, listValues in enumerate(listRows[1:], start=2):
        objProductRow: ProductRow | None = build_product_row(listValues, iRow)
        if objProductRow is not None:
            listProductRows.append(objProductRow)
    validate_unique_product_codes(listProductRows, pszSourceName)
    return listProductRows


def get_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0001.xlsxと_step0001.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0001")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def get_step0002_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0002.xlsxと_step0002.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0002")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def get_step0003_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0003.xlsxと_step0003.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0003")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def get_step0004_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0004.xlsxと_step0004.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0004")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def get_step0005_output_file_paths(pszInputFileFullPath: str) -> tuple[Path, Path]:
    """_step0005.xlsxと_step0005.tsvの出力パスを返します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    objBasePath: Path = objInputPath.with_name(objInputPath.stem + "_step0005")
    return objBasePath.with_suffix(".xlsx"), objBasePath.with_suffix(".tsv")


def create_temporary_path(objOutputPath: Path, pszSuffix: str) -> Path:
    """出力先と同じフォルダーに一意な一時ファイルパスを作ります。"""
    iFileDescriptor, pszTemporaryPath = tempfile.mkstemp(
        prefix=objOutputPath.stem + "_", suffix=pszSuffix, dir=objOutputPath.parent
    )
    os.close(iFileDescriptor)
    return Path(pszTemporaryPath)


def save_excel_template(objOutputPath: Path, listProductRows: list[ProductRow]) -> None:
    """3列の新規Excelテンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(EXPECTED_HEADERS))
    for objProductRow in listProductRows:
        objWorksheet.append(
            [objProductRow.product_code, objProductRow.product_name, objProductRow.spec]
        )
        objWorksheet.cell(row=objWorksheet.max_row, column=1).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_tsv_template(objOutputPath: Path, listProductRows: list[ProductRow]) -> None:
    """3列のTSVをUTF-8 BOMなし、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(EXPECTED_HEADERS)
        for objProductRow in listProductRows:
            objWriter.writerow(
                [objProductRow.product_code, objProductRow.product_name, objProductRow.spec]
            )


def build_step0002_row(objProductRow: ProductRow) -> list[str]:
    """処理0001の商品行を処理0002の14列へ変換します。"""
    return [
        "",
        "",
        "",
        "",
        objProductRow.product_code,
        objProductRow.product_code,
        objProductRow.product_name,
        "",
        objProductRow.spec,
        "",
        "",
        "",
        "",
        "",
    ]


def save_step0002_excel_template(
    objOutputPath: Path, listProductRows: list[ProductRow]
) -> None:
    """処理0002の14列を持つ新規Excelテンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(STEP0002_HEADERS))
    for objProductRow in listProductRows:
        objWorksheet.append(build_step0002_row(objProductRow))
        iOutputRow: int = objWorksheet.max_row
        objWorksheet.cell(row=iOutputRow, column=5).number_format = "@"
        objWorksheet.cell(row=iOutputRow, column=6).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_step0002_tsv_template(
    objOutputPath: Path, listProductRows: list[ProductRow]
) -> None:
    """処理0002の14列TSVをUTF-8 BOMなし、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(STEP0002_HEADERS)
        for objProductRow in listProductRows:
            objWriter.writerow(build_step0002_row(objProductRow))


def validate_step0001_outputs_match(
    listExcelRows: list[ProductRow], listTsvRows: list[ProductRow]
) -> None:
    """処理0001のXLSXとTSVの行数・行順・各値が一致することを確認します。"""
    if len(listExcelRows) != len(listTsvRows):
        raise ValueError(
            "step0001のXLSXとTSVの内容が一致しません。データ行数: XLSX = "
            + str(len(listExcelRows))
            + "、TSV = "
            + str(len(listTsvRows))
        )
    tupleColumns: tuple[tuple[str, str], ...] = (
        ("productCode", "product_code"),
        ("productName", "product_name"),
        ("spec", "spec"),
    )
    for iRow, (objExcelRow, objTsvRow) in enumerate(
        zip(listExcelRows, listTsvRows), start=2
    ):
        for pszColumnName, pszAttributeName in tupleColumns:
            pszExcelValue: str = getattr(objExcelRow, pszAttributeName)
            pszTsvValue: str = getattr(objTsvRow, pszAttributeName)
            if pszExcelValue != pszTsvValue:
                raise ValueError(
                    "step0001のXLSXとTSVの内容が一致しません。行 = "
                    + str(iRow)
                    + "、列 = "
                    + pszColumnName
                    + "、XLSX = "
                    + repr(pszExcelValue)
                    + "、TSV = "
                    + repr(pszTsvValue)
                )


def normalize_template_row(listValues: list[object]) -> list[str]:
    """処理0002の14列を、空セルを空文字にした比較・出力用文字列へ変換します。"""
    listNormalizedValues: list[str] = []
    for iColumn, objValue in enumerate(listValues[: len(STEP0002_HEADERS)]):
        if iColumn == 0 and isinstance(objValue, datetime):
            listNormalizedValues.append(objValue.date().strftime("%Y/%m/%d"))
        elif iColumn == 0 and isinstance(objValue, date):
            listNormalizedValues.append(objValue.strftime("%Y/%m/%d"))
        else:
            listNormalizedValues.append(normalize_text(objValue))
    return listNormalizedValues


def read_step0002_excel_rows(
    objExcelPath: Path, pszStepName: str = "step0002"
) -> list[list[str]]:
    """14列テンプレートExcelを検証してデータ行を読み取ります。"""
    objWorkbook: Workbook = load_workbook(objExcelPath, data_only=True)
    listTargetWorksheets: list[Worksheet] = []
    for objWorksheet in objWorkbook.worksheets:
        tupleHeaders: tuple[str, ...] = tuple(
            normalize_header(objWorksheet.cell(row=1, column=iColumn).value)
            for iColumn in range(1, objWorksheet.max_column + 1)
        )
        if tupleHeaders == STEP0002_HEADERS:
            listTargetWorksheets.append(objWorksheet)
    if len(listTargetWorksheets) == 0:
        raise ValueError(
            pszStepName + "の14列ヘッダーを持つExcelシートが見つかりません。"
        )
    if len(listTargetWorksheets) > 1:
        raise ValueError(
            pszStepName
            + "の対象シートが複数見つかりました。対象シート = "
            + ", ".join(objWorksheet.title for objWorksheet in listTargetWorksheets)
        )
    objWorksheet: Worksheet = listTargetWorksheets[0]
    listRows: list[list[str]] = []
    for iRow in range(2, objWorksheet.max_row + 1):
        listValues: list[object] = [
            objWorksheet.cell(row=iRow, column=iColumn).value
            for iColumn in range(1, len(STEP0002_HEADERS) + 1)
        ]
        listNormalizedRow: list[str] = normalize_template_row(listValues)
        if all(pszValue.strip() == "" for pszValue in listNormalizedRow):
            continue
        listRows.append(listNormalizedRow)
    return listRows


def read_step0002_tsv_rows(
    objTsvPath: Path, pszStepName: str = "step0002"
) -> list[list[str]]:
    """14列テンプレートのUTF-8 TSVを検証してデータ行を読み取ります。"""
    with objTsvPath.open(mode="r", encoding="utf-8", newline="") as objFile:
        listRows: list[list[str]] = list(
            csv.reader(objFile, delimiter="\t", strict=True)
        )
    if not listRows:
        raise ValueError(pszStepName + "のTSVファイルが空です。")
    if tuple(normalize_header(pszValue) for pszValue in listRows[0]) != STEP0002_HEADERS:
        raise ValueError(
            pszStepName + "のTSVヘッダーが仕様どおりの14列ではありません。"
        )
    listNormalizedRows: list[list[str]] = []
    for iRow, listValues in enumerate(listRows[1:], start=2):
        if len(listValues) != len(STEP0002_HEADERS):
            raise ValueError(
                str(iRow)
                + "行目の列数が14列ではありません。列数 = "
                + str(len(listValues))
            )
        listNormalizedRow: list[str] = normalize_template_row(listValues)
        if all(pszValue.strip() == "" for pszValue in listNormalizedRow):
            continue
        listNormalizedRows.append(listNormalizedRow)
    return listNormalizedRows


def validate_step0002_outputs_match(
    listExcelRows: list[list[str]],
    listTsvRows: list[list[str]],
    pszStepName: str = "step0002",
) -> None:
    """14列XLSX・TSVの行数、行順、すべての値が一致するか確認します。"""
    if len(listExcelRows) != len(listTsvRows):
        raise ValueError(
            pszStepName
            + "のXLSXとTSVの内容が一致しません。データ行数: XLSX = "
            + str(len(listExcelRows))
            + "、TSV = "
            + str(len(listTsvRows))
        )
    for iRow, (listExcelRow, listTsvRow) in enumerate(
        zip(listExcelRows, listTsvRows), start=2
    ):
        for iColumn, pszColumnName in enumerate(STEP0002_HEADERS):
            if listExcelRow[iColumn] != listTsvRow[iColumn]:
                raise ValueError(
                    pszStepName
                    + "のXLSXとTSVの内容が一致しません。行 = "
                    + str(iRow)
                    + "、列 = "
                    + pszColumnName
                    + "、XLSX = "
                    + repr(listExcelRow[iColumn])
                    + "、TSV = "
                    + repr(listTsvRow[iColumn])
                )


def build_step0003_rows(listStep0002Rows: list[list[str]]) -> list[list[str]]:
    """処理0002の商品1行を月～日の7行へ展開します。"""
    listStep0003Rows: list[list[str]] = []
    for listStep0002Row in listStep0002Rows:
        listMondayRow: list[str] = listStep0002Row.copy()
        listMondayRow[1] = WEEKDAYS[0]
        listStep0003Rows.append(listMondayRow)
        for pszWeekday in WEEKDAYS[1:]:
            listWeekdayRow: list[str] = [""] * len(STEP0002_HEADERS)
            listWeekdayRow[1] = pszWeekday
            listStep0003Rows.append(listWeekdayRow)
    return listStep0003Rows


def save_step0003_excel_template(
    objOutputPath: Path, listStep0003Rows: list[list[str]]
) -> None:
    """月～日に展開した処理0003のExcelテンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(STEP0002_HEADERS))
    for listValues in listStep0003Rows:
        objWorksheet.append(listValues)
        if listValues[1] == WEEKDAYS[0]:
            iOutputRow: int = objWorksheet.max_row
            objWorksheet.cell(row=iOutputRow, column=5).number_format = "@"
            objWorksheet.cell(row=iOutputRow, column=6).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_step0003_tsv_template(
    objOutputPath: Path, listStep0003Rows: list[list[str]]
) -> None:
    """月～日に展開した処理0003のTSVをUTF-8 BOMなし、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(STEP0002_HEADERS)
        objWriter.writerows(listStep0003Rows)


def validate_step0003_weekday_cycle(listStep0003Rows: list[list[str]]) -> None:
    """処理0003が7行周期で月～日の順になっていることを確認します。"""
    if len(listStep0003Rows) % len(WEEKDAYS) != 0:
        raise ValueError(
            "step0003のデータ行数が7の倍数ではありません。データ行数 = "
            + str(len(listStep0003Rows))
        )
    for iRowIndex, listValues in enumerate(listStep0003Rows):
        pszExpectedWeekday: str = WEEKDAYS[iRowIndex % len(WEEKDAYS)]
        pszActualWeekday: str = listValues[1]
        if pszActualWeekday != pszExpectedWeekday:
            raise ValueError(
                "step0003の曜日順が仕様と一致しません。行 = "
                + str(iRowIndex + 2)
                + "、期待値 = "
                + pszExpectedWeekday
                + "、実際値 = "
                + pszActualWeekday
            )


def build_step0004_rows(
    listStep0003Rows: list[list[str]], objStartMonday: date
) -> list[list[str]]:
    """7行ごとに同じ月～日の年月日を納品日列へ設定します。"""
    validate_start_monday(objStartMonday)
    validate_step0003_weekday_cycle(listStep0003Rows)
    listStep0004Rows: list[list[str]] = []
    for iRowIndex, listStep0003Row in enumerate(listStep0003Rows):
        objDeliveryDate: date = objStartMonday + timedelta(
            days=iRowIndex % len(WEEKDAYS)
        )
        listStep0004Row: list[str] = listStep0003Row.copy()
        listStep0004Row[0] = objDeliveryDate.strftime("%Y/%m/%d")
        listStep0004Rows.append(listStep0004Row)
    return listStep0004Rows


def save_step0004_excel_template(
    objOutputPath: Path, listStep0004Rows: list[list[str]]
) -> None:
    """納品日をExcel日付値として持つ処理0004テンプレートを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    objWorksheet.append(list(STEP0002_HEADERS))
    for listValues in listStep0004Rows:
        listExcelValues: list[object] = listValues.copy()
        listExcelValues[0] = datetime.strptime(listValues[0], "%Y/%m/%d").date()
        objWorksheet.append(listExcelValues)
        iOutputRow: int = objWorksheet.max_row
        objWorksheet.cell(row=iOutputRow, column=1).number_format = "yyyy/mm/dd"
        if listValues[1] == WEEKDAYS[0]:
            objWorksheet.cell(row=iOutputRow, column=5).number_format = "@"
            objWorksheet.cell(row=iOutputRow, column=6).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_step0004_tsv_template(
    objOutputPath: Path, listStep0004Rows: list[list[str]]
) -> None:
    """処理0004のTSVをUTF-8 BOMなし、タブ区切り、CRLFで保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow(STEP0002_HEADERS)
        objWriter.writerows(listStep0004Rows)


def validate_store_definitions() -> None:
    """店舗定義が100件で、店舗番号が一意かつ端点が仕様どおりか確認します。"""
    if len(STORE_DEFINITIONS) != 100:
        raise ValueError(
            "店舗定義数が100件ではありません。店舗定義数 = "
            + str(len(STORE_DEFINITIONS))
        )
    listStoreCodes: list[str] = [pszCode for pszCode, _ in STORE_DEFINITIONS]
    if len(set(listStoreCodes)) != len(listStoreCodes):
        raise ValueError("店舗番号が重複しています。")
    if STORE_DEFINITIONS[0] != ("2", "光の森"):
        raise ValueError("先頭の店舗定義が2／光の森ではありません。")
    if STORE_DEFINITIONS[-1] != ("151", "五日市"):
        raise ValueError("末尾の店舗定義が151／五日市ではありません。")


def build_step0005_random_seed(
    pszInputFileFullPath: str, objStartMonday: date
) -> int:
    """入力ベース名と開始月曜日から再現可能なSHA-256整数シードを作ります。"""
    pszSeedText: str = (
        Path(pszInputFileFullPath).stem + "|" + objStartMonday.isoformat()
    )
    objDigest: bytes = hashlib.sha256(pszSeedText.encode("utf-8")).digest()
    return int.from_bytes(objDigest, byteorder="big", signed=False)


def validate_product_delete_probability() -> None:
    """商品削除確率が0.0～1.0の有限数か確認します。"""
    if isinstance(PRODUCT_DELETE_PROBABILITY, bool) or not isinstance(
        PRODUCT_DELETE_PROBABILITY, (int, float)
    ):
        raise ValueError("商品削除確率が数値ではありません。")
    if not math.isfinite(PRODUCT_DELETE_PROBABILITY):
        raise ValueError("商品削除確率が有限数ではありません。")
    if not 0.0 <= PRODUCT_DELETE_PROBABILITY <= 1.0:
        raise ValueError(
            "商品削除確率は0.0～1.0で指定してください。Value = "
            + str(PRODUCT_DELETE_PROBABILITY)
        )


def build_product_delete_random_seed(
    pszInputFileFullPath: str, objStartMonday: date
) -> int:
    """入力名と開始月曜日から商品削除専用のSHA-256シードを作ります。"""
    pszSeedText: str = (
        Path(pszInputFileFullPath).stem
        + "|"
        + objStartMonday.isoformat()
        + "|product-deletion"
    )
    objDigest: bytes = hashlib.sha256(pszSeedText.encode("utf-8")).digest()
    return int.from_bytes(objDigest, byteorder="big", signed=False)


def validate_step0005_product_groups(listStep0004Rows: list[list[str]]) -> None:
    """処理0004のデータが商品ごとの月～日7行になっているか検証します。"""
    validate_step0003_weekday_cycle(listStep0004Rows)
    for iGroupStart in range(0, len(listStep0004Rows), len(WEEKDAYS)):
        listGroup: list[list[str]] = listStep0004Rows[
            iGroupStart:iGroupStart + len(WEEKDAYS)
        ]
        try:
            objMonday: date = datetime.strptime(listGroup[0][0], "%Y/%m/%d").date()
        except ValueError as objException:
            raise ValueError(
                "step0004の月曜日日付が不正です。行 = "
                + str(iGroupStart + 2)
            ) from objException
        if objMonday.weekday() != 0:
            raise ValueError(
                "step0004の開始日が月曜日ではありません。行 = "
                + str(iGroupStart + 2)
            )
        for iDay, listRow in enumerate(listGroup):
            try:
                objActualDate: date = datetime.strptime(listRow[0], "%Y/%m/%d").date()
            except ValueError as objException:
                raise ValueError(
                    "step0004の日付が不正です。行 = "
                    + str(iGroupStart + iDay + 2)
                ) from objException
            if objActualDate != objMonday + timedelta(days=iDay):
                raise ValueError(
                    "step0004の日付が月～日の連続日付ではありません。行 = "
                    + str(iGroupStart + iDay + 2)
                )
        for iColumn, pszColumnName in ((4, "Ｐ品番"), (5, "APEX品番"), (6, "商品名")):
            if not listGroup[0][iColumn].strip():
                raise ValueError(
                    "step0004の月曜日行の"
                    + pszColumnName
                    + "が空欄です。行 = "
                    + str(iGroupStart + 2)
                )


def select_step0005_product_rows(
    listStep0004Rows: list[list[str]],
    pszInputFileFullPath: str,
    objStartMonday: date,
) -> tuple[list[list[str]], int, int]:
    """商品ごとに独立抽選し、残す商品の月～日7行を返します。"""
    validate_product_delete_probability()
    validate_step0005_product_groups(listStep0004Rows)
    objDeleteRandom = random.Random(
        build_product_delete_random_seed(pszInputFileFullPath, objStartMonday)
    )
    listRetainedRows: list[list[str]] = []
    iKeptProductCount: int = 0
    iRemovedProductCount: int = 0
    for iGroupStart in range(0, len(listStep0004Rows), len(WEEKDAYS)):
        listGroup = listStep0004Rows[iGroupStart:iGroupStart + len(WEEKDAYS)]
        if objDeleteRandom.random() < PRODUCT_DELETE_PROBABILITY:
            iRemovedProductCount += 1
        else:
            listRetainedRows.extend(listRow.copy() for listRow in listGroup)
            iKeptProductCount += 1
    return listRetainedRows, iKeptProductCount, iRemovedProductCount


def generate_store_random_value(objRandom: random.Random) -> int | None:
    """空欄70%、1が20%、2～5が各2.5%となる値を1つ抽選します。"""
    iRandomBucket: int = objRandom.randint(1, 40)
    if iRandomBucket <= 28:
        return None
    if iRandomBucket <= 36:
        return 1
    return iRandomBucket - 35


def build_step0005_rows(
    listStep0004Rows: list[list[str]],
    pszInputFileFullPath: str,
    objStartMonday: date,
) -> list[list[object]]:
    """処理0004の14列へ、独立抽選した100店舗分の値を追加します。"""
    validate_store_definitions()
    objRandom = random.Random(
        build_step0005_random_seed(pszInputFileFullPath, objStartMonday)
    )
    listStep0005Rows: list[list[object]] = []
    iExpectedColumnCount: int = len(STEP0002_HEADERS) + len(STORE_DEFINITIONS)
    for listStep0004Row in listStep0004Rows:
        if len(listStep0004Row) != len(STEP0002_HEADERS):
            raise ValueError(
                "step0004のデータ行が14列ではありません。列数 = "
                + str(len(listStep0004Row))
            )
        listStoreValues: list[int | None] = [
            generate_store_random_value(objRandom) for _ in STORE_DEFINITIONS
        ]
        listStep0005Row: list[object] = listStep0004Row.copy() + listStoreValues
        if len(listStep0005Row) != iExpectedColumnCount:
            raise ValueError(
                "step0005のデータ行が114列ではありません。列数 = "
                + str(len(listStep0005Row))
            )
        listStep0005Rows.append(listStep0005Row)
    return listStep0005Rows


def save_step0005_excel_template(
    objOutputPath: Path, listStep0005Rows: list[list[object]]
) -> None:
    """2行ヘッダーと100店舗列を持つ処理0005のExcelを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    listStoreCodes: list[str] = [pszCode for pszCode, _ in STORE_DEFINITIONS]
    listStoreNames: list[str] = [pszName for _, pszName in STORE_DEFINITIONS]
    objWorksheet.append([""] * len(STEP0002_HEADERS) + listStoreCodes)
    objWorksheet.append(list(STEP0002_HEADERS) + listStoreNames)
    for iColumn in range(len(STEP0002_HEADERS) + 1, len(STEP0002_HEADERS) + 101):
        objWorksheet.cell(row=1, column=iColumn).number_format = "@"
    for listValues in listStep0005Rows:
        listExcelValues: list[object] = listValues.copy()
        listExcelValues[0] = datetime.strptime(str(listValues[0]), "%Y/%m/%d").date()
        objWorksheet.append(listExcelValues)
        iOutputRow: int = objWorksheet.max_row
        objWorksheet.cell(row=iOutputRow, column=1).number_format = "yyyy/mm/dd"
        objWorksheet.cell(row=iOutputRow, column=5).number_format = "@"
        objWorksheet.cell(row=iOutputRow, column=6).number_format = "@"
    if objWorksheet.max_column != 114:
        raise ValueError(
            "step0005 Excelの最終列がDJ列ではありません。列数 = "
            + str(objWorksheet.max_column)
        )
    objWorkbook.save(objOutputPath)


def save_step0005_tsv_template(
    objOutputPath: Path, listStep0005Rows: list[list[object]]
) -> None:
    """処理0005の114列TSVをUTF-8 BOMなし、CRLFで保存します。"""
    listStoreCodes: list[str] = [pszCode for pszCode, _ in STORE_DEFINITIONS]
    listStoreNames: list[str] = [pszName for _, pszName in STORE_DEFINITIONS]
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        objWriter = csv.writer(objFile, delimiter="\t", lineterminator="\r\n")
        objWriter.writerow([""] * len(STEP0002_HEADERS) + listStoreCodes)
        objWriter.writerow(list(STEP0002_HEADERS) + listStoreNames)
        for listValues in listStep0005Rows:
            if len(listValues) != 114:
                raise ValueError(
                    "step0005 TSVのデータ行が114列ではありません。列数 = "
                    + str(len(listValues))
                )
            objWriter.writerow(listValues)


def replace_output_files(
    objTemporaryExcelPath: Path,
    objTemporaryTsvPath: Path,
    objExcelOutputPath: Path,
    objTsvOutputPath: Path,
) -> None:
    """2出力を置換し、失敗時は可能な限り以前の状態へ戻します。"""
    listOutputPaths: list[Path] = [objExcelOutputPath, objTsvOutputPath]
    listTemporaryPaths: list[Path] = [objTemporaryExcelPath, objTemporaryTsvPath]
    dictBackupPaths: dict[Path, Path] = {}
    listReplacedPaths: list[Path] = []
    try:
        for objOutputPath in listOutputPaths:
            if objOutputPath.exists():
                objBackupPath: Path = create_temporary_path(objOutputPath, ".backup")
                shutil.copy2(objOutputPath, objBackupPath)
                dictBackupPaths[objOutputPath] = objBackupPath
        for objTemporaryPath, objOutputPath in zip(listTemporaryPaths, listOutputPaths):
            os.replace(objTemporaryPath, objOutputPath)
            listReplacedPaths.append(objOutputPath)
    except Exception:
        for objOutputPath in reversed(listReplacedPaths):
            objBackupPath = dictBackupPaths.get(objOutputPath)
            if objBackupPath is not None and objBackupPath.exists():
                os.replace(objBackupPath, objOutputPath)
            elif objOutputPath.exists():
                objOutputPath.unlink()
        raise
    finally:
        for objBackupPath in dictBackupPaths.values():
            if objBackupPath.exists():
                objBackupPath.unlink()
        for objTemporaryPath in listTemporaryPaths:
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()


def process_step0002_files(
    pszInputFileFullPath: str,
    objStep0001ExcelPath: Path,
    objStep0001TsvPath: Path,
) -> tuple[Path, Path, int]:
    """処理0001の両出力を比較し、処理0002のXLSXとTSVを作成します。"""
    try:
        listExcelRows: list[ProductRow] = read_excel_rows(str(objStep0001ExcelPath))
        listTsvRows: list[ProductRow] = read_delimited_rows(str(objStep0001TsvPath))
        validate_step0001_outputs_match(listExcelRows, listTsvRows)

        objExcelOutputPath, objTsvOutputPath = get_step0002_output_file_paths(
            pszInputFileFullPath
        )
        objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
        objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
        try:
            save_step0002_excel_template(objTemporaryExcelPath, listExcelRows)
            save_step0002_tsv_template(objTemporaryTsvPath, listExcelRows)
            replace_output_files(
                objTemporaryExcelPath,
                objTemporaryTsvPath,
                objExcelOutputPath,
                objTsvOutputPath,
            )
        finally:
            for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return objExcelOutputPath, objTsvOutputPath, len(listExcelRows)
    except Exception as objException:
        raise Step0002Error(str(objException)) from objException


def process_step0003_files(
    pszInputFileFullPath: str,
    objStep0002ExcelPath: Path,
    objStep0002TsvPath: Path,
) -> tuple[Path, Path, int, int]:
    """処理0002の両出力を比較し、月～日に展開した処理0003を作成します。"""
    try:
        listExcelRows: list[list[str]] = read_step0002_excel_rows(
            objStep0002ExcelPath
        )
        listTsvRows: list[list[str]] = read_step0002_tsv_rows(objStep0002TsvPath)
        validate_step0002_outputs_match(listExcelRows, listTsvRows)
        listStep0003Rows: list[list[str]] = build_step0003_rows(listExcelRows)

        objExcelOutputPath, objTsvOutputPath = get_step0003_output_file_paths(
            pszInputFileFullPath
        )
        objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
        objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
        try:
            save_step0003_excel_template(objTemporaryExcelPath, listStep0003Rows)
            save_step0003_tsv_template(objTemporaryTsvPath, listStep0003Rows)
            replace_output_files(
                objTemporaryExcelPath,
                objTemporaryTsvPath,
                objExcelOutputPath,
                objTsvOutputPath,
            )
        finally:
            for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return (
            objExcelOutputPath,
            objTsvOutputPath,
            len(listExcelRows),
            len(listStep0003Rows),
        )
    except Exception as objException:
        raise Step0003Error(str(objException)) from objException


def process_step0004_files(
    pszInputFileFullPath: str,
    objStep0003ExcelPath: Path,
    objStep0003TsvPath: Path,
    objStartMonday: date,
) -> tuple[Path, Path, int]:
    """処理0003の両出力を比較し、同じ1週間を繰り返す処理0004を作成します。"""
    try:
        listExcelRows: list[list[str]] = read_step0002_excel_rows(
            objStep0003ExcelPath, "step0003"
        )
        listTsvRows: list[list[str]] = read_step0002_tsv_rows(
            objStep0003TsvPath, "step0003"
        )
        validate_step0002_outputs_match(listExcelRows, listTsvRows, "step0003")
        listStep0004Rows: list[list[str]] = build_step0004_rows(
            listExcelRows, objStartMonday
        )

        objExcelOutputPath, objTsvOutputPath = get_step0004_output_file_paths(
            pszInputFileFullPath
        )
        objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
        objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
        try:
            save_step0004_excel_template(objTemporaryExcelPath, listStep0004Rows)
            save_step0004_tsv_template(objTemporaryTsvPath, listStep0004Rows)
            replace_output_files(
                objTemporaryExcelPath,
                objTemporaryTsvPath,
                objExcelOutputPath,
                objTsvOutputPath,
            )
        finally:
            for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return objExcelOutputPath, objTsvOutputPath, len(listStep0004Rows)
    except Exception as objException:
        raise Step0004Error(str(objException)) from objException


def process_step0005_files(
    pszInputFileFullPath: str,
    objStep0004ExcelPath: Path,
    objStep0004TsvPath: Path,
    objStartMonday: date,
) -> tuple[Path, Path, int, int, int, Path | None]:
    """処理0004の両出力を比較し、100店舗列を持つ処理0005を作成します。"""
    try:
        listExcelRows: list[list[str]] = read_step0002_excel_rows(
            objStep0004ExcelPath, "step0004"
        )
        listTsvRows: list[list[str]] = read_step0002_tsv_rows(
            objStep0004TsvPath, "step0004"
        )
        validate_step0002_outputs_match(listExcelRows, listTsvRows, "step0004")
        listRetainedRows, iKeptProductCount, iRemovedProductCount = (
            select_step0005_product_rows(
                listExcelRows, pszInputFileFullPath, objStartMonday
            )
        )
        listStep0005Rows: list[list[object]] = build_step0005_rows(
            listRetainedRows, pszInputFileFullPath, objStartMonday
        )

        objExcelOutputPath, objTsvOutputPath = get_step0005_output_file_paths(
            pszInputFileFullPath
        )
        objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
        objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
        objWarningPath: Path = get_step0005_warning_file_path(pszInputFileFullPath)
        objTemporaryWarningPath: Path | None = None
        try:
            save_step0005_excel_template(objTemporaryExcelPath, listStep0005Rows)
            save_step0005_tsv_template(objTemporaryTsvPath, listStep0005Rows)
            if iKeptProductCount == 0 and iRemovedProductCount > 0:
                objTemporaryWarningPath = create_temporary_path(objWarningPath, ".txt")
                pszWarningMessage: str = (
                    "処理結果: 警告\n"
                    + "入力ファイル: "
                    + os.path.abspath(pszInputFileFullPath)
                    + "\n発生した処理: 旭注文テンプレート処理0005\n"
                    + "警告内容: 商品削除抽選の結果、すべての商品が削除されました。"
                    + "処理0005はヘッダー2行だけで作成しました。\n"
                    + "商品削除確率: "
                    + str(PRODUCT_DELETE_PROBABILITY)
                    + " ("
                    + str(PRODUCT_DELETE_PROBABILITY * 100)
                    + "%)\n入力商品数: "
                    + str(iKeptProductCount + iRemovedProductCount)
                    + "\n残存商品数: 0\n削除商品数: "
                    + str(iRemovedProductCount)
                )
                write_warning_text(str(objTemporaryWarningPath), pszWarningMessage)
                replace_output_file_set(
                    {
                        objExcelOutputPath: objTemporaryExcelPath,
                        objTsvOutputPath: objTemporaryTsvPath,
                        objWarningPath: objTemporaryWarningPath,
                    }
                )
            else:
                replace_output_files(
                    objTemporaryExcelPath,
                    objTemporaryTsvPath,
                    objExcelOutputPath,
                    objTsvOutputPath,
                )
                if objWarningPath.exists():
                    objWarningPath.unlink()
        finally:
            for objTemporaryPath in (
                objTemporaryExcelPath,
                objTemporaryTsvPath,
                objTemporaryWarningPath,
            ):
                if objTemporaryPath is None:
                    continue
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return (
            objExcelOutputPath,
            objTsvOutputPath,
            len(listStep0005Rows),
            iKeptProductCount,
            iRemovedProductCount,
            objWarningPath if iKeptProductCount == 0 and iRemovedProductCount > 0 else None,
        )
    except Exception as objException:
        raise Step0005Error(str(objException)) from objException


def get_default_mapping_file_path() -> Path:
    """プログラムと同じフォルダーの対応表パスを返します。"""
    return Path(__file__).resolve().parent / AREA_STORE_MAPPING_FILE_NAME


def read_area_store_mapping(
    objMappingPath: Path,
) -> list[tuple[str, str, str]]:
    """配送センター・店舗コード・店舗略称の対応表を読み込みます。"""
    if not objMappingPath.is_file():
        raise ValueError("対応表が見つかりません。Path = " + str(objMappingPath))
    listRows: list[list[str]] | None = None
    objLastException: UnicodeDecodeError | None = None
    for pszEncoding in ("utf-8-sig", "cp932"):
        try:
            with objMappingPath.open(mode="r", encoding=pszEncoding, newline="") as objFile:
                listRows = list(csv.reader(objFile, delimiter="\t", strict=True))
            break
        except UnicodeDecodeError as objException:
            objLastException = objException
    if listRows is None:
        raise ValueError("対応表の文字コードを判定できません。") from objLastException
    listNonBlankRows: list[list[str]] = [
        listRow for listRow in listRows if any(str(pszValue).strip() for pszValue in listRow)
    ]
    if not listNonBlankRows:
        raise ValueError("対応表が空です。")
    tupleExpectedHeaders: tuple[str, str, str] = (
        "配送センター名", "店舗コード", "店舗略称"
    )
    if tuple(pszValue.strip() for pszValue in listNonBlankRows[0]) != tupleExpectedHeaders:
        raise ValueError("対応表のヘッダーが仕様どおりではありません。")
    listMappings: list[tuple[str, str, str]] = []
    setStoreCodes: set[str] = set()
    for iRow, listRow in enumerate(listNonBlankRows[1:], start=2):
        if len(listRow) != 3:
            raise ValueError(
                f"対応表の{iRow}行目が3列ではありません。列数 = {len(listRow)}"
            )
        pszCenter, pszCode, pszName = (pszValue.strip() for pszValue in listRow)
        if not pszCenter or not pszCode or not pszName:
            raise ValueError(f"対応表の{iRow}行目に空欄があります。")
        if pszCode in setStoreCodes:
            raise ValueError("対応表の店舗コードが重複しています。Code = " + pszCode)
        setStoreCodes.add(pszCode)
        listMappings.append((pszCenter, pszCode, pszName))
    if not listMappings:
        raise ValueError("対応表に店舗定義がありません。")
    return listMappings


def normalize_step0005_cell(objValue: object, iColumn: int) -> str:
    """処理0005のセルをXLSX・TSV比較用文字列へ変換します。"""
    if iColumn == 0 and isinstance(objValue, datetime):
        return objValue.date().strftime("%Y/%m/%d")
    if iColumn == 0 and isinstance(objValue, date):
        return objValue.strftime("%Y/%m/%d")
    return normalize_text(objValue)


def validate_step0005_table(listRows: list[list[str]], pszSource: str) -> None:
    """処理0005の2行ヘッダー、店舗コード、列数を検証します。"""
    if len(listRows) < 2:
        raise ValueError(pszSource + "に2行ヘッダーがありません。")
    iColumnCount: int = len(listRows[0])
    if iColumnCount < len(STEP0002_HEADERS):
        raise ValueError(pszSource + "の列数が14列未満です。")
    for iRow, listRow in enumerate(listRows, start=1):
        if len(listRow) != iColumnCount:
            raise ValueError(
                f"{pszSource}の{iRow}行目の列数が一致しません。"
            )
    if any(pszValue.strip() for pszValue in listRows[0][: len(STEP0002_HEADERS)]):
        raise ValueError(pszSource + "の1行目A～N列が空欄ではありません。")
    if tuple(pszValue.strip() for pszValue in listRows[1][: len(STEP0002_HEADERS)]) != STEP0002_HEADERS:
        raise ValueError(pszSource + "の2行目A～N列が仕様どおりではありません。")
    listStoreCodes: list[str] = [
        pszValue.strip() for pszValue in listRows[0][len(STEP0002_HEADERS):]
    ]
    if any(not pszCode for pszCode in listStoreCodes):
        raise ValueError(pszSource + "の店舗コードに空欄があります。")
    if len(set(listStoreCodes)) != len(listStoreCodes):
        raise ValueError(pszSource + "の店舗コードが重複しています。")


def read_step0005_excel_table(objExcelPath: Path) -> list[list[str]]:
    """処理0005 Excelの2行ヘッダーとデータを読み込みます。"""
    objWorkbook: Workbook = load_workbook(objExcelPath, data_only=True)
    if len(objWorkbook.worksheets) != 1:
        raise ValueError("step0005 Excelのシート数が1ではありません。")
    objWorksheet: Worksheet = objWorkbook.active
    listRows: list[list[str]] = [
        [
            normalize_step0005_cell(objWorksheet.cell(iRow, iColumn).value, iColumn - 1)
            for iColumn in range(1, objWorksheet.max_column + 1)
        ]
        for iRow in range(1, objWorksheet.max_row + 1)
    ]
    validate_step0005_table(listRows, "step0005 Excel")
    return listRows


def read_step0005_tsv_table(objTsvPath: Path) -> list[list[str]]:
    """処理0005 TSVの2行ヘッダーとデータを読み込みます。"""
    with objTsvPath.open(mode="r", encoding="utf-8", newline="") as objFile:
        listRawRows: list[list[str]] = list(csv.reader(objFile, delimiter="\t", strict=True))
    listRows: list[list[str]] = [
        [normalize_step0005_cell(pszValue, iColumn) for iColumn, pszValue in enumerate(listRow)]
        for listRow in listRawRows
    ]
    validate_step0005_table(listRows, "step0005 TSV")
    return listRows


def validate_step0005_tables_match(
    listExcelRows: list[list[str]], listTsvRows: list[list[str]]
) -> None:
    """処理0005のXLSXとTSVの全セルが一致することを確認します。"""
    if len(listExcelRows) != len(listTsvRows):
        raise ValueError("step0005のXLSXとTSVの行数が一致しません。")
    for iRow, (listExcelRow, listTsvRow) in enumerate(zip(listExcelRows, listTsvRows), start=1):
        if len(listExcelRow) != len(listTsvRow):
            raise ValueError(f"step0005の{iRow}行目の列数が一致しません。")
        for iColumn, (pszExcel, pszTsv) in enumerate(zip(listExcelRow, listTsvRow), start=1):
            if pszExcel != pszTsv:
                raise ValueError(
                    f"step0005のXLSXとTSVが一致しません。行 = {iRow}、列 = {iColumn}"
                )


def select_step0006_columns(
    listStep0005Rows: list[list[str]], setTargetStoreCodes: set[str]
) -> list[list[str]]:
    """A～N列と指定された店舗コードの列を元の順序で抽出します。"""
    listColumnIndexes: list[int] = list(range(len(STEP0002_HEADERS)))
    listColumnIndexes.extend(
        iColumn
        for iColumn, pszCode in enumerate(listStep0005Rows[0])
        if iColumn >= len(STEP0002_HEADERS) and pszCode.strip() in setTargetStoreCodes
    )
    return [[listRow[iColumn] for iColumn in listColumnIndexes] for listRow in listStep0005Rows]


def sanitize_delivery_center_filename(pszCenterName: str) -> str:
    """配送センター名をWindowsで安全なファイル名部分へ変換します。"""
    pszSafeName: str = pszCenterName.strip().replace("(", "_").replace("（", "_")
    pszSafeName = pszSafeName.replace(")", "").replace("）", "")
    pszSafeName = re.sub(r'[\\/:*?"<>|]', "_", pszSafeName)
    pszSafeName = re.sub(r"_+", "_", pszSafeName).strip(" ._")
    if not pszSafeName:
        raise ValueError("配送センター名をファイル名へ変換できません。")
    return pszSafeName


def save_step0006_excel_template(objOutputPath: Path, listRows: list[list[str]]) -> None:
    """抽出した2行ヘッダー形式の処理0006 Excelを保存します。"""
    objWorkbook: Workbook = Workbook()
    objWorksheet: Worksheet = objWorkbook.active
    for iRow, listRow in enumerate(listRows, start=1):
        listValues: list[object] = listRow.copy()
        if iRow >= 3 and listValues[0]:
            listValues[0] = datetime.strptime(str(listValues[0]), "%Y/%m/%d").date()
        for iColumn in range(len(STEP0002_HEADERS), len(listValues)):
            if iRow >= 3 and str(listValues[iColumn]).isdigit():
                listValues[iColumn] = int(str(listValues[iColumn]))
        objWorksheet.append(listValues)
        if iRow == 1:
            for iColumn in range(len(STEP0002_HEADERS) + 1, len(listValues) + 1):
                objWorksheet.cell(row=1, column=iColumn).number_format = "@"
        elif iRow >= 3:
            objWorksheet.cell(row=iRow, column=1).number_format = "yyyy/mm/dd"
            objWorksheet.cell(row=iRow, column=5).number_format = "@"
            objWorksheet.cell(row=iRow, column=6).number_format = "@"
    objWorkbook.save(objOutputPath)


def save_step0006_tsv_template(objOutputPath: Path, listRows: list[list[str]]) -> None:
    """抽出した2行ヘッダー形式の処理0006 TSVを保存します。"""
    with objOutputPath.open(mode="w", encoding="utf-8", newline="") as objFile:
        csv.writer(objFile, delimiter="\t", lineterminator="\r\n").writerows(listRows)


def replace_output_file_set(dictTemporaryOutputs: dict[Path, Path]) -> None:
    """処理0006の全出力を置換し、失敗時は以前の状態へ戻します。"""
    dictBackups: dict[Path, Path] = {}
    listReplaced: list[Path] = []
    try:
        for objOutputPath in dictTemporaryOutputs:
            if objOutputPath.exists():
                objBackupPath: Path = create_temporary_path(objOutputPath, ".backup")
                shutil.copy2(objOutputPath, objBackupPath)
                dictBackups[objOutputPath] = objBackupPath
        for objOutputPath, objTemporaryPath in dictTemporaryOutputs.items():
            os.replace(objTemporaryPath, objOutputPath)
            listReplaced.append(objOutputPath)
    except Exception:
        for objOutputPath in reversed(listReplaced):
            objBackupPath = dictBackups.get(objOutputPath)
            if objBackupPath is not None and objBackupPath.exists():
                os.replace(objBackupPath, objOutputPath)
            elif objOutputPath.exists():
                objOutputPath.unlink()
        raise
    finally:
        for objPath in [*dictBackups.values(), *dictTemporaryOutputs.values()]:
            if objPath.exists():
                objPath.unlink()


def build_step0006_all_stores_rows(
    listStep0005Rows: list[list[str]], listMappings: list[tuple[str, str, str]]
) -> list[list[str]]:
    """対応表にある全店舗を残した処理0006-1を構築します。"""
    return select_step0006_columns(
        listStep0005Rows, {pszCode for _, pszCode, _ in listMappings}
    )


def build_step0006_center_rows(
    listStep0005Rows: list[list[str]], setCenterStoreCodes: set[str]
) -> list[list[str]]:
    """1つの配送センターの店舗列を残した処理0006-2を構築します。"""
    return select_step0006_columns(listStep0005Rows, setCenterStoreCodes)


def center_has_order_quantity(listCenterRows: list[list[str]]) -> bool:
    """配送センターの店舗数量に1つでも値があるか確認します。"""
    return any(
        pszValue.strip()
        for listRow in listCenterRows[2:]
        for pszValue in listRow[len(STEP0002_HEADERS):]
    )


def process_step0006_files(
    pszInputFileFullPath: str,
    objStep0005ExcelPath: Path,
    objStep0005TsvPath: Path,
    objMappingPath: Path,
) -> tuple[list[Path], list[str]]:
    """処理0005と対応表から全店舗版と配送センター別の処理0006を作ります。"""
    try:
        listExcelRows = read_step0005_excel_table(objStep0005ExcelPath)
        listTsvRows = read_step0005_tsv_table(objStep0005TsvPath)
        validate_step0005_tables_match(listExcelRows, listTsvRows)
        listMappings = read_area_store_mapping(objMappingPath)
        setInputCodes: set[str] = set(listExcelRows[0][len(STEP0002_HEADERS):])
        setMappedInputCodes: set[str] = {
            pszCode for _, pszCode, _ in listMappings if pszCode in setInputCodes
        }
        if not setMappedInputCodes:
            raise ValueError("処理0005と対応表で一致する店舗コードがありません。")

        dictCenterCodes: dict[str, set[str]] = {}
        for pszCenter, pszCode, _ in listMappings:
            dictCenterCodes.setdefault(pszCenter, set()).add(pszCode)
        dictSafeNames: dict[str, str] = {
            pszCenter: sanitize_delivery_center_filename(pszCenter)
            for pszCenter in dictCenterCodes
        }
        if len(set(dictSafeNames.values())) != len(dictSafeNames):
            raise ValueError("配送センター名のファイル名が重複します。")

        objInputPath = Path(pszInputFileFullPath)
        objExcelOutputPath = objInputPath.with_name(objInputPath.stem + "_step0006.xlsx")
        objTsvOutputPath = objInputPath.with_name(objInputPath.stem + "_step0006.tsv")
        dictOutputRows: dict[Path, list[list[str]]] = {
            objExcelOutputPath: build_step0006_all_stores_rows(listExcelRows, listMappings),
        }
        dictOutputRows[objTsvOutputPath] = dictOutputRows[objExcelOutputPath]
        listWarnings: list[str] = []
        for pszCenter, setCodes in dictCenterCodes.items():
            listCenterRows = build_step0006_center_rows(listExcelRows, setCodes)
            if len(listCenterRows[0]) == len(STEP0002_HEADERS):
                listWarnings.append(
                    f'警告: 配送センター「{pszCenter}」の店舗コード列は処理0005にありません。'
                )
                continue
            if not center_has_order_quantity(listCenterRows):
                listWarnings.append(
                    f'警告: 配送センター「{pszCenter}」の店舗数量はすべて空欄ですが、XLSX・TSVを作成しました。'
                )
            pszCenterBaseName: str = (
                objInputPath.stem + "_step0006_" + dictSafeNames[pszCenter]
            )
            dictOutputRows[objInputPath.with_name(pszCenterBaseName + ".xlsx")] = listCenterRows
            dictOutputRows[objInputPath.with_name(pszCenterBaseName + ".tsv")] = listCenterRows

        setCenterOutputCodes: set[str] = set()
        for pszCenter, setCodes in dictCenterCodes.items():
            setCenterOutputCodes.update(setCodes & setInputCodes)
        if setCenterOutputCodes != setMappedInputCodes:
            raise ValueError("全店舗版と配送センター別版の店舗コードが一致しません。")

        dictTemporaryOutputs: dict[Path, Path] = {}
        try:
            for objOutputPath, listRows in dictOutputRows.items():
                objTemporaryPath = create_temporary_path(objOutputPath, objOutputPath.suffix)
                dictTemporaryOutputs[objOutputPath] = objTemporaryPath
                if objOutputPath.suffix == ".xlsx":
                    save_step0006_excel_template(objTemporaryPath, listRows)
                else:
                    save_step0006_tsv_template(objTemporaryPath, listRows)
            replace_output_file_set(dictTemporaryOutputs)
        finally:
            for objTemporaryPath in dictTemporaryOutputs.values():
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return list(dictOutputRows), listWarnings
    except Exception as objException:
        raise Step0006Error(str(objException)) from objException


def validate_step0007_product_groups(
    listRows: list[list[str]], pszCenterName: str
) -> None:
    """配送センター別データが商品ごとの月～日7行になっているか検証します。"""
    validate_step0005_table(listRows, "step0006「" + pszCenterName + "」")
    listDataRows: list[list[str]] = listRows[2:]
    if len(listDataRows) % len(WEEKDAYS) != 0:
        raise ValueError(
            f'配送センター「{pszCenterName}」のstep0006データ行数が7の倍数ではありません。'
        )
    for iGroupStart in range(0, len(listDataRows), len(WEEKDAYS)):
        listGroup: list[list[str]] = listDataRows[iGroupStart:iGroupStart + len(WEEKDAYS)]
        for iDay, (listRow, pszExpectedWeekday) in enumerate(zip(listGroup, WEEKDAYS)):
            iFileRow: int = iGroupStart + iDay + 3
            if listRow[1] != pszExpectedWeekday:
                raise ValueError(
                    f'配送センター「{pszCenterName}」の{iFileRow}行目の曜日が'
                    f'「{pszExpectedWeekday}」ではありません。'
                )
        try:
            objMonday: date = datetime.strptime(listGroup[0][0], "%Y/%m/%d").date()
        except ValueError as objException:
            raise ValueError(
                f'配送センター「{pszCenterName}」の月曜日日付が不正です。'
            ) from objException
        if objMonday.weekday() != 0:
            raise ValueError(
                f'配送センター「{pszCenterName}」の開始日が月曜日ではありません。'
            )
        for iDay, listRow in enumerate(listGroup):
            try:
                objActualDate: date = datetime.strptime(listRow[0], "%Y/%m/%d").date()
            except ValueError as objException:
                raise ValueError(
                    f'配送センター「{pszCenterName}」の日付が不正です。'
                ) from objException
            if objActualDate != objMonday + timedelta(days=iDay):
                raise ValueError(
                    f'配送センター「{pszCenterName}」の日付が月～日の連続日付ではありません。'
                )
        for iColumn, pszColumnName in ((4, "Ｐ品番"), (5, "APEX品番"), (6, "商品名")):
            if not listGroup[0][iColumn].strip():
                raise ValueError(
                    f'配送センター「{pszCenterName}」の月曜日行の{pszColumnName}が空欄です。'
                )


def product_week_has_order(listProductWeekRows: list[list[str]]) -> bool:
    """1商品分7行のO列以降に空欄以外の値があるか返します。"""
    return any(
        pszValue.strip() != ""
        for listRow in listProductWeekRows
        for pszValue in listRow[len(STEP0002_HEADERS):]
    )


def build_step0007_rows(
    listStep0006Rows: list[list[str]], pszCenterName: str
) -> tuple[list[list[str]], int, int]:
    """発注がある商品の月～日7行だけを残した処理0007を構築します。"""
    validate_step0007_product_groups(listStep0006Rows, pszCenterName)
    listOutputRows: list[list[str]] = [
        listStep0006Rows[0].copy(), listStep0006Rows[1].copy()
    ]
    iKeptProductCount: int = 0
    iRemovedProductCount: int = 0
    listDataRows: list[list[str]] = listStep0006Rows[2:]
    for iGroupStart in range(0, len(listDataRows), len(WEEKDAYS)):
        listGroup = listDataRows[iGroupStart:iGroupStart + len(WEEKDAYS)]
        if product_week_has_order(listGroup):
            listOutputRows.extend(listRow.copy() for listRow in listGroup)
            iKeptProductCount += 1
        else:
            iRemovedProductCount += 1
    return listOutputRows, iKeptProductCount, iRemovedProductCount


def build_product_step0007_rows(
    listStep0006Rows: list[list[str]], pszCenterName: str
) -> tuple[list[list[str]], int, int]:
    """商品別処理0006から発注のある店舗列だけを残した処理0007を構築します。"""
    validate_step0007_product_groups(listStep0006Rows, pszCenterName)
    if len(listStep0006Rows) != 2 + len(WEEKDAYS):
        raise ValueError(
            f'配送センター「{pszCenterName}」の商品別step0006が9行ではありません。'
        )
    listDataRows: list[list[str]] = listStep0006Rows[2:]
    listOutputColumnIndexes: list[int] = list(range(len(STEP0002_HEADERS)))
    iKeptStoreCount: int = 0
    iRemovedStoreCount: int = 0
    for iColumn in range(len(STEP0002_HEADERS), len(listStep0006Rows[0])):
        if any(listRow[iColumn].strip() != "" for listRow in listDataRows):
            listOutputColumnIndexes.append(iColumn)
            iKeptStoreCount += 1
        else:
            iRemovedStoreCount += 1
    listOutputRows: list[list[str]] = [
        [listRow[iColumn] for iColumn in listOutputColumnIndexes]
        for listRow in listStep0006Rows
    ]
    return listOutputRows, iKeptStoreCount, iRemovedStoreCount


def get_step0007_output_path(objStep0006Path: Path) -> Path:
    """配送センター別step0006パスからstep0007パスを作ります。"""
    pszMarker: str = "_step0006_"
    if pszMarker not in objStep0006Path.stem:
        raise ValueError(
            "配送センター別step0006のファイル名ではありません。Path = "
            + str(objStep0006Path)
        )
    pszOutputStem: str = objStep0006Path.stem.replace(pszMarker, "_step0007_", 1)
    return objStep0006Path.with_name(pszOutputStem + objStep0006Path.suffix)


def process_step0007_files(
    pszInputFileFullPath: str, listStep0006OutputPaths: list[Path]
) -> tuple[list[Path], list[str]]:
    """配送センター別step0006から発注のある商品だけのstep0007を作ります。"""
    try:
        objInputPath: Path = Path(pszInputFileFullPath)
        pszAllStoresStem: str = objInputPath.stem + "_step0006"
        dictCenterInputPaths: dict[str, dict[str, Path]] = {}
        for objPath in listStep0006OutputPaths:
            if objPath.stem == pszAllStoresStem:
                continue
            pszMarker: str = "_step0006_"
            if pszMarker not in objPath.stem:
                continue
            pszCenterName: str = objPath.stem.split(pszMarker, 1)[1]
            dictCenterInputPaths.setdefault(pszCenterName, {})[objPath.suffix] = objPath
        if not dictCenterInputPaths:
            raise ValueError("処理0007の対象となる配送センター別step0006がありません。")

        dictOutputRows: dict[Path, list[list[str]]] = {}
        listWarnings: list[str] = []
        for pszCenterName, dictPaths in dictCenterInputPaths.items():
            if set(dictPaths) != {".xlsx", ".tsv"}:
                raise ValueError(
                    f'配送センター「{pszCenterName}」のstep0006 XLSX・TSVが揃っていません。'
                )
            listExcelRows = read_step0005_excel_table(dictPaths[".xlsx"])
            listTsvRows = read_step0005_tsv_table(dictPaths[".tsv"])
            validate_step0005_tables_match(listExcelRows, listTsvRows)
            listStep0007Rows, iKeptCount, iRemovedCount = build_step0007_rows(
                listExcelRows, pszCenterName
            )
            if iKeptCount == 0:
                listWarnings.append(
                    f'警告: 配送センター「{pszCenterName}」には発注のある商品がありませんが、'
                    "ヘッダー2行のXLSX・TSVを作成しました。"
                )
            print(
                f'Step0007 Center: {pszCenterName}, Kept Products: {iKeptCount}, '
                f'Removed Products: {iRemovedCount}'
            )
            objExcelOutputPath = get_step0007_output_path(dictPaths[".xlsx"])
            objTsvOutputPath = get_step0007_output_path(dictPaths[".tsv"])
            dictOutputRows[objExcelOutputPath] = listStep0007Rows
            dictOutputRows[objTsvOutputPath] = listStep0007Rows

        dictTemporaryOutputs: dict[Path, Path] = {}
        try:
            for objOutputPath, listRows in dictOutputRows.items():
                objTemporaryPath = create_temporary_path(objOutputPath, objOutputPath.suffix)
                dictTemporaryOutputs[objOutputPath] = objTemporaryPath
                if objOutputPath.suffix == ".xlsx":
                    save_step0006_excel_template(objTemporaryPath, listRows)
                else:
                    save_step0006_tsv_template(objTemporaryPath, listRows)
            for pszCenterName, dictPaths in dictCenterInputPaths.items():
                objFinalExcelPath = get_step0007_output_path(dictPaths[".xlsx"])
                objFinalTsvPath = get_step0007_output_path(dictPaths[".tsv"])
                listSavedExcelRows = read_step0005_excel_table(
                    dictTemporaryOutputs[objFinalExcelPath]
                )
                listSavedTsvRows = read_step0005_tsv_table(
                    dictTemporaryOutputs[objFinalTsvPath]
                )
                validate_step0005_tables_match(listSavedExcelRows, listSavedTsvRows)
                validate_step0007_product_groups(listSavedExcelRows, pszCenterName)
                for iStart in range(2, len(listSavedExcelRows), len(WEEKDAYS)):
                    if not product_week_has_order(
                        listSavedExcelRows[iStart:iStart + len(WEEKDAYS)]
                    ):
                        raise ValueError(
                            f'配送センター「{pszCenterName}」のstep0007に発注のない商品が残っています。'
                        )
            replace_output_file_set(dictTemporaryOutputs)
        finally:
            for objTemporaryPath in dictTemporaryOutputs.values():
                if objTemporaryPath.exists():
                    objTemporaryPath.unlink()
        return list(dictOutputRows), listWarnings
    except Exception as objException:
        raise Step0007Error(str(objException)) from objException


def sanitize_product_filename_part(pszValue: str, pszColumnName: str, iRow: int) -> str:
    """商品のAPEX品番または商品名をWindowsで安全なファイル名部分へ変換します。"""
    pszTrimmedValue: str = pszValue.strip()
    if not pszTrimmedValue:
        raise ValueError(f"step0002の{iRow}行目の{pszColumnName}が空欄です。")
    pszSafeValue: str = re.sub(r'[\\/:*?"<>|\x00-\x1f]', "_", pszTrimmedValue)
    pszSafeValue = re.sub(r"_+", "_", pszSafeValue).rstrip(" .")
    if not pszSafeValue:
        raise ValueError(
            f"step0002の{iRow}行目の{pszColumnName}をファイル名へ変換できません。"
        )
    return pszSafeValue


def build_product_file_plans(
    pszInputFileFullPath: str,
    listStep0002Rows: list[list[str]],
    listMappings: list[tuple[str, str, str]],
) -> list[ProductFilePlan]:
    """全商品の識別情報とStep0003～Step0007出力パスを事前計算します。"""
    objInputPath: Path = Path(pszInputFileFullPath)
    listCenterNames: list[str] = list(dict.fromkeys(pszCenter for pszCenter, _, _ in listMappings))
    dictSafeCenterNames: dict[str, str] = {
        pszCenter: sanitize_delivery_center_filename(pszCenter)
        for pszCenter in listCenterNames
    }
    if len({pszName.casefold() for pszName in dictSafeCenterNames.values()}) != len(
        dictSafeCenterNames
    ):
        raise ValueError("配送センター名のファイル名が重複します。")

    listPlans: list[ProductFilePlan] = []
    dictPathOwners: dict[str, ProductFilePlan] = {}
    for iRow, listRow in enumerate(listStep0002Rows, start=2):
        pszApexCode: str = listRow[5].strip()
        pszProductName: str = listRow[6].strip()
        pszSafeApexCode = sanitize_product_filename_part(pszApexCode, "APEX品番", iRow)
        pszSafeProductName = sanitize_product_filename_part(pszProductName, "商品名", iRow)
        objPlan = ProductFilePlan(
            iRow,
            listRow.copy(),
            pszApexCode,
            pszProductName,
            pszSafeApexCode,
            pszSafeProductName,
        )
        for pszStep in ("step0003", "step0004", "step0005", "step0006"):
            pszStem: str = objInputPath.stem + "_" + pszStep + "_" + objPlan.file_identity
            objPlan.output_paths[pszStep + ".xlsx"] = objInputPath.with_name(pszStem + ".xlsx")
            objPlan.output_paths[pszStep + ".tsv"] = objInputPath.with_name(pszStem + ".tsv")
        objPlan.output_paths["step0005_warning.txt"] = objInputPath.with_name(
            objInputPath.stem + "_step0005_" + objPlan.file_identity + "_warning.txt"
        )
        for pszCenterName in listCenterNames:
            pszCenterStem: str = objPlan.file_identity + "_" + dictSafeCenterNames[pszCenterName]
            dictCenterPaths: dict[str, Path] = {}
            for pszStep in ("step0006", "step0007"):
                for pszSuffix in (".xlsx", ".tsv"):
                    dictCenterPaths[pszStep + pszSuffix] = objInputPath.with_name(
                        objInputPath.stem + "_" + pszStep + "_" + pszCenterStem + pszSuffix
                    )
            dictCenterPaths["step0007_warning.txt"] = objInputPath.with_name(
                objInputPath.stem
                + "_step0007_"
                + pszCenterStem
                + "_warning.txt"
            )
            objPlan.center_output_paths[pszCenterName] = dictCenterPaths
        for objOutputPath in [
            *objPlan.output_paths.values(),
            *(
                objPath
                for dictPaths in objPlan.center_output_paths.values()
                for objPath in dictPaths.values()
            ),
        ]:
            pszKey: str = os.path.normcase(os.path.abspath(objOutputPath)).casefold()
            if pszKey in dictPathOwners:
                raise ValueError(
                    "処理0003の商品別出力ファイル名が重複しています。\n"
                    + "APEX品番 = "
                    + pszApexCode
                    + "、商品名 = "
                    + pszProductName
                )
            dictPathOwners[pszKey] = objPlan
        listPlans.append(objPlan)
    return listPlans


def is_archived_step0003_path(objPath: Path) -> bool:
    """商品別Step0003ファイルが既にタイムスタンプ付きか返します。"""
    return re.search(r"_\d{14}(?:_\d+)?$", objPath.stem) is not None


def get_unique_path(objDesiredPath: Path) -> Path:
    """既存ファイルを上書きしない一意なパスを返します。"""
    if not objDesiredPath.exists():
        return objDesiredPath
    iSequence: int = 2
    while True:
        objCandidate = objDesiredPath.with_name(
            objDesiredPath.stem + "_" + str(iSequence) + objDesiredPath.suffix
        )
        if not objCandidate.exists():
            return objCandidate
        iSequence += 1


def archive_stale_step0003_files(
    pszInputFileFullPath: str,
    listPlans: list[ProductFilePlan],
    objProgress: ProductProcessingProgress,
) -> list[tuple[Path, Path]]:
    """今回対象外の商品別Step0003を%TEMP%へコピーし、最終更新日時付きへ変更します。"""
    objInputPath = Path(pszInputFileFullPath)
    setCurrentPaths: set[str] = {
        os.path.normcase(
            os.path.abspath(objPlan.output_paths["step0003" + pszSuffix])
        ).casefold()
        for objPlan in listPlans
        for pszSuffix in (".xlsx", ".tsv")
    }
    listStalePaths: list[Path] = []
    for pszSuffix in (".xlsx", ".tsv"):
        for objPath in objInputPath.parent.glob(objInputPath.stem + "_step0003_*" + pszSuffix):
            if is_archived_step0003_path(objPath):
                continue
            if os.path.normcase(os.path.abspath(objPath)).casefold() in setCurrentPaths:
                continue
            listStalePaths.append(objPath)
    objProgress.stale_files = len(listStalePaths)
    if not listStalePaths:
        return []
    iRequiredBytes: int = sum(objPath.stat().st_size for objPath in listStalePaths)
    iFreeBytes: int = shutil.disk_usage(tempfile.gettempdir()).free
    if iRequiredBytes > iFreeBytes:
        raise ValueError(
            "%TEMP%の空き容量が不足しています。必要容量 = "
            + str(iRequiredBytes)
            + "、空き容量 = "
            + str(iFreeBytes)
        )
    objTempRoot = Path(tempfile.gettempdir()) / "AsahiSingleOrderTemplateMaker"
    objTempRoot.mkdir(parents=True, exist_ok=True)
    pszRunTimestamp: str = datetime.now().strftime("%Y%m%d%H%M%S")
    objBackupDirectory = get_unique_path(objTempRoot / pszRunTimestamp)
    objBackupDirectory.mkdir()
    objProgress.temp_backup_directory = str(objBackupDirectory)
    for objPath in listStalePaths:
        objCopyPath = objBackupDirectory / objPath.name
        shutil.copy2(objPath, objCopyPath)
        if not objCopyPath.is_file() or objCopyPath.stat().st_size != objPath.stat().st_size:
            raise ValueError("%TEMP%へのバックアップ検証に失敗しました。Path = " + str(objPath))
        objProgress.temp_copies += 1
    listRenames: list[tuple[Path, Path]] = []
    try:
        for objPath in listStalePaths:
            pszTimestamp: str = datetime.fromtimestamp(objPath.stat().st_mtime).strftime(
                "%Y%m%d%H%M%S"
            )
            objArchivePath = get_unique_path(
                objPath.with_name(objPath.stem + "_" + pszTimestamp + objPath.suffix)
            )
            objPath.rename(objArchivePath)
            listRenames.append((objPath, objArchivePath))
            objProgress.renamed_files += 1
    except Exception:
        for objOriginalPath, objArchivePath in reversed(listRenames):
            try:
                objArchivePath.rename(objOriginalPath)
                objProgress.restore_successes += 1
            except OSError:
                objProgress.restore_failures += 1
                objProgress.failed_restore_paths.append(str(objArchivePath))
        raise
    return listRenames


def archive_existing_product_files(
    listCandidatePaths: list[Path], objProgress: ProductProcessingProgress
) -> list[Path]:
    """既存の商品別ファイルを%TEMP%へコピーして最終更新日時付きへ変更します。"""
    listOldPaths: list[Path] = [
        objPath for objPath in listCandidatePaths if objPath.is_file()
    ]
    if not listOldPaths:
        return []
    objProgress.stale_files += len(listOldPaths)
    iRequiredBytes: int = sum(objPath.stat().st_size for objPath in listOldPaths)
    iFreeBytes: int = shutil.disk_usage(tempfile.gettempdir()).free
    if iRequiredBytes > iFreeBytes:
        raise ValueError(
            "%TEMP%の空き容量が不足しています。必要容量 = "
            + str(iRequiredBytes)
            + "、空き容量 = "
            + str(iFreeBytes)
        )
    if objProgress.temp_backup_directory == "なし":
        objTempRoot = Path(tempfile.gettempdir()) / "AsahiSingleOrderTemplateMaker"
        objTempRoot.mkdir(parents=True, exist_ok=True)
        objBackupDirectory = get_unique_path(
            objTempRoot / datetime.now().strftime("%Y%m%d%H%M%S")
        )
        objBackupDirectory.mkdir()
        objProgress.temp_backup_directory = str(objBackupDirectory)
    else:
        objBackupDirectory = Path(objProgress.temp_backup_directory)
        objBackupDirectory.mkdir(parents=True, exist_ok=True)

    listCopiedPaths: list[Path] = []
    try:
        for objPath in listOldPaths:
            objCopyPath = objBackupDirectory / objPath.name
            shutil.copy2(objPath, objCopyPath)
            listCopiedPaths.append(objCopyPath)
            if not objCopyPath.is_file() or objCopyPath.stat().st_size != objPath.stat().st_size:
                raise ValueError(
                    "%TEMP%へのバックアップ検証に失敗しました。Path = " + str(objPath)
                )
            objProgress.temp_copies += 1
    except Exception:
        for objCopiedPath in listCopiedPaths:
            if objCopiedPath.exists():
                objCopiedPath.unlink()
        raise

    listRenames: list[tuple[Path, Path]] = []
    try:
        for objPath in listOldPaths:
            pszTimestamp: str = datetime.fromtimestamp(objPath.stat().st_mtime).strftime(
                "%Y%m%d%H%M%S"
            )
            objArchivePath = get_unique_path(
                objPath.with_name(objPath.stem + "_" + pszTimestamp + objPath.suffix)
            )
            objPath.rename(objArchivePath)
            listRenames.append((objPath, objArchivePath))
            objProgress.renamed_files += 1
    except Exception:
        for objOriginalPath, objArchivePath in reversed(listRenames):
            try:
                objArchivePath.rename(objOriginalPath)
                objProgress.restore_successes += 1
            except OSError:
                objProgress.restore_failures += 1
                objProgress.failed_restore_paths.append(str(objArchivePath))
        raise
    return [objArchivePath for _, objArchivePath in listRenames]


def archive_skipped_step0007_files(
    dictCenterPaths: dict[str, Path], objProgress: ProductProcessingProgress
) -> list[Path]:
    """対象データがない配送センターの古いStep0007を退避して名前を変更します。"""
    return archive_existing_product_files(
        [
            dictCenterPaths["step0007.xlsx"],
            dictCenterPaths["step0007.tsv"],
        ],
        objProgress,
    )


def format_product_processing_error(
    pszDetail: str, objProgress: ProductProcessingProgress
) -> str:
    """商品別処理の進行状況と復旧結果をエラー詳細へ追加します。"""
    return (
        pszDetail
        + "\n現在の商品: " + str(objProgress.current_product) + " / " + str(objProgress.total_products)
        + "\nAPEX品番: " + objProgress.apex_code
        + "\n商品名: " + objProgress.product_name
        + "\n配送センター: " + objProgress.center_name
        + "\n正常に完了した最終処理: " + objProgress.completed_process_name
        + "\n前処理までの正式出力: 保持"
        + "\n最大予定出力数: " + str(objProgress.planned_outputs)
        + "\n実際の出力対象数: " + str(objProgress.actual_output_targets)
        + "\n一時作成済み出力数: " + str(objProgress.temporary_outputs)
        + "\n検証済み出力数: " + str(objProgress.validated_outputs)
        + "\n確定済み出力数: " + str(objProgress.committed_outputs)
        + "\n正常スキップ配送センター数: " + str(objProgress.skipped_centers)
        + "\n警告ファイル数: " + str(objProgress.warning_files)
        + "\n古いファイル検出数: " + str(objProgress.stale_files)
        + "\n%TEMP%コピー済み数: " + str(objProgress.temp_copies)
        + "\nリネーム済み数: " + str(objProgress.renamed_files)
        + "\n復旧成功数: " + str(objProgress.restore_successes)
        + "\n復旧失敗数: " + str(objProgress.restore_failures)
        + "\n%TEMP%バックアップ: " + objProgress.temp_backup_directory
        + "\n復旧失敗パス: "
        + (
            ", ".join(objProgress.failed_restore_paths)
            if objProgress.failed_restore_paths
            else "なし"
        )
    )


def create_product_temporary_output(
    objFinalPath: Path,
    dictTemporaryOutputs: dict[Path, Path],
    objProgress: ProductProcessingProgress,
) -> Path:
    """商品別出力の一時パスを作成し、進行状況へ登録します。"""
    objTemporaryPath = create_temporary_path(objFinalPath, objFinalPath.suffix)
    dictTemporaryOutputs[objFinalPath] = objTemporaryPath
    if objFinalPath.suffix.lower() != ".txt":
        objProgress.temporary_outputs += 1
    return objTemporaryPath


def validate_saved_step0003_pair(objExcelPath: Path, objTsvPath: Path, pszStep: str) -> None:
    """Step0003またはStep0004の一時XLSX／TSVが一致することを確認します。"""
    listExcelRows = read_step0002_excel_rows(objExcelPath, pszStep)
    listTsvRows = read_step0002_tsv_rows(objTsvPath, pszStep)
    validate_step0002_outputs_match(listExcelRows, listTsvRows, pszStep)


def build_step0005_table_rows(listStep0005Rows: list[list[object]]) -> list[list[str]]:
    """処理0005のデータへ2行ヘッダーを加えた文字列表を返します。"""
    listStoreCodes: list[str] = [pszCode for pszCode, _ in STORE_DEFINITIONS]
    listStoreNames: list[str] = [pszName for _, pszName in STORE_DEFINITIONS]
    listTableRows: list[list[str]] = [
        [""] * len(STEP0002_HEADERS) + listStoreCodes,
        list(STEP0002_HEADERS) + listStoreNames,
    ]
    for listRow in listStep0005Rows:
        listTableRows.append([normalize_text(objValue) for objValue in listRow])
    return listTableRows


def commit_product_output_set(
    dictTemporaryOutputs: dict[Path, Path], objProgress: ProductProcessingProgress
) -> None:
    """1つの処理段階の出力を一括確定し、失敗時は同段階の出力を復旧します。"""
    dictBackups: dict[Path, Path] = {}
    listReplaced: list[Path] = []
    try:
        for objOutputPath in dictTemporaryOutputs:
            if objOutputPath.exists():
                objBackupPath = create_temporary_path(objOutputPath, ".backup")
                shutil.copy2(objOutputPath, objBackupPath)
                dictBackups[objOutputPath] = objBackupPath
        for objOutputPath, objTemporaryPath in dictTemporaryOutputs.items():
            os.replace(objTemporaryPath, objOutputPath)
            listReplaced.append(objOutputPath)
            if objOutputPath.suffix.lower() == ".txt":
                objProgress.warning_files += 1
            else:
                objProgress.committed_outputs += 1
    except Exception:
        for objOutputPath in reversed(listReplaced):
            try:
                objBackupPath = dictBackups.get(objOutputPath)
                if objBackupPath is not None and objBackupPath.exists():
                    os.replace(objBackupPath, objOutputPath)
                elif objOutputPath.exists():
                    objOutputPath.unlink()
                objProgress.restore_successes += 1
            except OSError:
                objProgress.restore_failures += 1
                objProgress.failed_restore_paths.append(str(objOutputPath))
            if objOutputPath.suffix.lower() == ".txt":
                objProgress.warning_files -= 1
            else:
                objProgress.committed_outputs -= 1
        raise
    finally:
        for objPath in [*dictBackups.values(), *dictTemporaryOutputs.values()]:
            if objPath.exists():
                objPath.unlink()


def process_product_step0003_files(
    objPlan: ProductFilePlan,
    objStep0002ExcelPath: Path,
    objStep0002TsvPath: Path,
    objProgress: ProductProcessingProgress,
) -> tuple[Path, Path]:
    """処理0002の正式な両出力から商品別処理0003を作成して確定します。"""
    listExcelRows = read_step0002_excel_rows(objStep0002ExcelPath)
    listTsvRows = read_step0002_tsv_rows(objStep0002TsvPath)
    validate_step0002_outputs_match(listExcelRows, listTsvRows)
    iProductIndex: int = objPlan.source_row - 2
    if iProductIndex < 0 or iProductIndex >= len(listExcelRows):
        raise ValueError("処理0002の商品行が見つかりません。")
    if listExcelRows[iProductIndex] != objPlan.step0002_row:
        raise ValueError("処理0002の商品情報が出力計画と一致しません。")

    objExcelPath = objPlan.output_paths["step0003.xlsx"]
    objTsvPath = objPlan.output_paths["step0003.tsv"]
    dictTemporaryOutputs: dict[Path, Path] = {}
    try:
        objExcelTemp = create_product_temporary_output(
            objExcelPath, dictTemporaryOutputs, objProgress
        )
        objTsvTemp = create_product_temporary_output(
            objTsvPath, dictTemporaryOutputs, objProgress
        )
        listRows = build_step0003_rows([listExcelRows[iProductIndex]])
        save_step0003_excel_template(objExcelTemp, listRows)
        save_step0003_tsv_template(objTsvTemp, listRows)
        validate_saved_step0003_pair(objExcelTemp, objTsvTemp, "step0003")
        objProgress.validated_outputs += 2
        commit_product_output_set(dictTemporaryOutputs, objProgress)
    finally:
        for objTemporaryPath in dictTemporaryOutputs.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    return objExcelPath, objTsvPath


def process_product_step0004_files(
    objPlan: ProductFilePlan,
    objStep0003ExcelPath: Path,
    objStep0003TsvPath: Path,
    objStartMonday: date,
    objProgress: ProductProcessingProgress,
) -> tuple[Path, Path]:
    """商品別処理0003の正式な両出力から処理0004を作成して確定します。"""
    listExcelRows = read_step0002_excel_rows(objStep0003ExcelPath, "step0003")
    listTsvRows = read_step0002_tsv_rows(objStep0003TsvPath, "step0003")
    validate_step0002_outputs_match(listExcelRows, listTsvRows, "step0003")
    listRows = build_step0004_rows(listExcelRows, objStartMonday)
    objExcelPath = objPlan.output_paths["step0004.xlsx"]
    objTsvPath = objPlan.output_paths["step0004.tsv"]
    dictTemporaryOutputs: dict[Path, Path] = {}
    try:
        objExcelTemp = create_product_temporary_output(
            objExcelPath, dictTemporaryOutputs, objProgress
        )
        objTsvTemp = create_product_temporary_output(
            objTsvPath, dictTemporaryOutputs, objProgress
        )
        save_step0004_excel_template(objExcelTemp, listRows)
        save_step0004_tsv_template(objTsvTemp, listRows)
        validate_saved_step0003_pair(objExcelTemp, objTsvTemp, "step0004")
        objProgress.validated_outputs += 2
        commit_product_output_set(dictTemporaryOutputs, objProgress)
    finally:
        for objTemporaryPath in dictTemporaryOutputs.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    return objExcelPath, objTsvPath


def process_product_step0005_files(
    pszInputFileFullPath: str,
    objPlan: ProductFilePlan,
    objStep0004ExcelPath: Path,
    objStep0004TsvPath: Path,
    objStartMonday: date,
    objProgress: ProductProcessingProgress,
) -> tuple[Path, Path]:
    """商品別処理0004の正式な両出力から処理0005を作成して確定します。"""
    listExcelRows = read_step0002_excel_rows(objStep0004ExcelPath, "step0004")
    listTsvRows = read_step0002_tsv_rows(objStep0004TsvPath, "step0004")
    validate_step0002_outputs_match(listExcelRows, listTsvRows, "step0004")
    pszProductSeedPath: str = str(
        Path(pszInputFileFullPath).with_name(
            Path(pszInputFileFullPath).stem + "_" + objPlan.file_identity
            + Path(pszInputFileFullPath).suffix
        )
    )
    listRetainedRows, _, iRemovedCount = select_step0005_product_rows(
        listExcelRows, pszProductSeedPath, objStartMonday
    )
    listRows = build_step0005_rows(
        listRetainedRows, pszProductSeedPath, objStartMonday
    )
    objExcelPath = objPlan.output_paths["step0005.xlsx"]
    objTsvPath = objPlan.output_paths["step0005.tsv"]
    objWarningPath = objPlan.output_paths["step0005_warning.txt"]
    dictTemporaryOutputs: dict[Path, Path] = {}
    try:
        objExcelTemp = create_product_temporary_output(
            objExcelPath, dictTemporaryOutputs, objProgress
        )
        objTsvTemp = create_product_temporary_output(
            objTsvPath, dictTemporaryOutputs, objProgress
        )
        save_step0005_excel_template(objExcelTemp, listRows)
        save_step0005_tsv_template(objTsvTemp, listRows)
        validate_step0005_tables_match(
            read_step0005_excel_table(objExcelTemp),
            read_step0005_tsv_table(objTsvTemp),
        )
        objProgress.validated_outputs += 2
        if iRemovedCount == 1:
            objWarningTemp = create_product_temporary_output(
                objWarningPath, dictTemporaryOutputs, objProgress
            )
            write_warning_text(
                str(objWarningTemp),
                "処理結果: 警告\n入力ファイル: "
                + os.path.abspath(pszInputFileFullPath)
                + "\n発生した処理: 旭注文テンプレート処理0005\nAPEX品番: "
                + objPlan.apex_code
                + "\n商品名: "
                + objPlan.product_name
                + "\n警告内容: 商品削除抽選の結果、この商品が削除されました。\n"
                + "商品削除確率: " + str(PRODUCT_DELETE_PROBABILITY)
                + " (" + str(PRODUCT_DELETE_PROBABILITY * 100) + "%)"
                + "\n入力商品数: 1\n残存商品数: 0\n削除商品数: 1",
            )
        commit_product_output_set(dictTemporaryOutputs, objProgress)
        if iRemovedCount != 1 and objWarningPath.exists():
            objWarningPath.unlink()
    finally:
        for objTemporaryPath in dictTemporaryOutputs.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    return objExcelPath, objTsvPath


def process_product_step0006_files(
    objPlan: ProductFilePlan,
    objStep0005ExcelPath: Path,
    objStep0005TsvPath: Path,
    listMappings: list[tuple[str, str, str]],
    objProgress: ProductProcessingProgress,
) -> tuple[Path, Path, list[str], list[str]]:
    """商品別処理0005の正式な両出力から処理0006を作成して確定します。"""
    listExcelRows = read_step0005_excel_table(objStep0005ExcelPath)
    listTsvRows = read_step0005_tsv_table(objStep0005TsvPath)
    validate_step0005_tables_match(listExcelRows, listTsvRows)
    dictCenterCodes: dict[str, set[str]] = {}
    for pszCenter, pszCode, _ in listMappings:
        dictCenterCodes.setdefault(pszCenter, set()).add(pszCode)
    objExcelPath = objPlan.output_paths["step0006.xlsx"]
    objTsvPath = objPlan.output_paths["step0006.tsv"]
    listWarnings: list[str] = []
    listOutputCenters: list[str] = []
    dictTemporaryOutputs: dict[Path, Path] = {}
    try:
        listAllStoreRows = build_step0006_all_stores_rows(listExcelRows, listMappings)
        objExcelTemp = create_product_temporary_output(
            objExcelPath, dictTemporaryOutputs, objProgress
        )
        objTsvTemp = create_product_temporary_output(
            objTsvPath, dictTemporaryOutputs, objProgress
        )
        save_step0006_excel_template(objExcelTemp, listAllStoreRows)
        save_step0006_tsv_template(objTsvTemp, listAllStoreRows)
        validate_step0005_tables_match(
            read_step0005_excel_table(objExcelTemp),
            read_step0005_tsv_table(objTsvTemp),
        )
        objProgress.validated_outputs += 2
        for pszCenterName, setCodes in dictCenterCodes.items():
            objProgress.center_name = pszCenterName
            listCenterRows = build_step0006_center_rows(listExcelRows, setCodes)
            if len(listCenterRows[0]) == len(STEP0002_HEADERS):
                listWarnings.append(
                    f'警告: 配送センター「{pszCenterName}」の店舗コード列は処理0005にありません。'
                )
                objProgress.actual_output_targets -= 4
                continue
            if not center_has_order_quantity(listCenterRows):
                listWarnings.append(
                    f'警告: 配送センター「{pszCenterName}」の商品「{objPlan.product_name}」の'
                    + "店舗数量はすべて空欄ですが、XLSX・TSVを作成しました。"
                )
            dictCenterPaths = objPlan.center_output_paths[pszCenterName]
            objCenterExcelTemp = create_product_temporary_output(
                dictCenterPaths["step0006.xlsx"], dictTemporaryOutputs, objProgress
            )
            objCenterTsvTemp = create_product_temporary_output(
                dictCenterPaths["step0006.tsv"], dictTemporaryOutputs, objProgress
            )
            save_step0006_excel_template(objCenterExcelTemp, listCenterRows)
            save_step0006_tsv_template(objCenterTsvTemp, listCenterRows)
            validate_step0005_tables_match(
                read_step0005_excel_table(objCenterExcelTemp),
                read_step0005_tsv_table(objCenterTsvTemp),
            )
            objProgress.validated_outputs += 2
            listOutputCenters.append(pszCenterName)
        commit_product_output_set(dictTemporaryOutputs, objProgress)
    finally:
        for objTemporaryPath in dictTemporaryOutputs.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    objProgress.center_name = "なし"
    return objExcelPath, objTsvPath, listOutputCenters, listWarnings


def create_step0007_skip_warning(
    pszInputFileFullPath: str,
    objPlan: ProductFilePlan,
    pszCenterName: str,
    iStep0006Rows: int,
    listArchivedPaths: list[Path],
    dictTemporaryOutputs: dict[Path, Path],
    objProgress: ProductProcessingProgress,
) -> Path:
    """処理0007を正常スキップした商品×配送センターの警告を一時作成します。"""
    objWarningPath = objPlan.center_output_paths[pszCenterName][
        "step0007_warning.txt"
    ]
    objWarningTemp = create_product_temporary_output(
        objWarningPath, dictTemporaryOutputs, objProgress
    )
    if iStep0006Rows == 2:
        pszReason = "処理0006がヘッダー2行だけで対象商品データがないため"
    else:
        pszReason = (
            "処理0006の月曜日～日曜日データは存在しますが、"
            + "発注がある店舗がないため"
        )
    write_warning_text(
        str(objWarningTemp),
        "処理結果: 警告\n入力ファイル: "
        + os.path.abspath(pszInputFileFullPath)
        + "\n発生した処理: 旭注文テンプレート処理0007\nAPEX品番: "
        + objPlan.apex_code
        + "\n商品名: "
        + objPlan.product_name
        + "\n配送センター: "
        + pszCenterName
        + "\n警告内容: "
        + pszReason
        + "、処理0007 XLSX・TSVを作成しませんでした。"
        + "\n処理0006行数: "
        + str(iStep0006Rows)
        + "\n発注あり店舗数: 0"
        + "\n処理0007 XLSX: 作成なし"
        + "\n処理0007 TSV: 作成なし"
        + "\n処理結果: 次の配送センターへ継続"
        + "\n古い処理0007バックアップ: "
        + ("あり" if listArchivedPaths else "なし")
        + "\n%TEMP%バックアップ: "
        + objProgress.temp_backup_directory,
    )
    return objWarningPath


def process_product_step0007_files(
    pszInputFileFullPath: str,
    objPlan: ProductFilePlan,
    listCenterNames: list[str],
    objProgress: ProductProcessingProgress,
) -> list[str]:
    """配送センター別処理0006の正式な両出力から処理0007を作成して確定します。"""
    listWarnings: list[str] = []
    dictTemporaryOutputs: dict[Path, Path] = {}
    setOldWarningPathsToArchive: set[Path] = set()
    try:
        for pszCenterName in listCenterNames:
            objProgress.center_name = pszCenterName
            dictCenterPaths = objPlan.center_output_paths[pszCenterName]
            listExcelRows = read_step0005_excel_table(
                dictCenterPaths["step0006.xlsx"]
            )
            listTsvRows = read_step0005_tsv_table(
                dictCenterPaths["step0006.tsv"]
            )
            validate_step0005_tables_match(listExcelRows, listTsvRows)
            if len(listExcelRows) == 2:
                listArchivedPaths = archive_skipped_step0007_files(
                    dictCenterPaths, objProgress
                )
                pszWarning = (
                    f'警告: 配送センター「{pszCenterName}」の商品「{objPlan.product_name}」は'
                    + "処理0006に対象データがないため、処理0007 XLSX・TSVを作成しませんでした。"
                )
                if listArchivedPaths:
                    pszWarning += (
                        " 古い処理0007ファイルは%TEMP%へバックアップし、"
                        + "最終更新日時付きファイル名へ変更しました。"
                    )
                if len(listArchivedPaths) == 1:
                    pszWarning += " 古い処理0007 XLSX・TSVは片方だけ存在していました。"
                create_step0007_skip_warning(
                    pszInputFileFullPath,
                    objPlan,
                    pszCenterName,
                    2,
                    listArchivedPaths,
                    dictTemporaryOutputs,
                    objProgress,
                )
                objProgress.actual_output_targets -= 2
                objProgress.skipped_centers += 1
                listWarnings.append(pszWarning)
                continue
            listRows, iKeptStoreCount, _ = build_product_step0007_rows(
                listExcelRows, pszCenterName
            )
            if iKeptStoreCount == 0:
                listArchivedPaths = archive_skipped_step0007_files(
                    dictCenterPaths, objProgress
                )
                create_step0007_skip_warning(
                    pszInputFileFullPath,
                    objPlan,
                    pszCenterName,
                    9,
                    listArchivedPaths,
                    dictTemporaryOutputs,
                    objProgress,
                )
                objProgress.actual_output_targets -= 2
                objProgress.skipped_centers += 1
                pszWarning = (
                    f'警告: 配送センター「{pszCenterName}」の商品「{objPlan.product_name}」には'
                    + "発注がある店舗がないため、処理0007 XLSX・TSVを作成しませんでした。"
                )
                if listArchivedPaths:
                    pszWarning += (
                        " 古い処理0007ファイルは%TEMP%へバックアップし、"
                        + "最終更新日時付きファイル名へ変更しました。"
                    )
                if len(listArchivedPaths) == 1:
                    pszWarning += " 古い処理0007 XLSX・TSVは片方だけ存在していました。"
                listWarnings.append(pszWarning)
                continue
            setOldWarningPathsToArchive.add(
                dictCenterPaths["step0007_warning.txt"]
            )
            objExcelTemp = create_product_temporary_output(
                dictCenterPaths["step0007.xlsx"], dictTemporaryOutputs, objProgress
            )
            objTsvTemp = create_product_temporary_output(
                dictCenterPaths["step0007.tsv"], dictTemporaryOutputs, objProgress
            )
            save_step0006_excel_template(objExcelTemp, listRows)
            save_step0006_tsv_template(objTsvTemp, listRows)
            listSavedExcelRows = read_step0005_excel_table(objExcelTemp)
            listSavedTsvRows = read_step0005_tsv_table(objTsvTemp)
            validate_step0005_tables_match(
                listSavedExcelRows, listSavedTsvRows
            )
            validate_step0005_tables_match(listRows, listSavedExcelRows)
            objProgress.validated_outputs += 2
        commit_product_output_set(dictTemporaryOutputs, objProgress)
        for objWarningPath in setOldWarningPathsToArchive:
            archive_existing_product_files([objWarningPath], objProgress)
    finally:
        for objTemporaryPath in dictTemporaryOutputs.values():
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()
    objProgress.center_name = "なし"
    return listWarnings


def process_product_file_pipeline(
    pszInputFileFullPath: str,
    objStep0002ExcelPath: Path,
    objStep0002TsvPath: Path,
    objStartMonday: date,
    objMappingPath: Path,
) -> tuple[list[ProductFilePlan], list[str], ProductProcessingProgress]:
    """各Stepの正式なXLSX／TSVを次Stepが読み込む商品別処理を制御します。"""
    objProgress = ProductProcessingProgress()
    listWarnings: list[str] = []
    try:
        listExcelRows = read_step0002_excel_rows(objStep0002ExcelPath)
        listTsvRows = read_step0002_tsv_rows(objStep0002TsvPath)
        validate_step0002_outputs_match(listExcelRows, listTsvRows)
        listMappings = read_area_store_mapping(objMappingPath)
        listPlans = build_product_file_plans(
            pszInputFileFullPath, listExcelRows, listMappings
        )
        objProgress.total_products = len(listPlans)
        objProgress.planned_outputs = sum(
            sum(
                1
                for objPath in objPlan.output_paths.values()
                if objPath.suffix.lower() != ".txt"
            )
            + sum(
                1
                for dictPaths in objPlan.center_output_paths.values()
                for objPath in dictPaths.values()
                if objPath.suffix.lower() != ".txt"
            )
            for objPlan in listPlans
        )
        objProgress.actual_output_targets = objProgress.planned_outputs
        archive_stale_step0003_files(
            pszInputFileFullPath, listPlans, objProgress
        )
        for iProduct, objPlan in enumerate(listPlans, start=1):
            objProgress.current_product = iProduct
            objProgress.apex_code = objPlan.apex_code
            objProgress.product_name = objPlan.product_name
            objProgress.center_name = "なし"
            objProgress.completed_process_name = "なし"
            objProgress.process_name = "処理0003"
            objStep0003ExcelPath, objStep0003TsvPath = process_product_step0003_files(
                objPlan, objStep0002ExcelPath, objStep0002TsvPath, objProgress
            )
            objProgress.completed_process_name = "処理0003"

            objProgress.process_name = "処理0004"
            objStep0004ExcelPath, objStep0004TsvPath = process_product_step0004_files(
                objPlan,
                objStep0003ExcelPath,
                objStep0003TsvPath,
                objStartMonday,
                objProgress,
            )
            objProgress.completed_process_name = "処理0004"

            objProgress.process_name = "処理0005"
            objStep0005ExcelPath, objStep0005TsvPath = process_product_step0005_files(
                pszInputFileFullPath,
                objPlan,
                objStep0004ExcelPath,
                objStep0004TsvPath,
                objStartMonday,
                objProgress,
            )
            objProgress.completed_process_name = "処理0005"

            objProgress.process_name = "処理0006"
            _, _, listCenterNames, listStep0006Warnings = process_product_step0006_files(
                objPlan,
                objStep0005ExcelPath,
                objStep0005TsvPath,
                listMappings,
                objProgress,
            )
            listWarnings.extend(listStep0006Warnings)
            objProgress.completed_process_name = "処理0006"

            objProgress.process_name = "処理0007"
            listWarnings.extend(
                process_product_step0007_files(
                    pszInputFileFullPath,
                    objPlan,
                    listCenterNames,
                    objProgress,
                )
            )
            objProgress.completed_process_name = "処理0007"
        if objProgress.temp_backup_directory != "なし":
            objBackupDirectory = Path(objProgress.temp_backup_directory)
            print("Temp Backup Directory: " + str(objBackupDirectory))
            print("Temp Backup Files: " + str(objProgress.temp_copies))
            print(
                "Temp Backup Bytes: "
                + str(sum(objPath.stat().st_size for objPath in objBackupDirectory.iterdir()))
            )
        return listPlans, listWarnings, objProgress
    except Exception as objException:
        pszDetail = format_product_processing_error(str(objException), objProgress)
        if objProgress.process_name == "処理0007":
            raise Step0007Error(pszDetail) from objException
        if objProgress.process_name == "処理0006":
            raise Step0006Error(pszDetail) from objException
        if objProgress.process_name == "処理0005":
            raise Step0005Error(pszDetail) from objException
        if objProgress.process_name == "処理0004":
            raise Step0004Error(pszDetail) from objException
        raise Step0003Error(pszDetail) from objException


def process_input_file(
    pszInputFileFullPath: str,
    objStartMonday: date,
    objMappingPath: Path | None = None,
) -> None:
    """入力から処理0001～処理0007のXLSX・TSVを作成します。"""
    validate_start_monday(objStartMonday)
    if objMappingPath is None:
        objMappingPath = get_default_mapping_file_path()
    pszValidatedPath: str = validate_input_path(pszInputFileFullPath)
    pszExtension: str = os.path.splitext(pszValidatedPath)[1].lower()
    if pszExtension == ".xlsx":
        listProductRows: list[ProductRow] = read_excel_rows(pszValidatedPath)
    else:
        listProductRows = read_delimited_rows(pszValidatedPath)
    validate_product_count(listProductRows)
    objExcelOutputPath, objTsvOutputPath = get_output_file_paths(pszValidatedPath)
    objTemporaryExcelPath: Path = create_temporary_path(objExcelOutputPath, ".xlsx")
    objTemporaryTsvPath: Path = create_temporary_path(objTsvOutputPath, ".tsv")
    try:
        save_excel_template(objTemporaryExcelPath, listProductRows)
        save_tsv_template(objTemporaryTsvPath, listProductRows)
        replace_output_files(
            objTemporaryExcelPath,
            objTemporaryTsvPath,
            objExcelOutputPath,
            objTsvOutputPath,
        )
    finally:
        for objTemporaryPath in (objTemporaryExcelPath, objTemporaryTsvPath):
            if objTemporaryPath.exists():
                objTemporaryPath.unlink()

    objStep0002ExcelPath, objStep0002TsvPath, _ = (
        process_step0002_files(
            pszValidatedPath,
            objExcelOutputPath,
            objTsvOutputPath,
        )
    )
    listProductPlans, listProductWarnings, objProductProgress = process_product_file_pipeline(
        pszValidatedPath,
        objStep0002ExcelPath,
        objStep0002TsvPath,
        objStartMonday,
        objMappingPath,
    )
    remove_old_error_file(pszValidatedPath)
    print("旭注文テンプレートファイルを作成しました。")
    print("Input: " + pszValidatedPath)
    print("Start Monday: " + objStartMonday.strftime("%Y/%m/%d"))
    print("Step0001 Excel: " + str(objExcelOutputPath))
    print("Step0001 TSV: " + str(objTsvOutputPath))
    print("Step0002 Excel: " + str(objStep0002ExcelPath))
    print("Step0002 TSV: " + str(objStep0002TsvPath))
    print("Step0005 Product Delete Probability: " + str(PRODUCT_DELETE_PROBABILITY))
    print(
        "Step0005 Product Delete Percent: "
        + str(PRODUCT_DELETE_PROBABILITY * 100)
        + "%"
    )
    print("Mapping: " + str(objMappingPath))
    for objPlan in listProductPlans:
        for pszKey, objOutputPath in objPlan.output_paths.items():
            if pszKey.endswith("warning.txt") and not objOutputPath.exists():
                continue
            print(pszKey + ": " + str(objOutputPath))
        for pszCenterName, dictPaths in objPlan.center_output_paths.items():
            for pszKey, objOutputPath in dictPaths.items():
                if objOutputPath.exists():
                    print(pszKey + " [" + pszCenterName + "]: " + str(objOutputPath))
    for pszWarning in listProductWarnings:
        print(pszWarning)
    print("Products: " + str(len(listProductPlans)))
    print("最大予定出力数: " + str(objProductProgress.planned_outputs))
    print("実際の出力対象数: " + str(objProductProgress.actual_output_targets))
    print("確定済み出力数: " + str(objProductProgress.committed_outputs))
    print("正常スキップ配送センター数: " + str(objProductProgress.skipped_centers))
    print("警告ファイル数: " + str(objProductProgress.warning_files))


def parse_command_line_arguments() -> tuple[str, str | None, Path]:
    """入力、開始月曜日、対応表のコマンドライン引数を解析します。"""
    pszStartMonday: str | None = None
    objMappingPath: Path = get_default_mapping_file_path()
    listInputPaths: list[str] = []
    iArgument: int = 1
    while iArgument < len(sys.argv):
        pszArgument: str = sys.argv[iArgument]
        if pszArgument in ("--start-monday", "--mapping-file"):
            if iArgument + 1 >= len(sys.argv):
                raise ValueError(pszArgument + "の値が指定されていません。")
            pszValue: str = sys.argv[iArgument + 1]
            if pszArgument == "--start-monday":
                if pszStartMonday is not None:
                    raise ValueError("--start-mondayが重複しています。")
                pszStartMonday = pszValue
            else:
                if objMappingPath != get_default_mapping_file_path():
                    raise ValueError("--mapping-fileが重複しています。")
                objMappingPath = Path(pszValue).expanduser().resolve()
            iArgument += 2
            continue
        if pszArgument.startswith("--"):
            raise ValueError("未対応のオプションです。Option = " + pszArgument)
        listInputPaths.append(pszArgument)
        iArgument += 1
    if len(listInputPaths) != 1:
        raise ValueError("入力ファイルパスは1つ指定してください。")
    return listInputPaths[0], pszStartMonday, objMappingPath


def main() -> int:
    """引数を確認して処理し、成功0・失敗1の終了コードを返します。"""
    try:
        pszInputFileFullPath, pszStartMondayArgument, objMappingPath = (
            parse_command_line_arguments()
        )
    except ValueError as objException:
        pszScriptFileName: str = os.path.basename(__file__)
        pszErrorMessage: str = (
            "Error: " + str(objException) + "\n"
            + "Usage: python "
            + pszScriptFileName
            + " [--mapping-file <mapping_path>] <input_file_path>\n"
            + "With date: python "
            + pszScriptFileName
            + " --start-monday YYYY-MM-DD [--mapping-file <mapping_path>] <input_file_path>\n"
        )
        print(pszErrorMessage, file=sys.stderr, end="")
        pszErrorFileFullPath: str = os.path.splitext(pszScriptFileName)[0] + "_error_argument.txt"
        try:
            write_error_text(pszErrorFileFullPath, pszErrorMessage)
        except OSError as objException:
            print(
                "Error: 引数エラーファイルを保存できません。Detail = " + str(objException),
                file=sys.stderr,
            )
        return 1
    try:
        try:
            if pszStartMondayArgument is not None:
                objStartMonday: date = parse_start_monday(pszStartMondayArgument)
            else:
                objSelectedMonday: date | None = select_start_monday()
                if objSelectedMonday is None:
                    show_start_monday_cancelled_message()
                    raise Step0004Error(
                        "開始月曜日の選択がキャンセルされたため、処理0004を中止しました。"
                    )
                objStartMonday = objSelectedMonday
        except Step0004Error:
            raise
        except Exception as objException:
            raise Step0004Error(str(objException)) from objException
        process_input_file(pszInputFileFullPath, objStartMonday, objMappingPath)
    except Step0007Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "旭注文テンプレート処理0007",
            str(objException),
        )
        return 1
    except Step0006Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "旭注文テンプレート処理0006",
            str(objException),
        )
        return 1
    except Step0005Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "旭注文テンプレート処理0005",
            str(objException),
        )
        return 1
    except Step0004Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "旭注文テンプレート処理0004",
            str(objException),
        )
        return 1
    except Step0003Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "旭注文テンプレート処理0003",
            str(objException),
        )
        return 1
    except Step0002Error as objException:
        report_processing_error(
            pszInputFileFullPath,
            "旭注文テンプレート処理0002",
            str(objException),
        )
        return 1
    except Exception as objException:
        report_processing_error(
            pszInputFileFullPath,
            "旭注文テンプレート処理0001",
            str(objException),
        )
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
